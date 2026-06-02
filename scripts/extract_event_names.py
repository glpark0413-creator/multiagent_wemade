#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Readdocs xlsx에서 탭별 이벤트 섹션 제목을 추출해 output/historical_event_names.json 저장.

B열의 "숫자. " 패턴(예: "1. 얼리썸머 14일 출석 이벤트!")을 이벤트 섹션 제목으로 인식.
Claude가 이 파일을 읽어 시즌 패턴을 학습하고, 신규 탭의 이벤트 명칭을 제안할 때 활용한다.
"""
import io
import json
import re
import sys
from pathlib import Path
from datetime import datetime

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# ─── 현재 프로젝트 설정 로드 (crawl_gdrive_project.py 가 생성) ─────────────
_BASE_DIR = Path(__file__).resolve().parent.parent
_CURRENT_PROJECT_FILE = _BASE_DIR / "output" / "json" / "current_project.json"
_NC_KR_FALLBACK = str(_BASE_DIR / "Readdocs" / "[NC_KR] 라이브 이벤트 문서.xlsx")

def _resolve_source() -> str:
    """current_project.json 에서 source_xlsx 읽기. 없으면 NC_KR 레거시 경로 반환."""
    if _CURRENT_PROJECT_FILE.exists():
        try:
            cfg = json.loads(_CURRENT_PROJECT_FILE.read_text(encoding="utf-8"))
            p = cfg.get("source_xlsx", "")
            if p and Path(p).exists():
                return p
        except Exception:
            pass
    return _NC_KR_FALLBACK

SOURCE = _resolve_source()

import sys as _sys
_sys.path.insert(0, str(Path(__file__).resolve().parent))
from _project_config import load_project_paths as _load_paths

_paths = _load_paths()

OUTPUT_DIR = Path("output")
OUTPUT_JSON_DIR = OUTPUT_DIR / "json"
# 프로젝트별 경로 우선, 없으면 레거시 경로 사용
OUTPUT_FILE = _paths.hist_event_names if _paths else OUTPUT_JSON_DIR / "historical_event_names.json"

# B열의 번호 섹션 패턴: "1.", "2.", "3." 등으로 시작
SECTION_PATTERN = re.compile(r"^\d+\.")

# 시즌·월 키워드 감지 패턴 (공통 + 장르별)
SEASON_RE = re.compile(
    r"(봄의|여름의|가을의|겨울의|한여름|초여름|얼리썸머|"
    r"새해|신년|설날|추석|크리스마스|핼러윈|기념|축제|"
    r"각성|전설|영웅|봉인|결전|"
    r"전반기|올스타|후반기|포스트시즌|개막|시즌|"
    r"\d+주년|\d+월의|\d+월 )"
)


def tab_name_to_date(tab_name: str) -> datetime | None:
    """탭명 YYMMDD → datetime. 숫자 6자리가 아니면 None."""
    if not re.fullmatch(r"\d{6}", tab_name):
        return None
    try:
        return datetime.strptime("20" + tab_name, "%Y%m%d")
    except ValueError:
        return None


def extract_section_titles(ws) -> list[dict]:
    """워크시트 B열에서 섹션 제목 추출."""
    sections = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.column != 2:  # B열
                continue
            val = cell.value
            if not val or not isinstance(val, str):
                continue
            if SECTION_PATTERN.match(val.strip()):
                sections.append({
                    "cell": cell.coordinate,
                    "title": val.strip(),
                    "season_keywords": SEASON_RE.findall(val),
                })
    return sections


def _build_summarized(tabs: list, season_by_month: dict) -> dict:
    """
    --summarize 모드: LLM에 전달할 컴팩트 요약만 생성한다.
    원문 행 데이터(tabs 전체)를 제외해 토큰을 대폭 절약한다.
    """
    from collections import Counter
    # 제목 구조 패턴 분석
    all_titles = [sec["title"] for tab in tabs for sec in tab["event_sections"]]
    # 시즌 키워드 빈도
    all_kws = [kw for tab in tabs for sec in tab["event_sections"] for kw in sec["season_keywords"]]
    kw_freq = dict(Counter(all_kws).most_common(20))

    # 이벤트 제목에서 공통 구조 단어 추출 (2자 이상 한국어)
    import re as _re
    word_counter: Counter = Counter()
    for title in all_titles:
        words = _re.findall(r"[가-힣]{2,}", title)
        word_counter.update(words)
    # 빈도 상위 30개 (섹션 번호 제외)
    top_words = {w: c for w, c in word_counter.most_common(40)
                 if not _re.match(r"^\d+$", w)}

    # 최근 탭 제목 예시 (원문 포함 — 최근 4탭만)
    recent_examples = []
    for tab in tabs[-4:]:
        recent_examples.append({
            "tab":    tab["tab"],
            "date":   tab["date"],
            "titles": [sec["title"] for sec in tab["event_sections"]],
        })

    return {
        "summary_mode":            True,
        "total_tabs":              len(tabs),
        "total_event_sections":    len(all_titles),
        "season_keywords_by_month": season_by_month,
        "keyword_frequency":       kw_freq,
        "top_title_words":         top_words,
        "recent_tab_titles":       recent_examples,
        "note": (
            "이 데이터는 --summarize 모드로 생성됨. "
            "원문 탭 데이터는 포함되지 않음 (토큰 절약)."
        ),
    }


def main():
    import argparse as _ap
    parser = _ap.ArgumentParser(description="이벤트 섹션 제목 추출")
    parser.add_argument("source", nargs="?", default=None, help="소스 xlsx 경로 (기본: current_project.json)")
    parser.add_argument("--output", default=None, help="출력 JSON 경로")
    parser.add_argument(
        "--summarize", action="store_true",
        help="요약 통계만 저장 (원문 행 제외, LLM 전달용 토큰 절약 모드)",
    )
    args = parser.parse_args()

    source = args.source or SOURCE
    out_file = Path(args.output) if args.output else OUTPUT_FILE

    if _paths:
        _paths.ensure_dirs()
    OUTPUT_DIR.mkdir(exist_ok=True)
    OUTPUT_JSON_DIR.mkdir(exist_ok=True)

    wb = openpyxl.load_workbook(source)
    tabs = []

    for name in wb.sheetnames:
        dt = tab_name_to_date(name)
        if dt is None:
            continue
        ws = wb[name]
        sections = extract_section_titles(ws)
        if not sections:
            continue
        tabs.append({
            "tab": name,
            "date": dt.strftime("%Y-%m-%d"),
            "month": dt.strftime("%Y-%m"),
            "event_sections": sections,
        })

    # 날짜 오름차순 정렬
    tabs.sort(key=lambda x: x["date"])

    # 월별 시즌 키워드 집계
    season_by_month: dict[str, list[str]] = {}
    for tab in tabs:
        month = tab["month"]
        kws: list[str] = []
        for sec in tab["event_sections"]:
            kws.extend(sec["season_keywords"])
        if kws:
            existing = set(season_by_month.get(month, []))
            existing.update(kws)
            season_by_month[month] = sorted(existing)

    if args.summarize:
        # ── 요약 모드: 원문 tabs 배열 제외 ──────────────────────────────────
        result = {
            "source":       source,
            "extracted_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
            **_build_summarized(tabs, season_by_month),
        }
    else:
        result = {
            "source":                   source,
            "extracted_at":             datetime.now().strftime("%Y-%m-%d %H:%M"),
            "total_tabs":               len(tabs),
            "tabs":                     tabs,
            "recent_tabs":              tabs[-4:],
            "season_keywords_by_month": season_by_month,
        }

    out_file.parent.mkdir(parents=True, exist_ok=True)
    with open(out_file, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    mode_label = "[요약 모드]" if args.summarize else "[전체 모드]"
    print(f"{mode_label} 저장: {out_file}")
    print(f"탭 수: {len(tabs)}")

    if args.summarize:
        kw_top = list(result.get("keyword_frequency", {}).items())[:5]
        print(f"시즌 키워드 Top 5: {kw_top}")
    else:
        print("\n── 최근 탭 이벤트 섹션 ──")
        for tab in tabs[-4:]:
            print(f"  [{tab['tab']} / {tab['date']}]")
            for sec in tab["event_sections"]:
                kw_str = f"  ← {sec['season_keywords']}" if sec["season_keywords"] else ""
                print(f"    {sec['cell']}: {sec['title']}{kw_str}")

        print("\n── 월별 시즌 키워드 ──")
        for month, kws in list(season_by_month.items())[-6:]:
            print(f"  {month}: {kws}")


if __name__ == "__main__":
    main()
