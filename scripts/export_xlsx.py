"""
export_xlsx.py
세션 버퍼를 xlsx 또는 csv로 Export.

사용법:
    python export_xlsx.py --format xlsx    # xlsx만
    python export_xlsx.py --format csv     # csv만
    python export_xlsx.py --format both    # 둘 다 (기본값)
"""

import json
import sys
import csv
from datetime import datetime, timezone, timedelta
from pathlib import Path

# Windows 콘솔 UTF-8 출력 설정
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR  = ROOT / "output" / "game-localizer"
BUFFER_PATH = OUTPUT_DIR / "session_buffer.json"

KST = timezone(timedelta(hours=9))

# Export 컬럼 정의
COLUMNS = ["#", "source_text", "text_type", "ko", "ja", "en", "zh", "unregistered_terms", "timestamp"]
TARGET_LANGS = ["ko", "ja", "en", "zh"]


def load_buffer() -> dict:
    if not BUFFER_PATH.exists():
        print("❌ session_buffer.json 이 없습니다. 번역 기록이 없습니다.")
        sys.exit(1)
    with open(BUFFER_PATH, encoding="utf-8") as f:
        return json.load(f)


def buffer_to_rows(entries: list) -> list[dict]:
    """버퍼 엔트리를 Export용 행 데이터로 변환"""
    rows = []
    for entry in entries:
        translations = entry.get("translations", {})
        unregistered = entry.get("unregistered_terms", [])
        row = {
            "#": entry.get("id", ""),
            "source_text": entry.get("source_text", ""),
            "text_type": entry.get("text_type", ""),
            "unregistered_terms": ", ".join(unregistered) if unregistered else "",
            "timestamp": entry.get("timestamp", "")[:19].replace("T", " "),
        }
        for lang in TARGET_LANGS:
            row[lang] = translations.get(lang, "")
        rows.append(row)
    return rows


def export_xlsx(rows: list[dict], filepath: Path):
    """xlsx 파일 생성 (서식 포함)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("❌ openpyxl 이 없습니다. 설치: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "번역 결과"

    # 헤더 스타일
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 언어별 헤더 색상
    lang_fills = {
        "ko": PatternFill("solid", fgColor="E8F0FE"),
        "ja": PatternFill("solid", fgColor="FEF9E8"),
        "en": PatternFill("solid", fgColor="E8FEE8"),
        "zh": PatternFill("solid", fgColor="FEE8E8"),
    }

    # 컬럼 너비 설정
    col_widths = {
        "#": 5, "source_text": 35, "text_type": 14,
        "ko": 30, "ja": 30, "en": 30, "zh": 30,
        "unregistered_terms": 25, "timestamp": 18,
    }

    # 헤더 행 출력
    col_labels = {
        "#": "#", "source_text": "원문", "text_type": "텍스트 유형",
        "ko": "🇰🇷 한국어", "ja": "🇯🇵 일본어", "en": "🇺🇸 영어", "zh": "🇨🇳 중국어",
        "unregistered_terms": "미지정 용어", "timestamp": "번역 시각",
    }

    for col_idx, col_key in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_labels.get(col_key, col_key))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_key, 15)

    ws.row_dimensions[1].height = 30

    # 데이터 행 출력
    for row_idx, row_data in enumerate(rows, start=2):
        is_even = row_idx % 2 == 0
        row_fill = PatternFill("solid", fgColor="F5F5F5") if is_even else None

        for col_idx, col_key in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_key, ""))
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if col_key in lang_fills:
                cell.fill = lang_fills[col_key]
            elif row_fill:
                cell.fill = row_fill

            if col_key == "#":
                cell.alignment = Alignment(horizontal="center", vertical="top")

        ws.row_dimensions[row_idx].height = 45

    # 헤더 행 고정
    ws.freeze_panes = "A2"

    wb.save(str(filepath))
    print(f"✅ xlsx 저장: {filepath}")


def export_csv(rows: list[dict], filepath: Path):
    """csv 파일 생성 (UTF-8 BOM, Excel 호환)"""
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    print(f"✅ csv 저장: {filepath}")


def main():
    fmt = "both"
    for arg in sys.argv[1:]:
        if arg.startswith("--format="):
            fmt = arg.split("=")[1]
        elif arg == "--format" and len(sys.argv) > sys.argv.index(arg) + 1:
            fmt = sys.argv[sys.argv.index(arg) + 1]

    buffer = load_buffer()
    entries = buffer.get("entries", [])

    if not entries:
        print("⚠️  세션 버퍼가 비어 있습니다. Export할 데이터가 없습니다.")
        sys.exit(0)

    rows = buffer_to_rows(entries)
    timestamp_str = datetime.now(KST).strftime("%Y%m%d_%H%M%S")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print(f"📤 Export 시작: {len(rows)}건 → {fmt.upper()}")

    if fmt in ("xlsx", "both"):
        xlsx_path = OUTPUT_DIR / f"translation_export_{timestamp_str}.xlsx"
        export_xlsx(rows, xlsx_path)

    if fmt in ("csv", "both"):
        csv_path = OUTPUT_DIR / f"translation_export_{timestamp_str}.csv"
        export_csv(rows, csv_path)

    print(f"🎉 Export 완료!")


if __name__ == "__main__":
    main()
