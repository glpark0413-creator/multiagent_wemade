#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
이벤트별 보상 추천 생성기 (v2 — 전체 이력 기반)

동작 흐름:
  1. 가장 많은 날짜형 탭을 포함한 xlsx 파일을 자동 탐색 (Readdocs/)
  2. 탐색된 파일을 scan_rewards_by_event 로직으로 완전 스캔 → 이벤트 유형별 보상 패턴
  3. 각 신규 탭의 이벤트 섹션에 대해:
       a) 참조 탭(직전 탭) 직접 비교
       b) 이력에서 유사 이벤트 섹션 N개 검색 → 보상 수량 통계
       c) 두 신호를 결합해 추천 결정
  4. reward_recommendation.json 저장 + 콘솔 테이블 출력

사용:
  python scripts/recommend_rewards.py                         # 자동 이력 탐색
  python scripts/recommend_rewards.py --hist-xlsx path.xlsx   # 이력 파일 명시
  python scripts/recommend_rewards.py --print                 # 테이블 콘솔 출력
  python scripts/recommend_rewards.py --tab 260625 260702     # 대상 탭 지정
"""
import importlib.util
import io
import json
import re
import sys
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl

# ─── 경로 설정 ────────────────────────────────────────────────────────────────
_BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _project_config import load_project_paths

_paths = load_project_paths()
_READDOCS = _BASE_DIR / "Readdocs"

# scan_rewards_by_event 모듈을 직접 import해 subprocess 없이 활용
_SCAN_MOD_PATH = Path(__file__).resolve().parent / "scan_rewards_by_event.py"
_spec = importlib.util.spec_from_file_location("_srbe", str(_SCAN_MOD_PATH))
_srbe = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_srbe)

# ─── 상수 ─────────────────────────────────────────────────────────────────────
CHANGE_THRESHOLD = 0.20        # ±20% 초과 시 변경 권장
TOP_N_SIMILAR    = 5           # 유사 이벤트 최대 검색 수
SEASON_YEAR_RE   = re.compile(r'\d{2}-\d{2}')   # '15-26' 등 시즌 연도 패턴

# ─── 신규 탭 → 참조 탭 매핑 ─────────────────────────────────────────────────
DEFAULT_SOURCE_MAP: dict[str, str] = {
    "260625": "260611",
    "260702": "260618",
}


# ══════════════════════════════════════════════════════════════════════════════
# 1. 이력 xlsx 자동 탐색
# ══════════════════════════════════════════════════════════════════════════════

def _count_date_tabs(xlsx_path: Path) -> int:
    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
        cnt = sum(1 for n in wb.sheetnames if re.match(r'^\d{6}$', str(n)))
        wb.close()
        return cnt
    except Exception:
        return 0


def find_best_historical_xlsx(project_id: str | None = None) -> Path | None:
    """
    Readdocs/ 디렉터리에서 가장 많은 날짜형 탭을 가진 xlsx 파일 반환.

    탐색 순서:
      1. Readdocs/[{project_id}]*.xlsx  (직접 업로드 파일)
      2. Readdocs/projects/{project_id}_*.xlsx  (다운로드 파일)
      현재 프로젝트(current_project.json) source_xlsx 는 제외 (신규 탭 파일과 혼동 방지)
    """
    pid = project_id or (
        _paths.project_id if _paths else None
    )

    candidates: list[tuple[Path, int]] = []

    # 1순위: Readdocs/ 루트의 직접 파일
    for f in _READDOCS.glob("*.xlsx"):
        if pid and f"[{pid}]" not in f.name and pid not in f.name:
            continue
        cnt = _count_date_tabs(f)
        if cnt >= 2:
            candidates.append((f, cnt))

    # 2순위: Readdocs/projects/ 다운로드 파일
    proj_dir = _READDOCS / "projects"
    if proj_dir.exists():
        for f in proj_dir.glob(f"{pid}_*.xlsx" if pid else "*.xlsx"):
            cnt = _count_date_tabs(f)
            if cnt >= 2:
                candidates.append((f, cnt))

    if not candidates:
        return None

    # 탭 수 내림차순 정렬 → 가장 많은 파일 반환
    candidates.sort(key=lambda x: -x[1])
    best_path, best_cnt = candidates[0]
    print(f"  [이력 xlsx] {best_path.name}  (날짜형 탭 {best_cnt}개)")
    return best_path


# ══════════════════════════════════════════════════════════════════════════════
# 2. 이력 스캔 (scan_rewards_by_event 로직 직접 호출)
# ══════════════════════════════════════════════════════════════════════════════

def scan_xlsx_for_history(xlsx_path: Path) -> dict:
    """
    xlsx_path 를 scan_rewards_by_event 로직으로 스캔.
    반환: { event_type_patterns: {...}, per_tab_sections: {...} }
    """
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    date_tabs = [n for n in wb.sheetnames if _srbe.is_date_tab(n)]

    all_tab_sections: dict[str, list] = {}
    for tab in date_tabs:
        sections = _srbe.scan_tab_by_section(wb[tab])
        all_tab_sections[tab] = sections

    patterns = _srbe.build_event_type_patterns(all_tab_sections)
    wb.close()

    return {
        "source":              str(xlsx_path),
        "date_tabs":           date_tabs,
        "event_type_patterns": patterns,
        "per_tab_sections":    all_tab_sections,
    }


# ══════════════════════════════════════════════════════════════════════════════
# 3. 유사 이벤트 섹션 검색
# ══════════════════════════════════════════════════════════════════════════════

def _title_similarity(a: str, b: str) -> float:
    """두 이벤트 제목 유사도 (SequenceMatcher + 공통 키워드 보너스)."""
    base = SequenceMatcher(None, a, b).ratio()
    # 공통 단어(2자 이상) 비율 보너스
    words_a = set(re.findall(r'[가-힣a-zA-Z0-9]{2,}', a))
    words_b = set(re.findall(r'[가-힣a-zA-Z0-9]{2,}', b))
    if words_a and words_b:
        overlap = len(words_a & words_b) / max(len(words_a), len(words_b))
        return base * 0.6 + overlap * 0.4
    return base


def find_similar_sections(
    target_title: str,
    target_etype: str,
    hist_per_tab: dict,
    top_n: int = TOP_N_SIMILAR,
) -> list[dict]:
    """
    이력 전체 탭에서 대상 섹션과 가장 유사한 섹션 top_n 개 반환.

    조건:
      - event_type 동일 (필수)
      - 제목 유사도 기준 상위 N개
    반환 항목:
      { tab, title, event_type, similarity, reward_rows, ... }
    """
    results: list[tuple[float, dict]] = []

    for tab, sections in hist_per_tab.items():
        for sec in sections:
            if sec["event_type"] != target_etype:
                continue
            sim = _title_similarity(target_title, sec["title"])
            results.append((sim, {**sec, "tab": tab, "similarity": sim}))

    results.sort(key=lambda x: -x[0])
    return [r for _, r in results[:top_n]]


# ══════════════════════════════════════════════════════════════════════════════
# 4. 보상 통계 산출 (유사 섹션 집합)
# ══════════════════════════════════════════════════════════════════════════════

def _extract_qty(rr: dict) -> int | None:
    nq = rr.get("nearest_quantity")
    if nq and nq.get("quantity"):
        return nq["quantity"].get("value")
    qic = rr.get("quantity_in_cell")
    if qic:
        return qic.get("value")
    return None


def _get_qty_cell(rr: dict) -> str | None:
    nq = rr.get("nearest_quantity")
    return nq.get("cell") if nq else None


def build_reward_stats_from_similar(
    similar_sections: list[dict],
    reward_type: str,
) -> dict | None:
    """
    유사 섹션 목록에서 특정 보상 유형의 수량 통계 산출.
    반환: { min, max, avg, samples, seen_tabs } 또는 None
    """
    samples = []
    seen = []
    for sec in similar_sections:
        for rr in sec.get("reward_rows", []):
            if rr.get("reward_type") != reward_type:
                continue
            qty = _extract_qty(rr)
            if qty is not None:
                samples.append(qty)
                seen.append(sec.get("tab", "?"))

    if not samples:
        return None
    return {
        "min":       min(samples),
        "max":       max(samples),
        "avg":       round(sum(samples) / len(samples)),
        "samples":   len(samples),
        "seen_tabs": sorted(set(seen)),
    }


# ══════════════════════════════════════════════════════════════════════════════
# 5-A. 아이템 명칭 추천 (이력 기반)
# ══════════════════════════════════════════════════════════════════════════════

def _normalize_name(s: str) -> str:
    """공백·특수문자 제거 후 소문자화 — 명칭 유사도 비교용."""
    return re.sub(r'[\s\-_·]', '', s).lower()


def find_suggested_name(
    reward_name:      str,
    reward_type:      str,
    similar_sections: list[dict],
    hist_per_tab:     dict,
) -> str | None:
    """
    이력에서 현재 아이템명의 추천 대안을 반환.

    탐색 기준 (엄격):
      1. 시즌 연도 패턴(`NN-NN`) 포함 → 베이스명(연도 제거) 동일 + 다른 연도인 이름을 최신 탭에서 탐색
      2. 시즌 연도 없음 → 유사 섹션에서 현재명과 정규화 유사도 ≥0.6인 다른 이름이 ≥2탭 등장 시만 제안

    변경 불필요 시 None 반환.
    """
    has_season = bool(SEASON_YEAR_RE.search(reward_name))
    norm_current = _normalize_name(reward_name)

    if has_season:
        base_name = SEASON_YEAR_RE.sub("", reward_name).strip()
        norm_base = _normalize_name(base_name)
        sorted_tabs = sorted(hist_per_tab.keys(), reverse=True)

        name_by_tab: dict[str, str] = {}
        for tab in sorted_tabs:
            for sec in hist_per_tab[tab]:
                for rr in sec.get("reward_rows", []):
                    rname = rr.get("reward_name", "")
                    rbase = SEASON_YEAR_RE.sub("", rname).strip()
                    # 동일 reward_type + 베이스명 동일 + 연도만 다름
                    if (rr.get("reward_type") == reward_type
                            and _normalize_name(rbase) == norm_base
                            and rname != reward_name):
                        if tab not in name_by_tab:
                            name_by_tab[tab] = rname

        if name_by_tab:
            latest_tab = sorted(name_by_tab.keys(), reverse=True)[0]
            return name_by_tab[latest_tab]

    # 시즌 연도가 없는 일반 아이템은 자동 명칭 추천 대상 아님
    # 이유: reward_type 이 동일해도 등급(일반/고급/최상급)·종류가 달라
    #       오탐이 빈발함 — 사용자가 직접 "아이템명: 구이름 → 새이름" 으로 입력
    return None


# ══════════════════════════════════════════════════════════════════════════════
# 5-B. 추천 결정
# ══════════════════════════════════════════════════════════════════════════════

def determine_recommendation(
    reward_name:   str,
    reward_type:   str,
    current_qty:   int | None,
    is_pack:       bool,
    source_qty:    int | None,
    sim_stats:     dict | None,   # 유사 섹션 통계
    hist_avg:      int | None,    # event_type 전체 평균 (보조)
) -> dict:
    """
    신호 우선순위:
      1. 팩형 / 수량 미감지 — 즉시 처리
      2. 시즌 명칭 포함 — 명칭 검토 플래그
      3. 참조 탭 직접 비교 (source_qty)
      4. 유사 이벤트 통계 (sim_stats) — 가장 신뢰도 높은 역사 신호
      5. event_type 전체 평균 (hist_avg) — 최후 폴백
    """
    name_review = bool(SEASON_YEAR_RE.search(reward_name))

    def _ctx():
        parts = []
        if source_qty is not None:
            parts.append(f"참조={source_qty:,}")
        if sim_stats:
            parts.append(
                f"유사{sim_stats['samples']}탭 avg={sim_stats['avg']:,}"
                f" [{sim_stats['min']:,}~{sim_stats['max']:,}]"
            )
        elif hist_avg is not None:
            parts.append(f"유형평균={hist_avg:,}")
        return " | ".join(parts) if parts else "데이터 없음"

    # ── 팩형 ─────────────────────────────────────────────────────────────────
    if is_pack:
        action = "명칭_검토" if name_review else "유지"
        icon   = "📝" if name_review else "✅"
        reason = "팩형 — 시즌 명칭 포함" if name_review else "팩형 보상"
        return {"action": action, "icon": icon, "reason": reason,
                "suggested_qty": None, "sim_stats": sim_stats,
                "source_qty": source_qty, "hist_avg": hist_avg}

    # ── 수량 미감지 ──────────────────────────────────────────────────────────
    if current_qty is None:
        return {"action": "수동_확인", "icon": "❓",
                "reason": "수량 미감지", "suggested_qty": None,
                "sim_stats": sim_stats, "source_qty": source_qty, "hist_avg": hist_avg}

    # ── 명칭 검토 플래그 ─────────────────────────────────────────────────
    if name_review:
        ctx = _ctx()
        return {"action": "명칭_검토", "icon": "📝",
                "reason": f"시즌 명칭 포함 ({ctx})",
                "suggested_qty": current_qty,
                "sim_stats": sim_stats, "source_qty": source_qty, "hist_avg": hist_avg}

    # ── 비교 우선순위 ──────────────────────────────────────────────────────
    #
    # 1순위: 참조 탭 직접 비교
    #   - 현재 == source_qty → 항상 "유지" (sim_stats는 참고만)
    #     이유: 신규 탭이 참조 탭을 그대로 복사한 경우, 티어별 보상을
    #           pooled avg와 비교하면 거짓 권장이 대거 발생함
    #   - 현재 != source_qty (사용자가 직접 수정) → sim_stats 로 검증
    #
    # 2순위: sim_stats (유사 이벤트, ≥2샘플)
    # 3순위: hist_avg (이벤트 유형 전체 평균, 보조)

    if source_qty is not None:
        if current_qty == source_qty:
            # 참조 탭과 동일 → 유지 (이력 avg를 reason에 참고로 표시)
            sim_note = (
                f" | 이력 avg {sim_stats['avg']:,}({sim_stats['samples']}탭)"
                if sim_stats else ""
            )
            return {"action": "유지", "icon": "✅",
                    "reason": f"참조 탭과 동일 ({current_qty:,}){sim_note}",
                    "suggested_qty": current_qty,
                    "sim_stats": sim_stats, "source_qty": source_qty, "hist_avg": hist_avg}

        # 참조 탭과 다른 값 → sim_stats 또는 source_qty 로 비교
        if sim_stats and sim_stats["samples"] >= 2:
            compare_to    = sim_stats["avg"]
            compare_label = f"유사 {sim_stats['samples']}탭 평균 {sim_stats['avg']:,}"
        else:
            compare_to    = source_qty
            compare_label = f"참조 탭 {source_qty:,}"
    elif sim_stats and sim_stats["samples"] >= 2:
        compare_to    = sim_stats["avg"]
        compare_label = f"유사 {sim_stats['samples']}탭 평균 {sim_stats['avg']:,}"
    elif hist_avg is not None:
        compare_to    = hist_avg
        compare_label = f"유형 평균 {hist_avg:,}"
    else:
        return {"action": "유지", "icon": "✅",
                "reason": "비교 데이터 없음 — 유지",
                "suggested_qty": current_qty,
                "sim_stats": sim_stats, "source_qty": source_qty, "hist_avg": hist_avg}

    diff = (current_qty - compare_to) / compare_to

    if diff > CHANGE_THRESHOLD:
        return {"action": "하향_검토", "icon": "↓",
                "reason": f"현재 {current_qty:,} > {compare_label} (+{diff*100:.0f}%)",
                "suggested_qty": compare_to,
                "sim_stats": sim_stats, "source_qty": source_qty, "hist_avg": hist_avg}
    if diff < -CHANGE_THRESHOLD:
        return {"action": "상향_권장", "icon": "↑",
                "reason": f"현재 {current_qty:,} < {compare_label} ({diff*100:.0f}%)",
                "suggested_qty": compare_to,
                "sim_stats": sim_stats, "source_qty": source_qty, "hist_avg": hist_avg}

    return {"action": "유지", "icon": "✅",
            "reason": f"현재 {current_qty:,} ≈ {compare_label} (±{diff*100:.0f}%)",
            "suggested_qty": current_qty,
            "sim_stats": sim_stats, "source_qty": source_qty, "hist_avg": hist_avg}


# ══════════════════════════════════════════════════════════════════════════════
# 6-0. 보상 순서 패턴 학습 및 추천
# ══════════════════════════════════════════════════════════════════════════════

def _order_key(rr: dict, use_type: bool) -> str:
    """
    보상 행의 순서 비교용 정규화 키 생성.
    use_type=True  → reward_type (카테고리 수준, 변동 작음)
    use_type=False → 시즌 연도 제거 + 앞 30자 (아이템명 수준)
    """
    if use_type:
        return rr.get("reward_type", "기타")
    name = SEASON_YEAR_RE.sub("", rr.get("reward_name", "")).strip()
    return name[:30] if name else rr.get("reward_type", "기타")


def build_canonical_reward_sequence(
    similar_sections: list[dict],
    use_type: bool = True,
) -> list[str]:
    """
    유사 이벤트 섹션들에서 보상 순서 패턴을 학습해 정식 순서 키 리스트 반환.

    알고리즘:
      각 유사 섹션에서 키(reward_type 또는 정규화 이름)의 첫 등장 상대 위치(0~1)를
      수집하고, 절반 이상 섹션에서 관찰된 키에 대해 평균 위치로 오름차순 정렬한다.
    """
    if not similar_sections:
        return []

    key_pos: dict[str, list[float]] = {}

    for sec in similar_sections:
        rows  = sec.get("reward_rows", [])
        total = len(rows)
        if total < 2:
            continue
        seen: dict[str, float] = {}
        for idx, rr in enumerate(rows):
            k = _order_key(rr, use_type)
            if k not in seen:
                seen[k] = idx / total          # 0=맨 앞, 1=맨 뒤 정규화
        for k, pos in seen.items():
            key_pos.setdefault(k, []).append(pos)

    threshold = max(1, len(similar_sections) // 2)
    avg = {
        k: sum(ps) / len(ps)
        for k, ps in key_pos.items()
        if len(ps) >= threshold
    }
    return sorted(avg.keys(), key=lambda k: avg[k])


def analyze_reward_order(
    current_rewards: list[dict],
    canonical_seq:   list[str],
    use_type:        bool = True,
) -> dict:
    """
    현재 보상 순서와 이력 정식 순서를 비교해 순서 이슈를 감지한다.

    반환 dict:
      match_score       float   Kendall-tau 쌍 순서 일치율 (0~1)
      has_order_issue   bool    True → 순서 변경 권장 (match_score < 0.75)
      current_type_seq  [str]   현재 키 첫 등장 순서
      canonical_type_seq[str]   이력 정식 순서 (현재에 없는 키 제외)
      out_of_order      [(key, cur_rank, rec_rank), ...]
      recommended_rows  [rr, ...] 권장 순서로 재배열된 보상 행
    """
    def _k(rr):
        return _order_key(rr, use_type)

    if not canonical_seq or len(current_rewards) < 3:
        return {
            "match_score": 1.0, "has_order_issue": False,
            "current_type_seq": [], "canonical_type_seq": [],
            "out_of_order": [], "recommended_rows": list(current_rewards),
        }

    # 현재 키 첫 등장 순서
    seen: set = set()
    cur_seq: list[str] = []
    for rr in current_rewards:
        k = _k(rr)
        if k not in seen:
            seen.add(k)
            cur_seq.append(k)

    # canonical에서 현재 존재하는 것만 → + 신규 키 뒤에
    target = [k for k in canonical_seq if k in cur_seq]
    novel  = [k for k in cur_seq if k not in canonical_seq]
    full_target = target + novel

    # Kendall-tau 쌍 순서 일치율
    n = len(cur_seq)
    correct = sum(
        1
        for i in range(n)
        for j in range(i + 1, n)
        if (full_target.index(cur_seq[i]) if cur_seq[i] in full_target else n)
         < (full_target.index(cur_seq[j]) if cur_seq[j] in full_target else n)
    )
    total_pairs = n * (n - 1) // 2
    score = round(correct / total_pairs, 2) if total_pairs > 0 else 1.0

    # 자리 이탈 항목
    out_of_order = [
        (k, cur_rank + 1, full_target.index(k) + 1)
        for cur_rank, k in enumerate(cur_seq)
        if k in full_target and full_target.index(k) != cur_rank
    ]

    # 권장 순서로 재배열 (같은 키 그룹 내 원본 상대 순서 유지)
    key_to_rows: dict[str, list] = {}
    for rr in current_rewards:
        key_to_rows.setdefault(_k(rr), []).append(rr)
    recommended_rows = []
    for k in full_target:
        recommended_rows.extend(key_to_rows.get(k, []))

    return {
        "match_score":        score,
        "has_order_issue":    score < 0.75,
        "current_type_seq":   cur_seq,
        "canonical_type_seq": full_target,
        "out_of_order":       out_of_order,
        "recommended_rows":   recommended_rows,
    }


# ══════════════════════════════════════════════════════════════════════════════
# 6. 섹션 비교 (신규 탭 × 이력)
# ══════════════════════════════════════════════════════════════════════════════

def compare_sections(
    new_sections:     list,
    hist_data:        dict,    # scan_xlsx_for_history 반환값
    source_sections:  list | None,
) -> list:
    """
    신규 탭 섹션 목록을 이력 데이터와 비교해 섹션별 보상 추천 반환.
    """
    hist_patterns  = hist_data.get("event_type_patterns", {})
    hist_per_tab   = hist_data.get("per_tab_sections", {})

    results = []

    for idx, section in enumerate(new_sections):
        etype = section["event_type"]
        title = section["title"]

        # 날짜 헤더 행 스킵 ('MM.DD_ Event' 형식)
        if re.match(r"^\d{2}\.\d{2}", title):
            continue

        reward_rows = section.get("reward_rows", [])

        # 참조 탭 같은 인덱스 섹션
        src_section = source_sections[idx] if (source_sections and idx < len(source_sections)) else None
        src_rewards = src_section.get("reward_rows", []) if src_section else []

        # 이력에서 유사 이벤트 섹션 검색
        similar = find_similar_sections(title, etype, hist_per_tab)

        # ── 보상 순서 패턴 학습 및 비교 ────────────────────────────────────
        # type 수준(카테고리): 시즌마다 이름이 바뀌어도 타입 순서는 유지됨
        canonical_type_seq  = build_canonical_reward_sequence(similar, use_type=True)
        order_analysis_type = analyze_reward_order(reward_rows, canonical_type_seq, use_type=True)
        # name 수준(아이템명): 보상 수가 적은 섹션(<= 20행)에서 추가 분석
        if len(reward_rows) <= 20:
            canonical_name_seq  = build_canonical_reward_sequence(similar, use_type=False)
            order_analysis_name = analyze_reward_order(reward_rows, canonical_name_seq, use_type=False)
        else:
            order_analysis_name = {"match_score": 1.0, "has_order_issue": False,
                                   "current_type_seq": [], "canonical_type_seq": [],
                                   "out_of_order": [], "recommended_rows": list(reward_rows)}

        sec_result = {
            "index":              idx,
            "event_title":        title,
            "event_type":         etype,
            "title_cell":         section.get("title_cell", ""),
            "start_row":          section.get("start_row"),
            "end_row":            section.get("end_row"),
            "similar_count":      len(similar),
            "similar_tabs":       [s["tab"] for s in similar],
            "order_type":         order_analysis_type,   # reward_type 기반 순서 분석
            "order_name":         order_analysis_name,   # 아이템명 기반 순서 분석 (소규모 섹션)
            "rewards":            [],
        }

        for ridx, rr in enumerate(reward_rows):
            rname      = rr.get("reward_name", "")
            rtype      = rr.get("reward_type", "기타")
            current_qty = _extract_qty(rr)
            qty_cell   = _get_qty_cell(rr)
            is_pack    = not rr.get("has_quantity", False) and current_qty is None

            # 참조 탭 같은 위치 보상
            source_qty: int | None = None
            if ridx < len(src_rewards):
                source_qty = _extract_qty(src_rewards[ridx])
            elif src_rewards:
                best_sim, best_qty = 0.0, None
                for sr in src_rewards:
                    sim = SequenceMatcher(None, rname, sr.get("reward_name", "")).ratio()
                    if sim > best_sim:
                        best_sim, best_qty = sim, _extract_qty(sr)
                source_qty = best_qty

            # 유사 섹션 통계 (같은 보상 유형)
            sim_stats = build_reward_stats_from_similar(similar, rtype)

            # event_type 전체 평균 (보조용)
            hist_avg: int | None = None
            tp = hist_patterns.get(etype, {})
            qs = tp.get("quantity_stats", {}).get(rtype)
            if qs:
                hist_avg = qs.get("avg")

            rec = determine_recommendation(
                rname, rtype, current_qty, is_pack,
                source_qty, sim_stats, hist_avg,
            )

            # ── 아이템 명칭 추천 (이력 기반) ──────────────────────────────
            suggested_name = find_suggested_name(rname, rtype, similar, hist_per_tab)
            # 추천 명칭이 있으면 action 에 명칭_변경 플래그 추가
            if suggested_name:
                rec["suggested_name"] = suggested_name
                # 명칭_검토 가 아닌 유지 상태에서도 명칭 변경 제안이 있으면 표시
                if rec["action"] == "유지":
                    rec["action"] = "명칭_변경_권장"
                    rec["icon"]   = "📝"
                    rec["reason"] = f"이력 기반 명칭 업데이트: '{rname}' → '{suggested_name}'"
            else:
                rec["suggested_name"] = None

            sec_result["rewards"].append({
                "reward_idx":    ridx,
                "reward_name":   rname,
                "reward_type":   rtype,
                "item_cell":     rr.get("cell", ""),
                "qty_cell":      qty_cell,
                "current_qty":   current_qty,
                "source_qty":    source_qty,
                "sim_stats":     sim_stats,
                "hist_avg":      hist_avg,
                "is_pack":       is_pack,
                "recommendation": rec,
            })

        results.append(sec_result)

    return results


# ══════════════════════════════════════════════════════════════════════════════
# 7. 메인
# ══════════════════════════════════════════════════════════════════════════════

def main(
    target_tabs:    list[str] | None = None,
    hist_xlsx_path: Path | None = None,
):
    # ── 경로 결정 ──────────────────────────────────────────────────────────
    if _paths:
        hist_json_path = _paths.reward_by_event        # 이미 스캔된 이력 JSON
        new_path       = _paths.reward_new_tabs
        out_path       = _paths.work_dir / "reward_recommendation.json"
        pid            = _paths.project_id
    else:
        hist_json_path = _BASE_DIR / "output" / "json" / "reward_by_event.json"
        new_path       = _BASE_DIR / "output" / "json" / "reward_new_tabs.json"
        out_path       = _BASE_DIR / "output" / "json" / "reward_recommendation.json"
        pid            = None

    if not new_path.exists():
        print(f"[오류] 신규 탭 보상 파일 없음: {new_path}")
        print("  → scan_rewards_by_event.py '{output_xlsx}' 를 먼저 실행하세요.")
        sys.exit(1)

    new_data = json.loads(new_path.read_text(encoding="utf-8"))

    # ── 이력 데이터 로드 ─────────────────────────────────────────────────
    # 우선순위:
    #  1) --hist-xlsx 명시 → 직접 스캔 (다른 형식의 xlsx를 강제 사용)
    #  2) reward_by_event.json 이 이미 존재하고 탭 수 ≥ 2 → 그대로 활용
    #  3) 자동으로 best historical xlsx 탐색 → 스캔
    if hist_xlsx_path is not None and hist_xlsx_path.exists():
        print(f"\n[이력 스캔]  --hist-xlsx 명시: {hist_xlsx_path.name}")
        hist_data = scan_xlsx_for_history(hist_xlsx_path)
        print(f"  → 날짜형 탭 {len(hist_data['date_tabs'])}개 스캔 완료")

    elif hist_json_path.exists():
        hist_data = json.loads(hist_json_path.read_text(encoding="utf-8"))
        # per_tab_sections 키가 없으면(구버전 포맷) 빈 dict 보완
        hist_data.setdefault("per_tab_sections", {})
        scanned_tabs = list(hist_data.get("per_tab_sections", {}).keys())
        print(f"\n[이력 데이터]  reward_by_event.json 사용  (탭: {scanned_tabs})")

    else:
        hist_xlsx_path = find_best_historical_xlsx(pid)
        if hist_xlsx_path and hist_xlsx_path.exists():
            print(f"\n[이력 스캔 시작]  {hist_xlsx_path.name}")
            hist_data = scan_xlsx_for_history(hist_xlsx_path)
            print(f"  → 날짜형 탭 {len(hist_data['date_tabs'])}개 스캔 완료")
        else:
            print("[경고] 이력 데이터 없음 — 참조 탭 비교만 수행합니다.")
            hist_data = {"event_type_patterns": {}, "per_tab_sections": {}}

    new_sections_all = new_data.get("per_tab_sections", {})
    hist_sections    = hist_data.get("per_tab_sections", {})

    tabs_to_process = target_tabs or list(new_sections_all.keys())

    print(f"\n[보상 추천 생성]  대상 탭: {tabs_to_process}")

    recommendation: dict[str, list] = {}

    for tab in tabs_to_process:
        if tab not in new_sections_all:
            print(f"  [{tab}] 신규 탭 데이터 없음, 건너뜀")
            continue

        source_tab  = DEFAULT_SOURCE_MAP.get(tab)
        source_secs = hist_sections.get(source_tab) if source_tab else None

        # 이력 파일에 source_tab 이 없으면 new_sections_all 에서 탐색
        if source_secs is None and source_tab:
            source_secs = new_sections_all.get(source_tab)

        if source_secs:
            print(f"  [{tab}] 참조 탭: {source_tab}")
        else:
            print(f"  [{tab}] 참조 탭 미지정 — 이력 패턴만 사용")

        sec_recs = compare_sections(
            new_sections_all[tab], hist_data, source_secs
        )
        recommendation[tab] = sec_recs

        total   = sum(len(s["rewards"]) for s in sec_recs)
        changes = sum(
            1 for s in sec_recs
            for r in s["rewards"]
            if r["recommendation"]["action"] in ("상향_권장", "하향_검토", "명칭_검토")
        )
        print(f"    섹션 {len(sec_recs)}개, 보상 {total}개, 변경권장 {changes}개")

    # 이력 출처 메타데이터 정리
    hist_tabs_scanned = hist_data.get("date_tabs") or list(hist_data.get("per_tab_sections", {}).keys())
    hist_source_label = (
        str(hist_xlsx_path) if hist_xlsx_path else
        str(hist_json_path) if hist_json_path.exists() else "없음"
    )

    result = {
        "generated_at":      __import__("datetime").datetime.now().isoformat(timespec="seconds"),
        "hist_source":       hist_source_label,
        "hist_tabs_scanned": hist_tabs_scanned,
        "new_source":        str(new_path),
        "source_map":        DEFAULT_SOURCE_MAP,
        "tabs":              recommendation,
    }
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n저장 완료: {out_path}")
    return result


# ══════════════════════════════════════════════════════════════════════════════
# 8. 콘솔 테이블 출력 (에이전트용)
# ══════════════════════════════════════════════════════════════════════════════

def print_recommendation_table(rec_data: dict, tab: str):
    """CLAUDE.md 형식의 이벤트별 보상 추천 표 출력."""
    sections     = rec_data["tabs"].get(tab, [])
    source_tab   = rec_data["source_map"].get(tab, "?")
    hist_tabs    = rec_data.get("hist_tabs_scanned", [])
    hist_xlsx    = rec_data.get("hist_source", "?")

    print(f"\n[이벤트별 보상 추천] — {tab} 탭")
    print(f"  참조 탭  : {source_tab}")
    print(f"  이력 기반: {hist_xlsx}  ({len(hist_tabs)}개 탭)")
    SEP = "─" * 115
    HDR = (
        f" {'#':>2} | {'이벤트명':<30} | {'보상 아이템':<26} | "
        f"{'현재':>9} | {'유사이력avg':>10} | {'참조탭':>9} | {'추천':<14} | 변경"
    )
    print(SEP)
    print(HDR)
    print(SEP)

    sec_num = 1
    for sec in sections:
        for ridx, r in enumerate(sec["rewards"]):
            icon   = r["recommendation"]["icon"]
            action = r["recommendation"]["action"]
            cur_q  = f"{r['current_qty']:,}"  if r["current_qty"] is not None else "(팩형)"
            src_q  = f"{r['source_qty']:,}"   if r["source_qty"]  is not None else "—"
            ss     = r.get("sim_stats")
            avg_q  = f"{ss['avg']:,}({ss['samples']}탭)" if ss else "—"
            ev_lbl = sec["event_title"][:30] if ridx == 0 else ""
            print(
                f" {sec_num:>2} | {ev_lbl:<30} | {r['reward_name'][:26]:<26} | "
                f"{cur_q:>9} | {avg_q:>10} | {src_q:>9} | {action:<14} | {icon}"
            )
        print(f"    {'':30}  {'유사 탭: ' + ','.join(sec['similar_tabs'][:3]):<40}")
        sec_num += 1

    print(SEP)
    # 변경 권장 항목만 요약
    flagged = [
        (si + 1, r)
        for si, sec in enumerate(sections)
        for r in sec["rewards"]
        if r["recommendation"]["action"] in ("상향_권장", "하향_검토", "명칭_검토")
    ]
    if flagged:
        print("\n[변경 권장 요약]")
        for sn, r in flagged:
            rec = r["recommendation"]
            print(f"  섹션{sn} {r['reward_name'][:24]:<24} {rec['icon']} {rec['reason']}")

    print(
        "\n→ 변경할 번호와 새 수량을 입력하거나 '권장' / '전체 승인' / '건너뜀' 을 입력하세요.\n"
        "  예)  '3번 다이아 500→1000'   '권장'   '전체 승인'   '건너뜀'"
    )


# ══════════════════════════════════════════════════════════════════════════════
# 9. 이벤트별 개별 테이블 출력 (스크린샷 형식)
# ══════════════════════════════════════════════════════════════════════════════

def print_per_event_table(rec_data: dict, tab: str):
    """
    이벤트 섹션마다 개별 보상 테이블을 출력한다.

    형식:
      ──── [1/7] 올스타 직행! 14일 출석 이벤트! ────
       보상 아이템                 │ 현재        │ 이력avg(N탭)  │ 추천
      ─────────────────────────────────────────────────
       골드                       │   500,000   │ 600,000(6탭)  │ ✅ 유지
       최상급 선수 카드 팩 티켓   │        10   │    12(6탭)    │ ✅ 유지
       컨셉카드 선수팩 티켓       │    (팩형)   │          —    │ ✅ 유지
       All 강화권 (A)             │       500   │   154(5탭)    │ ✅ 유지
      ─────────────────────────────────────────────────
      유사 이력: 260611, 260423, 260514
    """
    sections   = rec_data["tabs"].get(tab, [])
    source_tab = rec_data["source_map"].get(tab, "?")
    total      = len(sections)

    print(f"\n{'═' * 68}")
    print(f"[보상 추천] {tab} 탭  (참조: {source_tab} / 이력 {len(rec_data.get('hist_tabs_scanned', []))}개 탭)")
    print(f"{'═' * 68}")

    all_flagged:    list[tuple[int, str, dict]] = []  # (sec_idx, reward_name, rec)
    order_flagged:  list[tuple[int, str, dict]] = []  # (sec_idx, title, order_analysis)

    W_NAME = 30
    W_CUR  = 11
    W_AVG  = 14
    ROW_SEP = "─" * (W_NAME + W_CUR + W_AVG + 20)

    for si, sec in enumerate(sections, 1):
        title    = sec["event_title"]
        etype    = sec["event_type"]
        rewards  = sec["rewards"]
        sim_tabs = ", ".join(sec.get("similar_tabs", [])[:3]) or "—"

        # ── 순서 분석 결과 선택 (name 우선, type 보조) ────────────────────
        ord_n = sec.get("order_name", {})
        ord_t = sec.get("order_type", {})
        # name 분석이 있고 이슈가 있으면 name 사용, 아니면 type 사용
        if ord_n.get("has_order_issue"):
            ord_main = ord_n
            ord_label = "아이템명"
        elif ord_t.get("has_order_issue"):
            ord_main = ord_t
            ord_label = "보상유형"
        else:
            ord_main  = ord_t
            ord_label = "보상유형"

        has_order_issue = ord_main.get("has_order_issue", False)

        # ── 섹션 헤더 ─────────────────────────────────────────────────────
        hdr_text = f" [{si}/{total}] {title}"
        print(f"\n{hdr_text}")
        print(f"  유형: {etype}  │  유사: {sim_tabs}")

        # ── 순서 이슈 표시 ────────────────────────────────────────────────
        if has_order_issue:
            score_pct = int(ord_main["match_score"] * 100)
            cur_seq   = " → ".join(ord_main.get("current_type_seq", [])[:7])
            rec_seq   = " → ".join(ord_main.get("canonical_type_seq", [])[:7])
            print(f"  🔄 순서 변경 권장 ({ord_label} 일치율 {score_pct}%)")
            print(f"     현재  : {cur_seq}")
            print(f"     권장  : {rec_seq}")
            order_flagged.append((si, title, ord_main))

        # ── 현재 보상 테이블 ──────────────────────────────────────────────
        print(f"  {ROW_SEP}")
        print(
            f"  {'보상 아이템':<{W_NAME}} │ {'현재':>{W_CUR}} │ {'이력avg':>{W_AVG}} │ 추천"
        )
        print(f"  {ROW_SEP}")

        if not rewards:
            print(f"  {'(보상 없음 또는 미감지)':<{W_NAME}}")
        else:
            reward_key_set = {
                _order_key({"reward_type": r["reward_type"], "reward_name": r["reward_name"]},
                           use_type=(ord_label == "보상유형"))
                for r in rewards
            }
            # 권장 순서의 키 집합으로 순서 이탈 아이템 표시
            out_keys = {k for k, _, _ in ord_main.get("out_of_order", [])}

            for r in rewards:
                rec    = r["recommendation"]
                icon   = rec["icon"]
                action = rec["action"]
                cur_q  = f"{r['current_qty']:,}" if r["current_qty"] is not None else "(팩형)"
                ss     = r.get("sim_stats")
                avg_q  = f"{ss['avg']:,}({ss['samples']}탭)" if ss else "—"
                name   = r["reward_name"][:W_NAME]

                # 순서 이탈 아이템에 🔄 마킹
                rkey = _order_key(
                    {"reward_type": r["reward_type"], "reward_name": r["reward_name"]},
                    use_type=(ord_label == "보상유형"),
                )
                order_marker = " 🔄" if (has_order_issue and rkey in out_keys) else ""

                print(
                    f"  {name:<{W_NAME}} │ {cur_q:>{W_CUR}} │ {avg_q:>{W_AVG}} │ {icon} {action}{order_marker}"
                )

                if action not in ("유지",):
                    all_flagged.append((si, r["reward_name"], rec))

        print(f"  {ROW_SEP}")

        # ── 권장 순서 보상 테이블 (순서 이슈 있을 때만) ──────────────────
        if has_order_issue:
            rec_rows = ord_main.get("recommended_rows", [])
            print(f"  ↳ 권장 순서 ({ord_label} 기준, {len(rec_rows)}개 아이템)")
            print(f"  {'─' * (W_NAME + 4)}")
            for pos, rr in enumerate(rec_rows, 1):
                raw_name = rr.get("reward_name", "")
                qty      = rr.get("nearest_quantity") or {}
                qty_val  = qty.get("quantity", {}).get("value") if isinstance(qty, dict) else None
                qty_str  = f"{qty_val:,}" if qty_val is not None else "(팩형)"
                print(f"  {pos:>2}. {raw_name[:W_NAME-4]:<{W_NAME-4}} {qty_str:>{W_CUR}}")
            print(f"  {'─' * (W_NAME + 4)}")

    # ── 전체 변경 권장 요약 ───────────────────────────────────────────────
    print(f"\n{'═' * 68}")
    any_issue = bool(all_flagged or order_flagged)

    if all_flagged:
        print(f"[보상 수량·명칭 변경 권장] — {len(all_flagged)}건")
        for sn, rname, rec in all_flagged:
            print(f"  섹션{sn} {rname[:28]:<28}  {rec['icon']} {rec['reason']}")
    else:
        print("[보상 수량·명칭] 변경 권장 없음 ✅")

    if order_flagged:
        print(f"\n[보상 순서 변경 권장] — {len(order_flagged)}개 섹션")
        for sn, etitle, ord_info in order_flagged:
            score_pct  = int(ord_info["match_score"] * 100)
            out_items  = [k for k, _, _ in ord_info.get("out_of_order", [])]
            items_str  = ", ".join(out_items[:4])
            if len(out_items) > 4:
                items_str += f" 외 {len(out_items)-4}건"
            print(f"  섹션{sn} {etitle[:28]:<28}  🔄 일치율 {score_pct}%  ({items_str})")
    else:
        print("[보상 순서] 이력 패턴과 일치 ✅")

    if not any_issue:
        print("\n모든 이벤트 보상이 이력 패턴과 일치합니다.")

    print(
        f"\n{'─' * 68}\n"
        "→ 수정 방법:\n"
        "   '섹션번호 아이템명 수량'      예) '1 골드 600000'\n"
        "   '섹션번호 순서 적용'          예) '3 순서 적용' — 권장 순서로 xlsx 재배열\n"
        "   '권장'  — ↑/↓ 수량 권장 항목 일괄 적용\n"
        "   '전체 승인'  |  '건너뜀'"
    )


# ══════════════════════════════════════════════════════════════════════════════
# 10. AUTO-CONFIRM-HIGH 요약 출력
# ══════════════════════════════════════════════════════════════════════════════

_HIGH_ACTIONS = frozenset({"유지"})
_LOW_ACTIONS  = frozenset({"상향_권장", "하향_검토", "명칭_검토", "명칭_변경_권장", "수동_확인"})


def _print_auto_confirm_summary(rec_data: dict, tab: str) -> None:
    """
    HIGH 신뢰도(action=유지) 항목을 자동 확정으로 표시하고
    LOW 신뢰도 항목 수만 요약 출력한다.
    """
    sections = rec_data["tabs"].get(tab, [])

    high_count = sum(
        1 for sec in sections for r in sec["rewards"]
        if r["recommendation"]["action"] in _HIGH_ACTIONS
    )
    low_count = sum(
        1 for sec in sections for r in sec["rewards"]
        if r["recommendation"]["action"] in _LOW_ACTIONS
    )

    print(f"\n{'═' * 60}")
    print(f"[자동 확정 요약] {tab} 탭")
    print(f"  ✅ 자동 확정 (HIGH):  {high_count}개 항목")
    print(f"  ⚠  검토 필요 (LOW):  {low_count}개 항목")
    print(f"{'═' * 60}")

    if low_count == 0:
        print("  모든 항목이 자동 확정되었습니다.")
        return

    print("\n[검토 필요 항목]")
    for si, sec in enumerate(sections, 1):
        low_rewards = [r for r in sec["rewards"]
                       if r["recommendation"]["action"] in _LOW_ACTIONS]
        if not low_rewards:
            continue
        print(f"  섹션{si} {sec['event_title'][:30]}")
        for r in low_rewards:
            rec = r["recommendation"]
            print(f"    {rec['icon']} {r['reward_name'][:24]:<24}  {rec['reason']}")


def print_per_event_low_only(
    rec_data: dict,
    tab: str,
    auto_confirm_high: bool = True,
) -> None:
    """
    --per-event --low-only 조합:
    LOW 신뢰도 항목(변경권장)이 있는 섹션만 출력한다.
    HIGH 신뢰도 항목은 자동 확정으로 처리해 화면에 표시하지 않는다.
    """
    sections     = rec_data["tabs"].get(tab, [])
    source_tab   = rec_data["source_map"].get(tab, "?")
    hist_tabs    = rec_data.get("hist_tabs_scanned", [])
    total        = len(sections)

    # LOW 항목이 있는 섹션만 필터
    low_sections = [
        (si, sec) for si, sec in enumerate(sections, 1)
        if any(r["recommendation"]["action"] in _LOW_ACTIONS for r in sec["rewards"])
    ]

    high_total = sum(
        1 for sec in sections for r in sec["rewards"]
        if r["recommendation"]["action"] in _HIGH_ACTIONS
    )
    low_total = sum(
        1 for sec in sections for r in sec["rewards"]
        if r["recommendation"]["action"] in _LOW_ACTIONS
    )

    print(f"\n{'═' * 68}")
    print(f"[보상 검토 필요] {tab} 탭  (참조: {source_tab} / 이력 {len(hist_tabs)}개 탭)")
    print(f"  자동 확정 ✅ : {high_total}개 항목   검토 필요 ⚠ : {low_total}개 항목")
    print(f"{'═' * 68}")

    if not low_sections:
        print("  모든 보상 항목이 자동 확정되었습니다. 검토 불필요.")
        return

    W_NAME = 30
    W_CUR  = 11
    W_AVG  = 14
    ROW_SEP = "─" * (W_NAME + W_CUR + W_AVG + 20)

    for si, sec in low_sections:
        title    = sec["event_title"]
        etype    = sec["event_type"]
        sim_tabs = ", ".join(sec.get("similar_tabs", [])[:3]) or "—"
        low_r    = [r for r in sec["rewards"] if r["recommendation"]["action"] in _LOW_ACTIONS]
        high_r   = [r for r in sec["rewards"] if r["recommendation"]["action"] in _HIGH_ACTIONS]

        print(f"\n [{si}/{total}] {title}")
        print(f"  유형: {etype}  │  유사: {sim_tabs}")
        print(f"  ✅ 자동확정: {len(high_r)}개 항목 (검토 불필요)")
        print(f"  {ROW_SEP}")
        print(f"  {'보상 아이템':<{W_NAME}} │ {'현재':>{W_CUR}} │ {'이력avg':>{W_AVG}} │ 추천")
        print(f"  {ROW_SEP}")

        for r in low_r:
            rec   = r["recommendation"]
            icon  = rec["icon"]
            action = rec["action"]
            cur_q = f"{r['current_qty']:,}" if r["current_qty"] is not None else "(팩형)"
            ss    = r.get("sim_stats")
            avg_q = f"{ss['avg']:,}({ss['samples']}탭)" if ss else "—"
            name  = r["reward_name"][:W_NAME]
            print(f"  {name:<{W_NAME}} │ {cur_q:>{W_CUR}} │ {avg_q:>{W_AVG}} │ {icon} {action}")

        print(f"  {ROW_SEP}")

    print(f"\n{'─' * 68}")
    print(
        "→ 수정 방법:\n"
        "   '섹션번호 아이템명 수량'   예) '1 다이아 60'\n"
        "   '추천 수량으로'  — 추천값 적용\n"
        "   '건너뜀'"
    )


# ══════════════════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--hist-xlsx",         help="이력 xlsx 파일 명시 (없으면 자동 탐색)")
    parser.add_argument("--tab",               nargs="*", help="대상 탭 목록 (예: 260625 260702)")
    parser.add_argument("--print",             action="store_true", help="기존 통합 테이블 출력")
    parser.add_argument("--per-event",         action="store_true", help="이벤트별 개별 테이블 출력")
    parser.add_argument("--auto-confirm-high", action="store_true",
                        help="HIGH 신뢰도(action=유지) 항목 자동 확정 — 콘솔 출력 생략")
    parser.add_argument("--low-only",          action="store_true",
                        help="--per-event 시 LOW 신뢰도 항목(변경권장)만 출력")
    args = parser.parse_args()

    hist_path = Path(args.hist_xlsx) if args.hist_xlsx else None
    rec = main(args.tab, hist_path)

    if rec:
        if args.per_event:
            for tab in rec["tabs"]:
                if args.low_only:
                    print_per_event_low_only(rec, tab, auto_confirm_high=args.auto_confirm_high)
                else:
                    print_per_event_table(rec, tab)
        elif args.print:
            for tab in rec["tabs"]:
                print_recommendation_table(rec, tab)
        elif args.auto_confirm_high:
            # --auto-confirm-high 단독 사용: HIGH 자동 확정 요약만 출력
            for tab in rec["tabs"]:
                _print_auto_confirm_summary(rec, tab)
