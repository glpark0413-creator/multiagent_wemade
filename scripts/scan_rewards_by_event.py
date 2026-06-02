#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
이벤트 섹션별 보상 정밀 스캐너

동작:
  - xlsx 전체 날짜형 탭 스캔
  - B열의 "숫자." 패턴(이벤트 섹션 제목)으로 섹션 경계를 파악
  - 각 섹션 범위 내 보상 셀(명칭·수량·셀 좌표) 수집
  - 이벤트 유형별(출석·응모권·플레이미션·교환소 등) 역사적 패턴 통계 산출

사용:
  # 기본: 소스 xlsx → output/reward_by_event.json
  python scripts/scan_rewards_by_event.py

  # 커스텀 파일·출력 경로 지정
  python scripts/scan_rewards_by_event.py "output/이벤트기획_260611_260618.xlsx" output/reward_new_tabs.json

출력:
  - reward_by_event.json (또는 지정 경로)
    - event_type_patterns : 이벤트 유형별 역사적 보상 패턴 요약
    - per_tab_sections    : 탭별 섹션 단위 보상 상세
"""
import io
import json
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

# ─── 현재 프로젝트 설정 로드 ──────────────────────────────────────────────
_BASE_DIR = Path(__file__).resolve().parent.parent
_CURRENT_PROJECT_FILE = _BASE_DIR / "output" / "json" / "current_project.json"
_NC_KR_FALLBACK = str(_BASE_DIR / "Readdocs" / "[NC_KR] 라이브 이벤트 문서.xlsx")

def _resolve_source() -> str:
    """current_project.json 에서 source_xlsx 읽기. 없으면 NC_KR 레거시 경로 반환."""
    if _CURRENT_PROJECT_FILE.exists():
        try:
            cfg = json.loads(_CURRENT_PROJECT_FILE.read_text(encoding="utf-8"))
            p = cfg.get("source_xlsx", "")
            if p and Path(p).exists():
                return p
        except Exception:
            pass
    return _NC_KR_FALLBACK

SOURCE = _resolve_source()

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _project_config import load_project_paths

_paths = load_project_paths()

OUTPUT_DIR = Path("output")
OUTPUT_JSON_DIR = OUTPUT_DIR / "json"

# ─── 이벤트 섹션 패턴 ────────────────────────────────────────────────────────
# 두 가지 섹션 경계 형식 지원:
#   1) "숫자." 패턴  예: "1. 포인트 레이스 이벤트"
#   2) "이벤트 제목 :" 패턴  예: "이벤트 제목 : 올스타 직행! 14일 출석 이벤트!"
SECTION_PATTERN       = re.compile(r"^\d+\.")
EVENT_TITLE_PATTERN   = re.compile(r"^이벤트\s*제목\s*:\s*(.+)$")


def parse_section_title(val: str) -> "str | None":
    """B열 셀 값이 섹션 경계이면 정규화된 이벤트 제목을 반환, 아니면 None.

    - '숫자.' 패턴  → 값 그대로 반환  (예: '1. 포인트 레이스 이벤트')
    - '이벤트 제목 :' 패턴 → 콜론 이후 실제 제목만 추출
      (예: '이벤트 제목 : 올스타 직행! 14일 출석 이벤트!' → '올스타 직행! 14일 출석 이벤트!')
    """
    if SECTION_PATTERN.match(val):
        return val
    m = EVENT_TITLE_PATTERN.match(val)
    if m:
        return m.group(1).strip()
    return None

# ─── 이벤트 유형 정규화 (공통 + 장르별) ────────────────────────────────────
EVENT_TYPE_KEYWORDS = [
    # 공통
    ("출석",           "출석_이벤트"),
    ("미션",           "미션_이벤트"),
    ("패스",           "패스"),
    ("교환 상점",      "교환상점_이벤트"),
    ("교환소",         "교환상점_이벤트"),
    ("할인",           "할인_이벤트"),
    ("룰렛",           "룰렛_이벤트"),
    ("빙고",           "빙고_이벤트"),
    ("포인트 레이스",  "포인트레이스_이벤트"),
    ("응모권",         "응모권_이벤트"),
    ("승부 예측",      "승부예측_이벤트"),
    ("예측",           "승부예측_이벤트"),
    # 야구 스포츠
    ("야구공 찾기",    "탐색형_이벤트"),
    ("찾기",           "탐색형_이벤트"),
    ("플레이 미션",    "플레이미션_이벤트"),
    # NC_KR MMORPG
    ("던전",           "던전_이벤트"),
    ("레이드",         "레이드_이벤트"),
    ("보물 상자",      "보물상자_이벤트"),
    ("주사위",         "주사위_이벤트"),
    ("제작",           "제작_이벤트"),
    ("우편",           "우편지급_이벤트"),
    ("성장 가이드",    "성장가이드_이벤트"),
    ("지령",           "지령_이벤트"),
]

# ─── 보상 명칭 식별 키워드 (공통 + 장르별) ──────────────────────────────────
REWARD_TYPE_MAP = [
    # 야구 스포츠 공통
    ("선수 카드",      ["선수 카드", "선수카드"]),
    ("팩 티켓",        ["팩 티켓", "팩티켓"]),
    ("강화권",         ["강화권"]),
    ("행동력",         ["행동력"]),
    ("PP",             ["PP"]),
    ("XP",             ["XP"]),
    ("다이아",         ["다이아", "다이아몬드"]),
    ("골드",           ["골드"]),
    ("코인",           ["코인"]),
    ("상자",           ["상자", "박스"]),
    # NC_KR MMORPG
    ("원소 추출",      ["원소 추출"]),
    ("선택 상자",      ["선택 상자"]),
    ("주머니",         ["주머니"]),
    ("정수",           ["정수"]),
    ("파편",           ["파편"]),
    ("모리온",         ["모리온"]),
    ("강화 주문서",    ["강화 주문서"]),
    ("비약",           ["비약"]),
    ("기억",           ["기억"]),
    ("골드 상자",      ["골드 상자"]),
    ("축복",           ["축복"]),
    ("합금",           ["합금"]),
    ("룬",             ["룬"]),
    ("석판",           ["석판"]),
]

SKIP_PREFIXES = ("∎", "※", "·", "①", "②", "③", "④", "◆", "▶", "■", "○", "●", "—", "-")

# ─── 헤더 기반 보상 컬럼 탐지 ─────────────────────────────────────────────
# 앞으로 [보상 아이템] / [보상 수량] 컬럼명을 표준으로 사용.
# 현재 파일(보상 아이템 / 수량)도 하위 호환 지원.
REWARD_ITEM_HEADERS: frozenset = frozenset({"[보상 아이템]", "보상 아이템"})
REWARD_QTY_HEADERS:  frozenset = frozenset({"[보상 수량]",  "보상 수량", "수량"})

# ─── 수량 패턴 ───────────────────────────────────────────────────────────
QTY_PATTERNS = [
    re.compile(r'(\d[\d,]*)\s*(개|회|장|세트|EA|번)(?!\s*[%팩])', re.UNICODE),
    re.compile(r'(다이아|골드|코인|포인트)\s*(\d[\d,]+)(?!\s*[%팩])', re.UNICODE),
    re.compile(r'^(\d[\d,]+)$'),
]


# ─── 유틸 함수 ───────────────────────────────────────────────────────────

def is_date_tab(name: str) -> bool:
    return bool(re.match(r'^\d{6}$', str(name)))


def normalize_event_type(title: str) -> str:
    for keyword, etype in EVENT_TYPE_KEYWORDS:
        if keyword.lower() in title.lower():
            return etype
    return "기타"


def has_reward_keyword(text: str) -> bool:
    if not text:
        return False
    if any(text.startswith(p) for p in SKIP_PREFIXES):
        return False
    return any(kw in text for _, keywords in REWARD_TYPE_MAP for kw in keywords)


def classify_reward_type(text: str) -> str:
    for rtype, keywords in REWARD_TYPE_MAP:
        if any(kw in text for kw in keywords):
            return rtype
    return "기타"


def parse_quantity(text: str) -> dict | None:
    if not text:
        return None
    text = str(text).strip()

    m = QTY_PATTERNS[0].search(text)
    if m:
        return {
            "value": int(m.group(1).replace(',', '')),
            "unit": m.group(2),
            "raw": m.group(0).strip(),
            "confidence": "high",
        }

    m = QTY_PATTERNS[1].search(text)
    if m:
        return {
            "value": int(m.group(2).replace(',', '')),
            "unit": m.group(1),
            "raw": m.group(0).strip(),
            "confidence": "medium",
        }

    m = QTY_PATTERNS[2].match(text)
    if m:
        try:
            return {
                "value": int(m.group(1).replace(',', '')),
                "unit": "",
                "raw": m.group(1),
                "confidence": "low",
            }
        except ValueError:
            pass
    return None


# ─── 핵심 스캔 함수 ──────────────────────────────────────────────────────

def _scan_section_by_header(rows_data: dict, start_row: int, end_row: int) -> list:
    """
    [보상 아이템] / [보상 수량] 헤더 컬럼 기반 보상 수집.

    섹션 범위를 순회하면서 헤더 행을 찾고, 이후 데이터 행에서
    [보상 아이템] 컬럼 → 보상 이름, [보상 수량] 컬럼 → 수량을 읽는다.
    헤더가 한 섹션에 여러 번 등장할 수 있으므로 갱신 처리.

    반환: reward_rows 리스트 (헤더 없으면 빈 리스트)
    """
    result = []
    item_col: int | None = None
    qty_col:  int | None = None
    header_found = False

    for row_num in range(start_row, end_row + 1):
        cells = rows_data.get(row_num, {})
        if not cells:
            continue

        # 헤더 행 감지 (REWARD_ITEM_HEADERS 포함 여부)
        new_item_col: int | None = None
        new_qty_col:  int | None = None
        for col, (coord, val) in cells.items():
            v = val.strip()
            if v in REWARD_ITEM_HEADERS:
                new_item_col = col
            if v in REWARD_QTY_HEADERS:
                new_qty_col = col

        if new_item_col is not None:
            item_col   = new_item_col
            qty_col    = new_qty_col
            header_found = True
            continue  # 헤더 행 자체는 데이터 아님

        if item_col is None:
            continue  # 아직 헤더 미발견

        # 데이터 행: [보상 아이템] 컬럼에서 보상 이름 읽기
        item_cell = cells.get(item_col)
        if not item_cell:
            continue
        item_coord, item_name = item_cell
        item_name = item_name.strip()
        if not item_name:
            continue

        # [보상 수량] 컬럼에서 수량 읽기
        nearest: dict | None = None
        if qty_col is not None and qty_col in cells:
            qty_coord, qty_raw = cells[qty_col]
            # openpyxl이 숫자 셀을 '1000000.0' 형식으로 반환하는 경우 정수로 변환
            try:
                fval = float(qty_raw)
                if fval > 0 and fval == int(fval):
                    qty_raw = str(int(fval))
            except (ValueError, TypeError):
                pass
            qty_val = parse_quantity(qty_raw)
            if qty_val:
                # 헤더 기반 수량 컬럼은 명확히 지정된 값 → 항상 고신뢰도
                qty_val["confidence"] = "high"
                nearest = {
                    "cell": qty_coord,
                    "col": qty_col,
                    "value": qty_raw,
                    "quantity": qty_val,
                    "col_distance": abs(qty_col - item_col),
                }

        result.append({
            "cell":             item_coord,
            "col":              item_col,
            "row":              row_num,
            "reward_name":      item_name,
            "reward_type":      classify_reward_type(item_name),
            "quantity_in_cell": None,   # 헤더 기반: 수량은 별도 컬럼
            "nearest_quantity": nearest,
            "has_quantity":     nearest is not None,
            "scan_mode":        "header",
        })

    return result if header_found else []


def _scan_section_by_keyword(rows_data: dict, start_row: int, end_row: int) -> list:
    """
    키워드 기반 보상 수집 (헤더가 없는 섹션의 폴백).
    기존 has_reward_keyword() 로직 — REWARD_TYPE_MAP 키워드로 보상 셀 감지.
    """
    result = []
    for row_num in range(start_row, end_row + 1):
        cells = rows_data.get(row_num)
        if not cells:
            continue
        for col, (coord, val) in cells.items():
            if not has_reward_keyword(val):
                continue
            # 인접 셀에서 수량 탐색
            nearby: list = []
            for other_col, (other_coord, other_val) in cells.items():
                if other_col == col:
                    continue
                qty = parse_quantity(other_val)
                if qty and qty["confidence"] in ("high", "medium"):
                    nearby.append({
                        "cell": other_coord,
                        "col": other_col,
                        "value": other_val,
                        "quantity": qty,
                        "col_distance": abs(other_col - col),
                    })
            nearby.sort(key=lambda x: x["col_distance"])
            nearest = nearby[0] if nearby else None
            self_qty = parse_quantity(val) if re.search(r'\d', val) else None
            result.append({
                "cell":             coord,
                "col":              col,
                "row":              row_num,
                "reward_name":      val,
                "reward_type":      classify_reward_type(val),
                "quantity_in_cell": self_qty,
                "nearest_quantity": nearest,
                "has_quantity":     bool(self_qty or nearest),
                "scan_mode":        "keyword",
            })
    return result


def scan_tab_by_section(ws) -> list:
    """
    워크시트를 이벤트 섹션 단위로 스캔.
    B열(col=2)의 "숫자." 패턴으로 섹션 경계를 파악하고,
    각 섹션 범위 내 보상 셀을 수집한다.
    """
    # ① 전체 행 읽기: row_num → {col_int: (coord, value)}
    rows_data: dict[int, dict] = {}
    for row in ws.iter_rows():
        row_dict: dict[int, tuple] = {}
        row_num = None
        for cell in row:
            if row_num is None:
                row_num = cell.row
            if cell.value is not None:
                val = str(cell.value).strip()
                if val:
                    row_dict[cell.column] = (cell.coordinate, val)
        if row_dict and row_num is not None:
            rows_data[row_num] = row_dict

    if not rows_data:
        return []

    # ② 섹션 경계 탐색 (B열 = column 2)
    # '숫자.' 패턴 또는 '이벤트 제목 :' 패턴 모두 인식
    section_starts: list[tuple[int, str, str]] = []  # (row_num, coord, title)
    for row_num in sorted(rows_data.keys()):
        b_cell = rows_data[row_num].get(2)
        if not b_cell:
            continue
        title = parse_section_title(b_cell[1])
        if title is not None:
            section_starts.append((row_num, b_cell[0], title))

    if not section_starts:
        return []

    max_row = max(rows_data.keys())

    # ③ 각 섹션 범위에서 보상 셀 수집
    sections = []
    for i, (start_row, title_coord, title) in enumerate(section_starts):
        end_row = section_starts[i + 1][0] - 1 if (i + 1) < len(section_starts) else max_row

        # [보상 아이템]/[보상 수량] 헤더 기반 우선, 없으면 키워드 폴백
        reward_rows = _scan_section_by_header(rows_data, start_row + 1, end_row)
        if not reward_rows:
            reward_rows = _scan_section_by_keyword(rows_data, start_row + 1, end_row)

        sections.append({
            "title": title,
            "event_type": normalize_event_type(title),
            "title_cell": title_coord,
            "start_row": start_row,
            "end_row": end_row,
            "reward_count": len(reward_rows),
            "reward_rows": reward_rows,
        })

    return sections


def build_event_type_patterns(all_tab_sections: dict) -> dict:
    """
    이벤트 유형별 역사적 보상 패턴 집계.
    각 유형에서 자주 사용된 보상 명칭, 수량 범위·평균, 최근 구성 예시 포함.
    """
    type_data: dict[str, dict] = {}

    for tab, sections in all_tab_sections.items():
        for section in sections:
            etype = section["event_type"]
            if etype not in type_data:
                type_data[etype] = {
                    "seen_tabs": [],
                    "compositions": [],        # 탭별 실제 보상 구성
                    "reward_type_freq": {},    # 보상 유형 등장 횟수
                    "qty_samples": {},         # 보상 유형 → 수량 샘플 목록
                    "reward_name_freq": {},    # 보상 명칭 등장 횟수
                }

            entry = type_data[etype]
            entry["seen_tabs"].append(tab)

            composition = []
            for rr in section["reward_rows"]:
                rtype = rr["reward_type"]
                rname = rr["reward_name"]

                # 수량 결정: 셀 내 > 인접 셀
                qty = rr.get("quantity_in_cell") or (
                    rr["nearest_quantity"]["quantity"]
                    if rr.get("nearest_quantity") else None
                )

                composition.append({
                    "reward_name": rname,
                    "reward_type": rtype,
                    "quantity": qty,
                    "cell": rr["cell"],
                    "quantity_cell": rr["nearest_quantity"]["cell"] if rr.get("nearest_quantity") else None,
                })

                # 빈도 집계
                entry["reward_type_freq"][rtype] = entry["reward_type_freq"].get(rtype, 0) + 1
                entry["reward_name_freq"][rname] = entry["reward_name_freq"].get(rname, 0) + 1

                # 수량 샘플 (high/medium 신뢰도만)
                if qty and qty.get("confidence") in ("high", "medium"):
                    entry["qty_samples"].setdefault(rtype, []).append(qty["value"])

            if composition:
                entry["compositions"].append({
                    "tab": tab,
                    "event_title": section["title"],
                    "rewards": composition,
                })

    # 통계 요약
    patterns = {}
    for etype, data in type_data.items():
        qty_stats = {}
        for rtype, samples in data["qty_samples"].items():
            qty_stats[rtype] = {
                "min": min(samples),
                "max": max(samples),
                "avg": round(sum(samples) / len(samples)),
                "samples": len(samples),
            }

        # 가장 자주 등장한 보상 명칭 Top-5
        top_names = sorted(
            data["reward_name_freq"].items(), key=lambda x: -x[1]
        )[:5]

        patterns[etype] = {
            "seen_count": len(set(data["seen_tabs"])),
            "seen_tabs": sorted(set(data["seen_tabs"])),
            "top_reward_names": [n for n, _ in top_names],
            "reward_type_frequency": data["reward_type_freq"],
            "quantity_stats": qty_stats,
            "recent_compositions": data["compositions"][-3:],   # 최근 3탭 구성 예시
        }

    return patterns


def build_summarized_patterns(patterns: dict) -> dict:
    """
    --summarize 모드: event_type_patterns에서 median/range/trend만 추출.
    per_tab_sections(원문 행 데이터)를 제외해 토큰을 절약한다.
    """
    import statistics as _stat

    summary: dict[str, dict] = {}
    for etype, info in patterns.items():
        qty_summary: dict[str, dict] = {}
        for rtype, qs in info.get("quantity_stats", {}).items():
            samples_raw = qs.get("samples", 0)
            avg = qs.get("avg")
            mn  = qs.get("min")
            mx  = qs.get("max")
            # trend 판별: 최근 3개 구성 예시에서 수량 추이
            recent_comps = info.get("recent_compositions", [])
            recent_vals: list[int] = []
            for comp in recent_comps:
                for rw in comp.get("rewards", []):
                    if rw.get("reward_type") == rtype:
                        qty = rw.get("quantity")
                        if qty and isinstance(qty.get("value"), (int, float)):
                            recent_vals.append(int(qty["value"]))
            if len(recent_vals) >= 2 and avg:
                recent_avg = sum(recent_vals) / len(recent_vals)
                if recent_avg > avg * 1.05:
                    trend = "increasing"
                elif recent_avg < avg * 0.95:
                    trend = "decreasing"
                else:
                    trend = "stable"
            else:
                trend = "unknown"

            qty_summary[rtype] = {
                "median":  avg,   # build_event_type_patterns 가 avg를 산출
                "range":   [mn, mx],
                "samples": samples_raw,
                "trend":   trend,
            }

        summary[etype] = {
            "seen_tabs":       info.get("seen_count", 0),
            "top_reward_names": info.get("top_reward_names", []),
            "reward_summary":  qty_summary,
        }
    return summary


def main(source: str = SOURCE, out_path: str | None = None, summarize: bool = False):
    if _paths:
        _paths.ensure_dirs()
    OUTPUT_DIR.mkdir(exist_ok=True)
    OUTPUT_JSON_DIR.mkdir(exist_ok=True)

    if out_path is None:
        if _paths:
            out_path = str(_paths.reward_by_event)
        else:
            out_path = str(OUTPUT_JSON_DIR / "reward_by_event.json")

    print("[이벤트 섹션별 보상 스캔]")
    print(f"  소스: {source}")
    print(f"  출력: {out_path}")
    if summarize:
        print("  모드: 요약 (per_tab_sections 제외, 토큰 절약)")

    wb = openpyxl.load_workbook(source, data_only=True)
    date_tabs = [n for n in wb.sheetnames if is_date_tab(n)]
    print(f"  날짜형 탭: {date_tabs}")

    all_tab_sections: dict[str, list] = {}
    for tab in date_tabs:
        sections = scan_tab_by_section(wb[tab])
        all_tab_sections[tab] = sections
        reward_total = sum(s["reward_count"] for s in sections)
        print(f"  [{tab}] 섹션 {len(sections)}개, 보상 셀 {reward_total}개")

    patterns = build_event_type_patterns(all_tab_sections)

    if summarize:
        # 요약 모드: per_tab_sections 제외, 통계 요약만 저장
        result = {
            "scanned_at":              datetime.now().isoformat(timespec="seconds"),
            "source":                  source,
            "summary_mode":            True,
            "event_type_patterns":     patterns,          # 유형별 패턴 (recent_compositions 포함)
            "event_type_summary":      build_summarized_patterns(patterns),  # median/range/trend
            # per_tab_sections 는 의도적으로 제외
        }
    else:
        result = {
            "scanned_at":          datetime.now().isoformat(timespec="seconds"),
            "source":              source,
            "event_type_patterns": patterns,
            "per_tab_sections":    all_tab_sections,
        }

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"\n[이벤트 유형별 보상 패턴 요약]")
    for etype, info in patterns.items():
        qty_str = ", ".join(
            f"{rt} 평균 {s['avg']}개 ({s['samples']}샘플)"
            for rt, s in info["quantity_stats"].items()
        )
        if not qty_str:
            qty_str = "수량 데이터 없음 (팩형)"
        tabs_str = f"{info['seen_count']}개 탭"
        print(f"  {etype:20s} | {tabs_str} | {qty_str}")

    print(f"\n저장: {out_path}")


if __name__ == "__main__":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

    import argparse as _ap
    _parser = _ap.ArgumentParser(description="이벤트 섹션별 보상 스캔")
    _parser.add_argument("source", nargs="?", default=None)
    _parser.add_argument("out_path", nargs="?", default=None)
    _parser.add_argument("--summarize", action="store_true",
                         help="요약 통계만 저장 (per_tab_sections 제외, 토큰 절약)")
    _args = _parser.parse_args()

    _src = _args.source or SOURCE
    if _args.out_path:
        _out = _args.out_path
    elif _paths:
        _out = str(_paths.reward_by_event)
    else:
        _out = None
    main(_src, _out, summarize=_args.summarize)
