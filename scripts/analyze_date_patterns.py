#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
날짜별 이벤트 패턴 분석기

기존 xlsx / Google Sheets의 모든 날짜형 탭을 스캔해
  - 날짜 × 이벤트 유형 매트릭스 생성
  - 주기(7일/14일/월간) 패턴 감지
  - 누락 이벤트 (역사적으로 해당 날짜에 있어야 하는데 없는 것) 알림
  - 신규 이벤트 (역사에 없던 이벤트가 처음 등장) 알림
  - 이벤트 연속 등장 기간(streak) 분석

출력:
  output/projects/event-planner/work/date_pattern_analysis.json

사용:
  python scripts/analyze_date_patterns.py
  python scripts/analyze_date_patterns.py "path/to/source.xlsx"
  python scripts/analyze_date_patterns.py "https://docs.google.com/spreadsheets/d/..."
"""
import io
import json
import re
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from statistics import median

import openpyxl

# ─── 경로 설정 ────────────────────────────────────────────────────────────────
_BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _project_config import load_project_paths

_paths = load_project_paths()

# ─── 이벤트 유형 키워드 매핑 ─────────────────────────────────────────────────
EVENT_TYPE_KEYWORDS = [
    ("출석",          "출석이벤트"),
    ("미션",          "미션이벤트"),
    ("패스",          "패스"),
    ("교환 상점",     "교환상점"),
    ("교환소",        "교환상점"),
    ("할인",          "할인이벤트"),
    ("룰렛",          "룰렛이벤트"),
    ("빙고",          "빙고이벤트"),
    ("포인트 레이스", "포인트레이스"),
    ("응모권",        "응모권이벤트"),
    ("승부 예측",     "승부예측"),
    ("예측",          "승부예측"),
    ("찾기",          "탐색이벤트"),
    ("플레이 미션",   "플레이미션"),
    ("던전",          "던전이벤트"),
    ("레이드",        "레이드이벤트"),
    ("보물 상자",     "보물상자이벤트"),
    ("주사위",        "주사위이벤트"),
    ("제작",          "제작이벤트"),
    ("우편",          "우편지급"),
    ("성장 가이드",   "성장가이드"),
    ("지령",          "지령이벤트"),
    ("포토",          "포토이벤트"),
    ("스탬프",        "스탬프이벤트"),
    ("뽑기",          "뽑기이벤트"),
    ("로그인",        "로그인이벤트"),
    # 스포츠 게임 특화
    ("PvP",           "PvP이벤트"),
    ("pvp",           "PvP이벤트"),
    ("핫타임",        "핫타임이벤트"),
    ("오더 경쟁",     "오더경쟁이벤트"),
    ("오더경쟁",      "오더경쟁이벤트"),
    ("쿠폰",          "쿠폰이벤트"),
    ("개막",          "시즌이벤트"),
    ("폐막",          "시즌이벤트"),
    ("올스타",        "올스타이벤트"),
    ("포스트시즌",    "포스트시즌이벤트"),
    ("드래프트",      "드래프트이벤트"),
    ("시즌",          "시즌이벤트"),
    ("챌린지",        "챌린지이벤트"),
    ("랭킹",          "랭킹이벤트"),
    ("선물",          "선물이벤트"),
    ("보상",          "보상이벤트"),
]

# 탭당 최소 이벤트 수 기준
MIN_EVENTS_WARN = 3

# ─── 유틸 ─────────────────────────────────────────────────────────────────────

def is_date_tab(name: str) -> bool:
    return bool(re.fullmatch(r"\d{6}", str(name)))

def tab_to_date(tab: str) -> datetime | None:
    try:
        return datetime.strptime("20" + tab, "%Y%m%d")
    except Exception:
        return None

def normalize_event_type(title: str) -> str:
    for kw, etype in EVENT_TYPE_KEYWORDS:
        if kw.lower() in title.lower():
            return etype
    return "기타"

SECTION_PATTERN = re.compile(r"^\d+\.")
TITLE_PATTERN   = re.compile(r"^이벤트\s*제목\s*:\s*(.+)$")
# 탭 헤더 행 패턴 — 실제 이벤트 섹션이 아닌 것 제외
#   예: "04.16_ Event", "06.11_ Event", "2026.04.16 Event" 등
TAB_HEADER_PATTERN = re.compile(r"^\d{1,2}[./-]\d{1,2}[._\s]")

def extract_sections(ws) -> list[dict]:
    """워크시트 B열에서 이벤트 섹션 추출."""
    sections = []
    for row in ws.iter_rows():
        for cell in row:
            col = getattr(cell, "column", None)
            if col != 2:
                continue
            val = getattr(cell, "value", None)
            if not val:
                continue
            s = str(val).strip()
            # 탭 헤더 패턴(날짜_이름) 제외
            if TAB_HEADER_PATTERN.match(s):
                continue
            title = None
            if SECTION_PATTERN.match(s):
                title = s
            else:
                m = TITLE_PATTERN.match(s)
                if m:
                    title = m.group(1).strip()
            if title:
                sections.append({
                    "title": title,
                    "event_type": normalize_event_type(title),
                    "row": getattr(cell, "row", None),
                })
    return sections


# ─── 핵심 분석 ───────────────────────────────────────────────────────────────

def build_date_matrix(wb) -> tuple[dict[str, list[str]], dict[str, list[dict]]]:
    """
    날짜탭 → 등장 이벤트 유형 리스트 매핑.
    Returns: (matrix, raw_sections_by_tab)
    """
    matrix: dict[str, list[str]] = {}
    raw: dict[str, list[dict]] = {}
    for tab in wb.sheetnames:
        if not is_date_tab(tab):
            continue
        try:
            secs = extract_sections(wb[tab])
            etypes = list(dict.fromkeys(s["event_type"] for s in secs))  # 순서 유지, 중복 제거
            matrix[tab] = etypes
            raw[tab] = secs
        except Exception as e:
            print(f"  [{tab}] 읽기 오류: {e} — 건너뜀", file=sys.stderr)
    return dict(sorted(matrix.items())), dict(sorted(raw.items()))


def frequency_map(matrix: dict[str, list[str]]) -> dict[str, dict]:
    """이벤트 유형별 전체 등장률·탭 목록."""
    total = len(matrix)
    cnt: dict[str, int] = defaultdict(int)
    tabs_by_type: dict[str, list[str]] = defaultdict(list)
    for tab, etypes in matrix.items():
        for et in etypes:
            cnt[et] += 1
            tabs_by_type[et].append(tab)
    result = {}
    for et, c in cnt.items():
        rate = c / total if total else 0
        if rate >= 0.80:
            priority = "anchor"
        elif rate >= 0.50:
            priority = "common"
        elif rate >= 0.30:
            priority = "optional"
        else:
            priority = "rare"
        result[et] = {
            "count": c,
            "total": total,
            "rate": round(rate, 3),
            "rate_pct": f"{rate*100:.0f}%",
            "priority": priority,
            "tabs": sorted(tabs_by_type[et]),
        }
    return result


def detect_interval(sorted_tabs: list[str]) -> int:
    """탭 날짜 간격 중앙값(일) 반환."""
    dates = [tab_to_date(t) for t in sorted_tabs]
    dates = [d for d in dates if d]
    if len(dates) < 2:
        return 7
    gaps = [(dates[i+1]-dates[i]).days for i in range(len(dates)-1)
            if 3 <= (dates[i+1]-dates[i]).days <= 35]
    if not gaps:
        return 7
    return int(median(gaps))


def detect_missing_and_new(
    matrix: dict[str, list[str]],
    freq: dict[str, dict],
    interval_days: int,
) -> tuple[list[dict], list[dict]]:
    """
    누락 이벤트: anchor/common 이벤트가 연속 2+ 탭 이상 등장하지 않은 경우.
    신규 이벤트: 첫 등장 이후 3탭 미만 등장한 이벤트.
    반환: (missing_alerts, new_alerts)
    """
    tabs = sorted(matrix.keys())
    missing_alerts: list[dict] = []
    new_alerts: list[dict] = []

    for et, info in freq.items():
        if info["priority"] not in ("anchor", "common"):
            continue
        # 누락 감지: 연속 부재 탭 구간
        absence_start = None
        max_absence = 0
        cur_absence = 0
        absence_periods: list[dict] = []

        for tab in tabs:
            if et in matrix[tab]:
                if absence_start and cur_absence >= 2:
                    absence_periods.append({
                        "from_tab": absence_start,
                        "to_tab": tab,
                        "count": cur_absence,
                        "days": cur_absence * interval_days,
                    })
                    if cur_absence > max_absence:
                        max_absence = cur_absence
                cur_absence = 0
                absence_start = None
            else:
                if absence_start is None:
                    absence_start = tab
                cur_absence += 1

        # 마지막 구간도 체크 (최근까지 계속 없는 경우)
        if absence_start and cur_absence >= 2:
            absence_periods.append({
                "from_tab": absence_start,
                "to_tab": "(현재)",
                "count": cur_absence,
                "days": cur_absence * interval_days,
            })
            if cur_absence > max_absence:
                max_absence = cur_absence

        if absence_periods:
            worst = max(absence_periods, key=lambda x: x["count"])
            missing_alerts.append({
                "event_type": et,
                "priority": info["priority"],
                "rate_pct": info["rate_pct"],
                "absence_periods": absence_periods,
                "longest_absence_tabs": worst["count"],
                "longest_absence_days": worst["days"],
                "severity": "high" if info["priority"] == "anchor" else "medium",
            })

    # 신규 이벤트 감지: 전체 탭 중 등장 비율 < 20% 이면서 최근 2탭에 등장
    recent_tabs = tabs[-3:] if len(tabs) >= 3 else tabs
    for et, info in freq.items():
        if info["rate"] < 0.20 and et != "기타":
            # 최근 탭에 등장하는지 확인
            recent_appearances = [t for t in recent_tabs if et in matrix.get(t, [])]
            if recent_appearances:
                first_tab = info["tabs"][0] if info["tabs"] else None
                new_alerts.append({
                    "event_type": et,
                    "first_seen": first_tab,
                    "count": info["count"],
                    "total": info["total"],
                    "rate_pct": info["rate_pct"],
                    "recent_tabs": recent_appearances,
                })

    missing_alerts.sort(key=lambda x: (0 if x["severity"] == "high" else 1))
    return missing_alerts, new_alerts


def build_streak_map(matrix: dict[str, list[str]]) -> dict[str, dict]:
    """이벤트 유형별 연속 등장 / 비등장 스트릭 분석."""
    tabs = sorted(matrix.keys())
    result: dict[str, dict] = {}

    all_types = {et for etypes in matrix.values() for et in etypes if et != "기타"}
    for et in all_types:
        presence = [et in matrix[t] for t in tabs]
        # 현재(마지막) 스트릭
        cur_streak = 0
        cur_state = presence[-1] if presence else False
        for p in reversed(presence):
            if p == cur_state:
                cur_streak += 1
            else:
                break
        # 최장 연속 등장
        max_on = max_cur = 0
        for p in presence:
            if p:
                max_cur += 1
                max_on = max(max_on, max_cur)
            else:
                max_cur = 0

        result[et] = {
            "current_streak": cur_streak,
            "current_state": "등장 중" if cur_state else "비등장 중",
            "max_consecutive_on": max_on,
            "last_seen": ([t for t in reversed(tabs) if et in matrix.get(t, [])] or [None])[0],
        }
    return result


def monthly_distribution(matrix: dict[str, list[str]]) -> dict[str, dict]:
    """월별 이벤트 유형 등장 현황."""
    monthly: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    tabs_per_month: dict[str, int] = defaultdict(int)
    for tab, etypes in matrix.items():
        dt = tab_to_date(tab)
        if not dt:
            continue
        ym = dt.strftime("%Y-%m")
        tabs_per_month[ym] += 1
        for et in etypes:
            monthly[ym][et] += 1
    return {
        ym: {
            "tab_count": tabs_per_month[ym],
            "events": dict(sorted(monthly[ym].items(), key=lambda x: -x[1])),
        }
        for ym in sorted(monthly.keys())
    }


# ─── 메인 ────────────────────────────────────────────────────────────────────

def resolve_source(url_or_path: str) -> str:
    """Google Sheets URL → xlsx 다운로드. 로컬 경로는 그대로 반환."""
    from gdrive_utils import is_google_url, resolve_to_local_file
    if is_google_url(url_or_path):
        cache_dir = _BASE_DIR / "output" / "gdrive_cache"
        cache_dir.mkdir(parents=True, exist_ok=True)
        local = resolve_to_local_file(url_or_path, dest_dir=cache_dir)
        return str(local)
    return url_or_path


def analyze(source: str) -> dict:
    """전체 분석 실행. source는 로컬 xlsx 경로."""
    if not Path(source).exists():
        return {"error": f"파일 없음: {source}", "source": source}

    wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
    matrix, raw_sections = build_date_matrix(wb)
    wb.close()

    # "기타" 이벤트의 실제 제목 목록 추출 (분류 개선용)
    unclassified_titles: dict[str, list[str]] = defaultdict(list)
    for tab, secs in raw_sections.items():
        for sec in secs:
            if sec["event_type"] == "기타":
                t = sec["title"]
                if t not in unclassified_titles[tab]:
                    unclassified_titles[tab].append(t)

    if not matrix:
        return {"error": "날짜형 탭이 없습니다.", "source": source}

    tabs = sorted(matrix.keys())
    interval = detect_interval(tabs)
    freq = frequency_map(matrix)
    missing_alerts, new_alerts = detect_missing_and_new(matrix, freq, interval)
    streaks = build_streak_map(matrix)
    monthly = monthly_distribution(matrix)

    # 탭당 이벤트 수 통계
    counts = [len(v) for v in matrix.values()]
    avg_count = round(sum(counts) / len(counts), 1) if counts else 0

    return {
        "analyzed_at": datetime.now().isoformat(timespec="seconds"),
        "source": source,
        "total_tabs": len(tabs),
        "tab_date_range": {
            "first": tabs[0],
            "last": tabs[-1],
        },
        "interval_days": interval,
        "avg_events_per_tab": avg_count,
        "min_events_per_tab": min(counts) if counts else 0,
        "max_events_per_tab": max(counts) if counts else 0,
        # 날짜 × 이벤트 매트릭스 (프론트 시각화용)
        "date_matrix": {tab: etypes for tab, etypes in matrix.items()},
        # "기타" 이벤트 실제 제목 (분류 확인용)
        "unclassified_titles": {tab: titles for tab, titles in unclassified_titles.items() if titles},
        # 전체 이벤트 유형 목록 (매트릭스 컬럼 헤더)
        "all_event_types": sorted(
            {et for etypes in matrix.values() for et in etypes},
            key=lambda et: -freq.get(et, {}).get("rate", 0),
        ),
        # 이벤트 유형별 빈도 통계
        "frequency": freq,
        # ⚠ 누락 알림
        "missing_alerts": missing_alerts,
        # ★ 신규 이벤트 알림
        "new_event_alerts": new_alerts,
        # 스트릭 분석
        "streaks": streaks,
        # 월별 분포
        "monthly_distribution": monthly,
    }


def main():
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

    raw_source = sys.argv[1] if len(sys.argv) > 1 else None

    # 소스 경로 결정
    if raw_source:
        source = resolve_source(raw_source) if raw_source.startswith("http") else raw_source
    else:
        # current_project.json 에서 읽기
        cp = _BASE_DIR / "output" / "json" / "current_project.json"
        if cp.exists():
            try:
                cfg = json.loads(cp.read_text(encoding="utf-8"))
                source = cfg.get("source_xlsx", "")
            except Exception:
                source = ""
        else:
            source = ""

    if not source or not Path(source).exists():
        print(f"[오류] 소스 파일을 찾을 수 없습니다: {source}", file=sys.stderr)
        print("사용법: python scripts/analyze_date_patterns.py <xlsx경로 또는 Sheets URL>",
              file=sys.stderr)
        sys.exit(1)

    print(f"[날짜별 이벤트 패턴 분석]")
    print(f"  소스: {source}")

    result = analyze(source)

    if "error" in result:
        print(f"[오류] {result['error']}", file=sys.stderr)
        sys.exit(1)

    # 출력 경로
    if _paths:
        _paths.ensure_dirs()
        out_path = _paths.work_dir / "date_pattern_analysis.json"
    else:
        out_path = _BASE_DIR / "output" / "projects" / "event-planner" / "work" / "date_pattern_analysis.json"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")

    # 콘솔 요약
    print(f"\n[분석 완료] 탭 {result['total_tabs']}개 / 탭 간격 {result['interval_days']}일")
    print(f"  탭당 평균 이벤트: {result['avg_events_per_tab']}개 "
          f"(최소 {result['min_events_per_tab']} ~ 최대 {result['max_events_per_tab']})")

    print(f"\n[이벤트 유형별 등장률] (상위 15개)")
    freq_sorted = sorted(result["frequency"].items(), key=lambda x: -x[1]["rate"])[:15]
    PRIORITY_ICON = {"anchor": "❗", "common": "⚠", "optional": "〇", "rare": "  "}
    for et, info in freq_sorted:
        bar = "█" * int(info["rate"] * 10)
        icon = PRIORITY_ICON.get(info["priority"], " ")
        print(f"  {icon} {et:<22} {bar:<10} {info['rate_pct']:>4}  [{info['count']}/{info['total']}]")

    print(f"\n[⚠ 누락 알림] ({len(result['missing_alerts'])}건)")
    for alert in result["missing_alerts"]:
        worst = alert["absence_periods"][-1] if alert["absence_periods"] else {}
        print(f"  [{alert['severity'].upper()}] {alert['event_type']} — "
              f"최장 {alert['longest_absence_tabs']}탭 연속 미등장 "
              f"({alert['longest_absence_days']}일 추정)")

    print(f"\n[★ 신규 이벤트 알림] ({len(result['new_event_alerts'])}건)")
    for alert in result["new_event_alerts"]:
        print(f"  ★ {alert['event_type']} — 첫 등장 {alert['first_seen']}, "
              f"등장 {alert['count']}회/{alert['total']}탭")

    print(f"\n저장: {out_path}")
    return result


if __name__ == "__main__":
    main()
