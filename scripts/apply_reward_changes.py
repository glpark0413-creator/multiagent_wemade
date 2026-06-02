#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
승인된 보상 변경 사항을 xlsx 에 반영하는 스크립트

사용 방법:
  python scripts/apply_reward_changes.py --xlsx "{xlsx_path}" --changes "{changes_json_path}"

changes_json 형식:
  {
    "260625": [
      {
        "section_title": "올스타 직행! 14일 출석 이벤트!",
        "reward_name":   "다이아",
        "qty_cell":      "AR18",
        "new_qty":       1000,
        "old_qty":       500
      }
    ]
  }

  또는 에이전트가 직접 인라인으로 호출:
  apply_changes(xlsx_path, changes_dict)
"""
import io
import json
import sys
from pathlib import Path

import openpyxl

_BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _project_config import load_project_paths

_paths = load_project_paths()


def apply_changes(xlsx_path: str | Path, changes: dict[str, list]) -> dict:
    """
    xlsx_path  : 수정할 Excel 파일 경로
    changes    : { 탭명: [ {qty_cell, new_qty, reward_name, ...}, ... ] }

    반환: { "applied": [ {tab, cell, reward_name, old, new} ], "errors": [...] }
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        return {"applied": [], "errors": [f"파일 없음: {xlsx_path}"]}

    wb = openpyxl.load_workbook(str(xlsx_path))
    applied = []
    errors  = []

    for tab_name, rows in changes.items():
        if tab_name not in wb.sheetnames:
            errors.append(f"탭 없음: {tab_name}")
            continue
        ws = wb[tab_name]

        for change in rows:
            qty_cell   = change.get("qty_cell")
            new_qty    = change.get("new_qty")
            reward_name = change.get("reward_name", "?")
            old_qty    = change.get("old_qty")

            # 보상 명칭 변경
            item_cell  = change.get("item_cell")
            new_name   = change.get("new_name")

            if qty_cell and new_qty is not None:
                try:
                    cell = ws[qty_cell]
                    old_val = cell.value
                    cell.value = new_qty
                    applied.append({
                        "tab":         tab_name,
                        "type":        "수량 변경",
                        "cell":        qty_cell,
                        "reward_name": reward_name,
                        "old":         old_val,
                        "new":         new_qty,
                    })
                    print(f"  [{tab_name}] {qty_cell} ({reward_name}): {old_val} → {new_qty}")
                except Exception as e:
                    errors.append(f"셀 수정 오류 [{tab_name}] {qty_cell}: {e}")

            if item_cell and new_name:
                try:
                    cell = ws[item_cell]
                    old_val = cell.value
                    cell.value = new_name
                    applied.append({
                        "tab":      tab_name,
                        "type":     "명칭 변경",
                        "cell":     item_cell,
                        "old":      old_val,
                        "new":      new_name,
                    })
                    print(f"  [{tab_name}] {item_cell} (명칭): '{old_val}' → '{new_name}'")
                except Exception as e:
                    errors.append(f"명칭 수정 오류 [{tab_name}] {item_cell}: {e}")

    if applied:
        wb.save(str(xlsx_path))
        print(f"\n  저장: {xlsx_path}  ({len(applied)}건 적용)")
    else:
        print("  변경 사항 없음 — xlsx 저장 생략")

    return {"applied": applied, "errors": errors}


def parse_agent_response(response: str, rec_data: dict, tab: str) -> list[dict]:
    """
    에이전트 응답 문자열 → changes 리스트 변환.

    ── 수량 변경 ──
      "권장"                      → 상향_권장 / 하향_검토 항목 전부
      "전체 승인"                  → 모든 비-유지 항목 (수량 변경만)
      "건너뜀"                     → 빈 리스트
      "1,3,5"                     → 해당 보상 번호의 권장 수량으로 변경
      "3번 다이아 500→1000"         → 보상 번호 3, 다이아, 수량 1000으로 직접 지정

    ── 아이템 명칭 변경 ──
      "이름 변경"                  → 명칭_검토 / 명칭_변경_권장 항목 전부 suggested_name 으로 적용
      "이름 3번"                   → 보상 번호 3번의 suggested_name 적용
      "이름 3번: 새이름"            → 보상 번호 3번을 '새이름'으로 직접 지정
      "아이템명: 구이름 → 새이름"    → 구이름과 일치하는 보상의 명칭 변경
    """
    import re as _re
    response = response.strip()
    sections = rec_data.get("tabs", {}).get(tab, [])
    # reward_review_queue 형식도 지원
    queue_rewards = None
    if "queue" in rec_data:
        for q in rec_data["queue"]:
            if q.get("tab") == tab:
                queue_rewards = q.get("rewards", [])
                break

    # 건너뜀
    if response in ("건너뜀", "skip", "스킵"):
        return []

    changes = []

    # ── 내부 헬퍼: sections 형식(recommend_rewards.json) 에서 수량 변경 추가 ──
    def _add_qty_changes(secs):
        for sec in secs:
            for r in sec.get("rewards", []):
                action = r.get("recommendation", {}).get("action", "")
                if action in ("상향_권장", "하향_검토"):
                    sq = r.get("recommendation", {}).get("suggested_qty")
                    if sq is not None and r.get("qty_cell"):
                        changes.append({
                            "section_title": sec.get("event_title", ""),
                            "reward_name":   r.get("reward_name", ""),
                            "qty_cell":      r["qty_cell"],
                            "new_qty":       sq,
                            "old_qty":       r.get("current_qty"),
                        })

    # ── 내부 헬퍼: 명칭 변경 (sections / queue 양쪽 지원) ──
    def _add_name_change(item_cell, reward_name, current_name, new_name, section_title=""):
        if item_cell and new_name and new_name != current_name:
            changes.append({
                "section_title": section_title,
                "reward_name":   reward_name,
                "item_cell":     item_cell,
                "new_name":      new_name,
                "old_name":      current_name,
            })

    def _resolve_rewards(sections):
        """sections 형식의 보상 목록 → 플랫 리스트."""
        flat = []
        for sec in sections:
            for r in sec.get("rewards", []):
                flat.append((sec, r))
        return flat

    # ══ 명칭 일괄 변경: "이름 변경" or "명칭 변경" ══════════════════════════
    if _re.match(r'^(이름|명칭)\s*(변경|추천|권장)$', response):
        # sections 또는 queue_rewards 중 available 한 소스 사용
        source = queue_rewards or []
        if not source and sections:
            for sec, r in _resolve_rewards(sections):
                action = r.get("recommendation", {}).get("action", "")
                if action in ("명칭_검토", "명칭_변경_권장"):
                    sn = r.get("recommendation", {}).get("suggested_name")
                    if sn:
                        _add_name_change(r.get("item_cell"), r.get("reward_name", ""),
                                         r.get("reward_name", ""), sn, sec.get("event_title", ""))
        else:
            for r in source:
                action = r.get("action", "")
                if action in ("명칭_검토", "명칭_변경_권장"):
                    sn = r.get("suggested_name")
                    if sn:
                        _add_name_change(r.get("item_cell"), r.get("reward_name", ""),
                                         r.get("reward_name", ""), sn)
        return changes

    # ══ 수량 권장 일괄 ════════════════════════════════════════════════════════
    if response in ("권장", "추천"):
        _add_qty_changes(sections)
        return changes

    # ══ 전체 승인 (수량만) ════════════════════════════════════════════════════
    if response in ("전체 승인", "전체승인", "전체"):
        _add_qty_changes(sections)
        return changes

    # ══ "이름 N번" — 번호 지정 명칭 변경 ════════════════════════════════════
    name_num = _re.match(r'^(이름|명칭)\s+(\d+)번\s*(?:[:：]\s*(.+))?$', response)
    if name_num:
        _, num_str, override_name = name_num.groups()
        num = int(num_str)
        src = queue_rewards or []
        if src and 1 <= num <= len(src):
            r = src[num - 1]
            new_name = override_name.strip() if override_name else r.get("suggested_name")
            if new_name:
                _add_name_change(r.get("item_cell"), r.get("reward_name", ""),
                                 r.get("reward_name", ""), new_name)
        return changes

    # ══ "아이템명: 구이름 → 새이름" ══════════════════════════════════════════
    inline_name = _re.search(r'아이템명\s*[:：]\s*(.+?)\s*[→\-]+\s*(.+)', response)
    if inline_name:
        old_name_kw, new_name = inline_name.group(1).strip(), inline_name.group(2).strip()
        src = queue_rewards or []
        for r in src:
            if old_name_kw.lower() in r.get("reward_name", "").lower():
                _add_name_change(r.get("item_cell"), r.get("reward_name", ""),
                                 r.get("reward_name", ""), new_name)
        if not src and sections:
            for sec, r in _resolve_rewards(sections):
                if old_name_kw.lower() in r.get("reward_name", "").lower():
                    _add_name_change(r.get("item_cell"), r.get("reward_name", ""),
                                     r.get("reward_name", ""), new_name, sec.get("event_title", ""))
        return changes

    # ══ 번호 선택 (수량): "1,3,5" ════════════════════════════════════════════
    nums_only = _re.match(r'^[\d,\s]+$', response)
    if nums_only and sections:
        selected = [int(n) for n in _re.findall(r'\d+', response)]
        for n in selected:
            if 1 <= n <= len(sections):
                sec = sections[n - 1]
                for r in sec.get("rewards", []):
                    action = r.get("recommendation", {}).get("action", "")
                    if action in ("상향_권장", "하향_검토"):
                        sq = r.get("recommendation", {}).get("suggested_qty")
                        if sq is not None and r.get("qty_cell"):
                            changes.append({
                                "section_title": sec.get("event_title", ""),
                                "reward_name":   r.get("reward_name", ""),
                                "qty_cell":      r["qty_cell"],
                                "new_qty":       sq,
                                "old_qty":       r.get("current_qty"),
                            })
        return changes

    # ══ 직접 수정: "3번 다이아 500→1000" 또는 "3번 다이아=1000" ═══════════
    direct = _re.findall(r'(\d+)번\s+([^\s,→=]+)\s*[→=]\s*(\d[\d,]*)', response)
    for sec_num_str, rname_kw, new_val_str in direct:
        sec_num = int(sec_num_str)
        new_val = int(new_val_str.replace(',', ''))
        if sections and 1 <= sec_num <= len(sections):
            sec = sections[sec_num - 1]
            for r in sec.get("rewards", []):
                if rname_kw.lower() in r.get("reward_name", "").lower():
                    if r.get("qty_cell"):
                        changes.append({
                            "section_title": sec.get("event_title", ""),
                            "reward_name":   r.get("reward_name", ""),
                            "qty_cell":      r["qty_cell"],
                            "new_qty":       new_val,
                            "old_qty":       r.get("current_qty"),
                        })
    return changes


# ─── CLI 진입점 ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx",    required=True, help="수정할 xlsx 파일 경로")
    parser.add_argument("--changes", required=True, help="변경 사항 JSON 파일 경로")
    args = parser.parse_args()

    changes_path = Path(args.changes)
    if not changes_path.exists():
        print(f"[오류] 변경 파일 없음: {changes_path}")
        sys.exit(1)

    changes_data = json.loads(changes_path.read_text(encoding="utf-8"))
    result = apply_changes(args.xlsx, changes_data)

    if result["errors"]:
        print("\n[오류 목록]")
        for e in result["errors"]:
            print(f"  - {e}")

    print(f"\n적용 완료: {len(result['applied'])}건, 오류: {len(result['errors'])}건")
