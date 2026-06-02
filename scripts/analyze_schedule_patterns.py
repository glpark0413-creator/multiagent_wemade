#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
이벤트 일정 패턴 분석기

기존 xlsx의 날짜형 탭 전체를 분석해 이벤트 일정 배치 규칙을 학습하고
schedule_patterns.json 으로 저장한다.

분석 항목:
  - 탭당 평균 이벤트 수
  - 이벤트 유형별 등장률 (anchor 이벤트 식별)
  - 탭 간격 패턴 (7일, 14일 주기)
  - 이벤트 유형 순서 패턴 (섹션 내 위치)
  - 이벤트 유형 동시 등장 규칙
  - 월별 특이사항 (골든위크, 연말 등)
  - 이벤트 유형별 평균 섹션 비중

사용:
  python scripts/analyze_schedule_patterns.py
  python scripts/analyze_schedule_patterns.py "path/to/source.xlsx" \\
      --output output/schedule_patterns.json
"""
import argparse
import importlib.util
import io
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

# ─── 경로 설정 ────────────────────────────────────────────────────────────────
_BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _project_config import load_project_paths

_paths = load_project_paths()

# scan_rewards_by_event 의 섹션 파싱 로직 재활용
_SCAN_MOD = Path(__file__).resolve().parent / "scan_rewards_by_event.py"
_spec = importlib.util.spec_from_file_location("_srbe", str(_SCAN_MOD))
_srbe = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_srbe)

# ─── 상수 ─────────────────────────────────────────────────────────────────────
ANCHOR_RATE   = 0.80   # 이 비율 이상 등장 → anchor 이벤트
COMMON_RATE   = 0.50   # 이 비율 이상 등장 → 일반 이벤트
OPTIONAL_RATE = 0.30   # 이 비율 이상 등장 → 선택 이벤트

# 월별 시즌 노트 (고정 규칙)
MONTH_SEASON_NOTES: dict[str, str] = {
    "01": "신년·설날 이벤트 필수 (1~3일 출석 이벤트 강화)",
    "02": "발렌타인 이벤트 선택",
    "03": "봄 시즌 이벤트 (봄의 XX 문구 권장)",
    "04": "봄 시즌 이벤트 지속",
    "05": "황금연휴(일본 골든위크) 집중 배치 — 5/3~5/6 기간 이벤트 강화",
    "06": "여름 시즌 이벤트 시작 (얼리썸머)",
    "07": "여름 시즌 이벤트 피크",
    "08": "여름 시즌 이벤트 마무리",
    "09": "추석·가을 시즌 이벤트",
    "10": "핼러윈 이벤트 선택 (하순)",
    "11": "연말 준비 이벤트",
    "12": "크리스마스·연말 이벤트 필수",
}


# ─── 유틸 함수 ────────────────────────────────────────────────────────────────

def tab_to_date(tab_name: str) -> datetime | None:
    if not re.fullmatch(r"\d{6}", tab_name):
        return None
    try:
        return datetime.strptime("20" + tab_name, "%Y%m%d")
    except ValueError:
        return None


def scan_date_cells(ws) -> list[dict]:
    """
    워크시트에서 날짜 관련 셀을 추출한다.
    openpyxl datetime 객체 또는 'MM/DD', 'M/D' 텍스트 패턴을 인식.
    """
    date_re = re.compile(r"(\d{1,2})[/\.](\d{1,2})")
    dates_found = []
    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if val is None:
                continue
            # openpyxl이 datetime으로 파싱한 경우
            if hasattr(val, "year") and hasattr(val, "month") and hasattr(val, "day"):
                try:
                    dates_found.append({
                        "cell": cell.coordinate,
                        "date": datetime(val.year, val.month, val.day),
                        "source": "datetime",
                    })
                except Exception:
                    pass
                continue
            # 텍스트에서 MM/DD 패턴 추출
            text = str(val).strip()
            m = date_re.search(text)
            if m:
                try:
                    mm, dd = int(m.group(1)), int(m.group(2))
                    if 1 <= mm <= 12 and 1 <= dd <= 31:
                        dates_found.append({
                            "cell": cell.coordinate,
                            "text": text,
                            "month": mm,
                            "day": dd,
                            "source": "text",
                        })
                except ValueError:
                    pass
    return dates_found


def estimate_tab_interval(sorted_dates: list[datetime]) -> int:
    """연속 탭 날짜들의 중앙값 간격(일)을 반환."""
    if len(sorted_dates) < 2:
        return 7
    gaps = [(sorted_dates[i + 1] - sorted_dates[i]).days
            for i in range(len(sorted_dates) - 1)
            if 3 <= (sorted_dates[i + 1] - sorted_dates[i]).days <= 30]
    if not gaps:
        return 7
    gaps.sort()
    return gaps[len(gaps) // 2]  # 중앙값


def build_cooccurrence(tab_etypes: list[list[str]]) -> dict[str, dict[str, str]]:
    """
    이벤트 유형 쌍별 동시 등장 규칙을 분석한다.
    두 유형이 항상 함께 등장하면 '허용', 한 번도 함께 등장하지 않으면 '금지'.
    """
    all_types = sorted({et for etypes in tab_etypes for et in etypes})
    rules: dict[str, dict[str, str]] = {}

    for i, a in enumerate(all_types):
        for b in all_types[i + 1:]:
            tabs_with_a = sum(1 for et in tab_etypes if a in et)
            tabs_with_b = sum(1 for et in tab_etypes if b in et)
            tabs_with_both = sum(1 for et in tab_etypes if a in et and b in et)

            if tabs_with_a == 0 or tabs_with_b == 0:
                continue

            # 두 유형이 모두 등장한 탭 중 함께 등장한 비율
            overlap_rate = tabs_with_both / max(tabs_with_a, tabs_with_b)

            if overlap_rate >= 0.7:
                rule = "허용"
            elif overlap_rate == 0:
                rule = "독립"  # 서로 배타적이지만 금지는 아님
            else:
                rule = "선택적"

            key = f"{a}+{b}"
            rules[key] = rule

    return rules


def build_section_order_pattern(tab_sections: dict[str, list]) -> dict[str, dict]:
    """
    이벤트 유형별 섹션 내 위치(앞/중/뒤) 패턴을 분석한다.
    반환: { "출석_이벤트": { "position": "앞", "avg_index": 0.1, "seen": 21 }, ... }
    """
    type_positions: dict[str, list[float]] = defaultdict(list)

    for tab, sections in tab_sections.items():
        total = len(sections)
        if total == 0:
            continue
        for idx, sec in enumerate(sections):
            etype = sec.get("event_type", "기타")
            relative_pos = idx / total  # 0=맨 앞, 1=맨 뒤
            type_positions[etype].append(relative_pos)

    result = {}
    for etype, positions in type_positions.items():
        avg = sum(positions) / len(positions)
        if avg < 0.33:
            pos_label = "앞"
        elif avg < 0.67:
            pos_label = "중간"
        else:
            pos_label = "뒤"
        result[etype] = {
            "position":  pos_label,
            "avg_index": round(avg, 2),
            "seen":      len(positions),
        }

    return result


# ─── 메인 분석 함수 ──────────────────────────────────────────────────────────

def analyze(source_xlsx: str) -> dict:
    wb = openpyxl.load_workbook(source_xlsx, data_only=True)
    date_tabs_raw = [n for n in wb.sheetnames if _srbe.is_date_tab(n)]

    tab_dates: list[tuple[str, datetime]] = []
    for name in date_tabs_raw:
        dt = tab_to_date(name)
        if dt:
            tab_dates.append((name, dt))
    tab_dates.sort(key=lambda x: x[1])

    if not tab_dates:
        return {"error": "날짜형 탭 없음", "source": source_xlsx}

    sorted_dts = [dt for _, dt in tab_dates]
    tab_interval = estimate_tab_interval(sorted_dts)

    # ── 탭별 섹션 스캔 ───────────────────────────────────────────────────────
    tab_sections: dict[str, list] = {}
    for name, _ in tab_dates:
        sections = _srbe.scan_tab_by_section(wb[name])
        tab_sections[name] = sections

    # ── 이벤트 유형 빈도 분석 ────────────────────────────────────────────────
    total_tabs = len(tab_dates)
    tab_etypes: list[list[str]] = []
    etype_counter: Counter = Counter()

    for name, sections in tab_sections.items():
        etypes = [s.get("event_type", "기타") for s in sections]
        tab_etypes.append(etypes)
        etype_counter.update(set(etypes))  # 탭당 1회만 카운트

    event_type_stats: dict[str, dict] = {}
    anchor_events: list[dict] = []

    for etype, count in etype_counter.most_common():
        rate = count / total_tabs
        if rate >= ANCHOR_RATE:
            priority = "anchor"
        elif rate >= COMMON_RATE:
            priority = "common"
        elif rate >= OPTIONAL_RATE:
            priority = "optional"
        else:
            priority = "rare"

        # 해당 유형의 이벤트 제목 예시 (최근 탭 기준 3개)
        title_examples: list[str] = []
        for name, sections in list(tab_sections.items())[-6:]:
            for sec in sections:
                if sec.get("event_type") == etype and sec["title"] not in title_examples:
                    title_examples.append(sec["title"])
                    if len(title_examples) >= 3:
                        break

        event_type_stats[etype] = {
            "seen_tabs":      count,
            "total_tabs":     total_tabs,
            "rate":           round(rate, 3),
            "rate_pct":       f"{rate * 100:.0f}%",
            "priority":       priority,
            "title_examples": title_examples,
        }

        if priority == "anchor":
            # anchor 이벤트의 전형적 시작 위치 파악
            first_positions: list[str] = []
            for (tab_name, tab_dt), etypes_in_tab in zip(tab_dates, tab_etypes):
                if etype in etypes_in_tab:
                    # 탭 시작일 기준 상대 위치
                    dom = tab_dt.day
                    if dom <= 7:
                        first_positions.append("월초 1~7일")
                    elif dom <= 14:
                        first_positions.append("월초 8~14일")
                    else:
                        first_positions.append("월중·하순")

            pos_counter = Counter(first_positions)
            typical_start = pos_counter.most_common(1)[0][0] if pos_counter else "월초"
            anchor_events.append({
                "type":          etype,
                "typical_start": typical_start,
                "rate_pct":      f"{rate * 100:.0f}%",
            })

    # ── 탭당 평균 이벤트 수 ───────────────────────────────────────────────────
    counts = [len(secs) for secs in tab_sections.values()]
    avg_events = round(sum(counts) / len(counts), 1) if counts else 0
    min_events = min(counts) if counts else 0
    max_events = max(counts) if counts else 0

    # ── 이벤트 유형별 섹션 위치 패턴 ────────────────────────────────────────
    order_pattern = build_section_order_pattern(tab_sections)

    # ── 동시 등장 규칙 ────────────────────────────────────────────────────────
    cooccurrence = build_cooccurrence(tab_etypes)

    # ── 갭 분석 (탭 간격 기반 추정) ─────────────────────────────────────────
    # 연속된 탭에서 동일 이벤트 유형이 몇 탭 연속 등장하는지 → 지속 기간 추정
    type_streaks: dict[str, list[int]] = defaultdict(list)
    for etype in etype_counter:
        streak = 0
        for etypes_in_tab in tab_etypes:
            if etype in etypes_in_tab:
                streak += 1
            else:
                if streak > 0:
                    type_streaks[etype].append(streak)
                    streak = 0
        if streak > 0:
            type_streaks[etype].append(streak)

    avg_duration_by_type: dict[str, int] = {}
    for etype, streaks in type_streaks.items():
        avg_streak = sum(streaks) / len(streaks)
        # 스트릭(연속 탭 수) × 탭 간격(일) = 추정 지속 기간
        avg_duration_by_type[etype] = round(avg_streak * tab_interval)

    # ── 월별 특이사항 ────────────────────────────────────────────────────────
    # 기존 탭에서 특정 월에만 등장하는 이벤트 유형 감지
    month_etypes: dict[str, set] = defaultdict(set)
    for (tab_name, tab_dt), etypes_in_tab in zip(tab_dates, tab_etypes):
        month_key = f"{tab_dt.month:02d}"
        month_etypes[month_key].update(etypes_in_tab)

    # 특정 월에만 등장하는 유형 감지 (전체 등장률 < 50% 이면서 해당 월에서는 높음)
    month_specific_events: dict[str, list[str]] = {}
    for month, etypes_in_month in month_etypes.items():
        special = []
        for etype in etypes_in_month:
            global_rate = etype_counter.get(etype, 0) / total_tabs
            # 해당 월 탭 수
            month_tabs = sum(
                1 for _, dt in tab_dates if f"{dt.month:02d}" == month
            )
            month_count = sum(
                1 for (_, dt), etypes_in_tab in zip(tab_dates, tab_etypes)
                if f"{dt.month:02d}" == month and etype in etypes_in_tab
            )
            month_rate = month_count / month_tabs if month_tabs > 0 else 0
            if global_rate < 0.40 and month_rate >= 0.60:
                special.append(etype)
        if special:
            month_specific_events[month] = special

    # 고정 시즌 노트와 병합
    month_specific: dict[str, dict] = {}
    all_months = sorted(set(list(MONTH_SEASON_NOTES.keys()) + list(month_specific_events.keys())))
    for m in all_months:
        month_specific[m] = {
            "notes":          MONTH_SEASON_NOTES.get(m, ""),
            "special_events": month_specific_events.get(m, []),
        }

    # ── 최근 탭 요약 (LLM 참고용) ─────────────────────────────────────────────
    recent_tabs_summary: list[dict] = []
    for tab_name, tab_dt in tab_dates[-4:]:
        sections = tab_sections.get(tab_name, [])
        recent_tabs_summary.append({
            "tab":         tab_name,
            "date":        tab_dt.strftime("%Y-%m-%d"),
            "event_count": len(sections),
            "event_types": [s.get("event_type", "기타") for s in sections],
        })

    return {
        "analyzed_at":          datetime.now().isoformat(timespec="seconds"),
        "source":               source_xlsx,
        "total_tabs_analyzed":  total_tabs,
        "tab_date_range":       {
            "first": tab_dates[0][1].strftime("%Y-%m-%d"),
            "last":  tab_dates[-1][1].strftime("%Y-%m-%d"),
        },
        "tab_interval_days":    tab_interval,
        "avg_events_per_tab":   avg_events,
        "min_events_per_tab":   min_events,
        "max_events_per_tab":   max_events,
        "event_type_stats":     event_type_stats,
        "anchor_events":        anchor_events,
        "avg_duration_by_type": avg_duration_by_type,
        "section_order_pattern": order_pattern,
        "cooccurrence_rules":   cooccurrence,
        "month_specific":       month_specific,
        "recent_tabs_summary":  recent_tabs_summary,
    }


# ─── CLI ─────────────────────────────────────────────────────────────────────

def _resolve_source() -> str:
    cp = _BASE_DIR / "output" / "json" / "current_project.json"
    if cp.exists():
        try:
            cfg = json.loads(cp.read_text(encoding="utf-8"))
            p = cfg.get("source_xlsx", "")
            if p and Path(p).exists():
                return p
        except Exception:
            pass
    fallback = _BASE_DIR / "Readdocs" / "[NC_KR] 라이브 이벤트 문서.xlsx"
    return str(fallback)


def main():
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

    parser = argparse.ArgumentParser(description="이벤트 일정 패턴 분석")
    parser.add_argument("source", nargs="?", default=None, help="소스 xlsx 경로")
    parser.add_argument("--output", default=None, help="출력 JSON 경로")
    args = parser.parse_args()

    source = args.source or _resolve_source()
    if not Path(source).exists():
        print(f"[오류] 파일 없음: {source}", file=sys.stderr)
        sys.exit(1)

    # 출력 경로 결정
    if args.output:
        out_path = Path(args.output)
    elif _paths:
        _paths.ensure_dirs()
        out_path = _paths.work_dir / "schedule_patterns.json"
    else:
        out_path = _BASE_DIR / "output" / "event-planner" / "work" / "schedule_patterns.json"

    out_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"[일정 패턴 분석 시작]")
    print(f"  소스: {source}")

    result = analyze(source)

    if "error" in result:
        print(f"[오류] {result['error']}", file=sys.stderr)
        sys.exit(1)

    out_path.write_text(
        json.dumps(result, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    # 콘솔 요약
    print(f"\n[분석 완료] 탭 {result['total_tabs_analyzed']}개")
    print(f"  탭 간격: {result['tab_interval_days']}일")
    print(f"  탭당 평균 이벤트: {result['avg_events_per_tab']}개 "
          f"(최소 {result['min_events_per_tab']} ~ 최대 {result['max_events_per_tab']})")

    print("\n[이벤트 유형별 등장률]")
    for etype, info in sorted(
        result["event_type_stats"].items(),
        key=lambda x: -x[1]["rate"],
    ):
        bar = "█" * int(info["rate"] * 10)
        print(f"  {etype:<22} {bar:<10} {info['rate_pct']:>4}  [{info['priority']}]")

    print("\n[Anchor 이벤트] (항상 포함해야 하는 이벤트)")
    for ev in result["anchor_events"]:
        print(f"  {ev['type']:<22} 전형 시작: {ev['typical_start']}")

    print(f"\n저장: {out_path}")


if __name__ == "__main__":
    main()
