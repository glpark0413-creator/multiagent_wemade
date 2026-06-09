#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
이벤트 패턴 분석기 — 소스 xlsx 전체 탭 학습 + 신규 탭 갭 탐지

동작:
  - xlsx 전체 날짜형 탭 스캔
  - B열 "숫자." 패턴으로 이벤트 섹션 추출
  - 이벤트 유형별 등장 빈도·월별 분포·탭당 이벤트 수 집계
  - 신규 탭(output xlsx)과 비교해 누락/추가 가능 이벤트 식별
  - 우선순위(required ≥80% / recommended 50~80% / optional 30~50%) 분류

사용:
  # 기본: 소스 xlsx 분석만
  python scripts/analyze_event_patterns.py

  # 신규 탭과 갭 비교
  python scripts/analyze_event_patterns.py \
      "Readdocs/[NC_KR] 라이브 이벤트 문서.xlsx" \
      "output/file/이벤트기획_260611_260618.xlsx" \
      "260611,260618"

출력:
  output/json/event_pattern_analysis.json
    - event_type_frequency  : 이벤트 유형별 등장 빈도·월별 분포·제목 예시
    - tab_count_stats       : 탭당 이벤트 수 통계 (avg/min/max)
    - new_tabs_gap_analysis : 신규 탭별 갭 분석 (누락 유형·우선순위)
"""
import io
import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

# ─── 경로 상수 ────────────────────────────────────────────────────────────────
_BASE_DIR_AP = Path(__file__).resolve().parent.parent
_CURRENT_PROJECT_FILE_AP = _BASE_DIR_AP / "output" / "json" / "current_project.json"
_NC_KR_FALLBACK_AP = str(_BASE_DIR_AP / "Readdocs" / "[NC_KR] 라이브 이벤트 문서.xlsx")

def _resolve_source_ap() -> str:
    """current_project.json 에서 source_xlsx 읽기. 없으면 NC_KR 레거시 경로 반환."""
    if _CURRENT_PROJECT_FILE_AP.exists():
        try:
            cfg = json.loads(_CURRENT_PROJECT_FILE_AP.read_text(encoding="utf-8"))
            p = cfg.get("source_xlsx", "")
            if p and Path(p).exists():
                return p
        except Exception:
            pass
    return _NC_KR_FALLBACK_AP

SOURCE = _resolve_source_ap()

import sys as _sys_ap
_sys_ap.path.insert(0, str(Path(__file__).resolve().parent))
from _project_config import load_project_paths as _load_paths_ap
_paths_ap = _load_paths_ap()

OUTPUT_DIR = Path("output")
OUTPUT_JSON_DIR = OUTPUT_DIR / "json"

# ─── 섹션 패턴 ────────────────────────────────────────────────────────────────
# 섹션 경계 형식:
#   1) "숫자." / "숫자)" 패턴  예: "1. 포인트 레이스 이벤트"
#   2) "이벤트 제목 :" 패턴    예: "이벤트 제목 : 올스타 직행! 14일 출석 이벤트!"
#   3) "N월 이달의 쿠폰" 패턴   예: "5월 이달의 쿠폰"
#   4) 커스텀 쿠폰 헤더         예: "폴리볼 쿠폰", "KBO 쿠폰"
SECTION_PATTERN       = re.compile(r"^\d+[.)]")
EVENT_TITLE_PATTERN   = re.compile(r"^이벤트\s*제목\s*:\s*(.+)$")
COUPON_TITLE_PATTERN  = re.compile(r"^\d+월\s+이달의\s+쿠폰")
CUSTOM_COUPON_PATTERN = re.compile(r"^(?!■|∎)[가-힣A-Za-z0-9]+\s+쿠폰$")


def parse_section_title(val: str) -> "str | None":
    """B열 셀 값이 섹션 경계이면 정규화된 이벤트 제목을 반환, 아니면 None."""
    if SECTION_PATTERN.match(val):
        return val
    m = EVENT_TITLE_PATTERN.match(val)
    if m:
        return m.group(1).strip()
    if COUPON_TITLE_PATTERN.match(val):
        return val
    if CUSTOM_COUPON_PATTERN.match(val):
        return val
    return None

# ─── 이벤트 유형 정규화 키워드 (공통 + 장르별) ───────────────────────────────
EVENT_TYPE_KEYWORDS = [
    # 공통
    ("출석",          "출석_이벤트"),
    ("미션",          "미션_이벤트"),
    ("패스",          "패스"),
    ("교환 상점",     "교환상점_이벤트"),
    ("교환소",        "교환상점_이벤트"),
    ("할인",          "할인_이벤트"),
    ("룰렛",          "룰렛_이벤트"),
    ("빙고",          "빙고_이벤트"),
    ("포인트 레이스", "포인트레이스_이벤트"),
    ("응모권",        "응모권_이벤트"),
    ("쿠폰",          "쿠폰_이벤트"),          # ← 매월 쿠폰 / 커스텀 쿠폰 포함
    ("승부 예측",     "승부예측_이벤트"),
    ("예측",          "승부예측_이벤트"),
    # 야구 스포츠
    ("야구공 찾기",   "탐색형_이벤트"),
    ("찾기",          "탐색형_이벤트"),
    ("플레이 미션",   "플레이미션_이벤트"),
    # NC_KR MMORPG
    ("던전",          "던전_이벤트"),
    ("레이드",        "레이드_이벤트"),
    ("보물 상자",     "보물상자_이벤트"),
    ("주사위",        "주사위_이벤트"),
    ("제작",          "제작_이벤트"),
    ("우편",          "우편지급_이벤트"),
    ("성장 가이드",   "성장가이드_이벤트"),
    ("지령",          "지령_이벤트"),
]

# 이벤트 유형 중요도 — 역사적 빈도와 무관하게 항상 추적할 핵심 유형
CRITICAL_EVENT_TYPES: set[str] = {"출석_이벤트", "미션_이벤트", "패스"}

# ─── 갭 분석 우선순위 기준 ────────────────────────────────────────────────────
RATE_THRESHOLDS: dict[str, float] = {
    "required":    0.80,   # ≥ 80%  → ❗ 누락 확인 필요
    "recommended": 0.50,   # 50~80% → ⚠ 추가 권장
    "optional":    0.30,   # 30~50% → 〇 선택 사항
    # < 30% : rare — 자동 추천 제외
}

PRIORITY_ICON: dict[str, str] = {
    "required":    "❗ 누락 확인 필요",
    "recommended": "⚠ 추가 권장",
    "optional":    "〇 선택 사항",
    "rare":        "",
}

BAR_CHARS = ("○", "◐", "●")  # 등장률 시각화용


# ─── 유틸 함수 ────────────────────────────────────────────────────────────────

def is_date_tab(name: str) -> bool:
    return bool(re.fullmatch(r"\d{6}", str(name)))


def tab_to_year_month(tab_name: str) -> str | None:
    """YYMMDD → YYYY-MM 문자열. 파싱 실패 시 None."""
    if not re.fullmatch(r"\d{6}", tab_name):
        return None
    try:
        dt = datetime.strptime("20" + tab_name, "%Y%m%d")
        return dt.strftime("%Y-%m")
    except ValueError:
        return None


def normalize_event_type(title: str) -> str:
    """이벤트 섹션 제목 → 표준 유형명."""
    for keyword, etype in EVENT_TYPE_KEYWORDS:
        if keyword.lower() in title.lower():
            return etype
    return "기타"


def rate_bar(rate: float, width: int = 4) -> str:
    """0.0~1.0 → ████○○ 형식 막대."""
    filled = round(rate * width)
    return "█" * filled + "○" * (width - filled)


# ─── 섹션 추출 ────────────────────────────────────────────────────────────────

def extract_event_sections(ws) -> list[dict]:
    """워크시트 B열에서 이벤트 섹션 제목 추출.

    read_only 모드에서 EmptyCell 객체가 섞일 수 있으므로 column/value 접근 전 안전 체크.
    """
    sections = []
    for row in ws.iter_rows():
        for cell in row:
            # read_only 모드의 EmptyCell 방어
            col = getattr(cell, "column", None)
            if col != 2:
                continue
            val = getattr(cell, "value", None)
            if not val:
                continue
            val_str = str(val).strip()
            title = parse_section_title(val_str)
            if title is not None:
                sections.append({
                    "title": title,
                    "event_type": normalize_event_type(title),
                    "cell": getattr(cell, "coordinate", f"B{getattr(cell, 'row', '?')}"),
                    "row": getattr(cell, "row", None),
                })
    return sections


# ─── 빈도 통계 ────────────────────────────────────────────────────────────────

def build_frequency_stats(all_tab_sections: dict[str, list]) -> dict:
    """
    이벤트 유형별 등장 빈도 및 월별 분포 집계.

    반환:
      {
        "frequency_stats": { 이벤트유형: {count, rate, priority, ...} },
        "tab_count_stats": { avg, min, max, recent }
      }
    """
    total_tabs = len(all_tab_sections)
    if total_tabs == 0:
        return {"frequency_stats": {}, "tab_count_stats": {}}

    # 유형별 등장 탭 목록 / 제목 예시 / 월별 분포
    type_tabs: dict[str, list[str]] = defaultdict(list)
    type_titles: dict[str, list[str]] = defaultdict(list)
    monthly: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))

    events_per_tab: list[int] = []

    for tab in sorted(all_tab_sections.keys()):
        sections = all_tab_sections[tab]
        events_per_tab.append(len(sections))
        ym = tab_to_year_month(tab) or "unknown"

        seen_types: set[str] = set()
        for sec in sections:
            etype = sec["event_type"]
            title = sec["title"]

            if etype not in seen_types:
                type_tabs[etype].append(tab)
                seen_types.add(etype)
                monthly[etype][ym] += 1

            # 제목 예시 최대 5개 수집
            if title not in type_titles[etype] and len(type_titles[etype]) < 5:
                type_titles[etype].append(title)

    # 모든 유형 = 실제 등장 유형 + 항상 추적할 핵심 유형
    all_types = set(type_tabs.keys()) | CRITICAL_EVENT_TYPES

    frequency_stats: dict[str, dict] = {}
    for etype in sorted(all_types):
        tabs = type_tabs.get(etype, [])
        rate = len(tabs) / total_tabs

        if rate >= RATE_THRESHOLDS["required"]:
            priority = "required"
        elif rate >= RATE_THRESHOLDS["recommended"]:
            priority = "recommended"
        elif rate >= RATE_THRESHOLDS["optional"]:
            priority = "optional"
        else:
            priority = "rare"

        frequency_stats[etype] = {
            "count":                len(tabs),
            "total_tabs":           total_tabs,
            "rate":                 round(rate, 3),
            "rate_pct":             f"{rate * 100:.0f}%",
            "priority":             priority,
            "seen_tabs":            sorted(tabs),
            "recent_tabs":          sorted(tabs)[-5:],
            "title_examples":       type_titles.get(etype, [])[-3:],
            "monthly_distribution": dict(sorted(monthly.get(etype, {}).items())),
        }

    # 탭당 이벤트 수 통계
    tab_count_stats: dict = {}
    if events_per_tab:
        tab_count_stats = {
            "avg":    round(sum(events_per_tab) / len(events_per_tab), 1),
            "min":    min(events_per_tab),
            "max":    max(events_per_tab),
            "recent": events_per_tab[-5:],
        }

    return {"frequency_stats": frequency_stats, "tab_count_stats": tab_count_stats}


# ─── 갭 분석 ─────────────────────────────────────────────────────────────────

def analyze_gaps(
    new_tab_sections: list[dict],
    frequency_stats: dict,
    tab_count_stats: dict,
) -> dict:
    """
    신규 탭의 이벤트 유형 목록 vs 역사적 패턴 비교 → 갭 분석.

    반환:
      {
        present_types, event_count, historical_avg,
        is_below_avg, gap_events, has_required_gaps, has_recommended_gaps
      }
    """
    present_types: set[str] = {sec["event_type"] for sec in new_tab_sections}

    gaps: list[dict] = []
    for etype, stats in frequency_stats.items():
        if etype == "기타":
            continue
        if etype in present_types:
            continue
        if stats["priority"] == "rare":
            continue

        gaps.append({
            "event_type":          etype,
            "historical_rate":     stats["rate"],
            "historical_rate_pct": stats["rate_pct"],
            "priority":            stats["priority"],
            "priority_label":      PRIORITY_ICON.get(stats["priority"], ""),
            "title_examples":      stats["title_examples"],
            "count_out_of":        f"{stats['count']}/{stats['total_tabs']}",
        })

    # required → recommended → optional 순 정렬
    priority_order = {"required": 0, "recommended": 1, "optional": 2}
    gaps.sort(key=lambda x: priority_order.get(x["priority"], 9))

    current_count = len(new_tab_sections)
    avg_count = tab_count_stats.get("avg", 0)

    return {
        "present_types":           sorted(present_types),
        "event_count":             current_count,
        "historical_avg":          avg_count,
        "historical_min":          tab_count_stats.get("min", 0),
        "historical_max":          tab_count_stats.get("max", 0),
        "is_below_avg":            current_count < avg_count,
        "gap_events":              gaps,
        "has_required_gaps":       any(g["priority"] == "required"    for g in gaps),
        "has_recommended_gaps":    any(g["priority"] == "recommended" for g in gaps),
    }


# ─── 콘솔 출력 ───────────────────────────────────────────────────────────────

def _print_frequency_table(frequency_stats: dict, total_tabs: int) -> None:
    print("\n[이벤트 유형별 역사 등장률]")
    print(f"  {'이벤트 유형':<22} | {'막대':>4} | {'등장률':>5} | {'횟수':>7} | 우선순위")
    print("  " + "─" * 68)
    sorted_items = sorted(frequency_stats.items(), key=lambda x: -x[1]["rate"])
    for etype, info in sorted_items:
        bar = rate_bar(info["rate"])
        icon = {"required": "❗", "recommended": "⚠", "optional": "〇", "rare": " "}.get(info["priority"], " ")
        print(
            f"  {etype:<22} | {bar} | {info['rate_pct']:>4} | "
            f"{info['count']:>3}/{total_tabs:<3} | {icon} {info['priority']}"
        )


def _print_gap_table(tab: str, gap: dict) -> None:
    print(f"\n  [{tab}] 현재 이벤트 수: {gap['event_count']}개  "
          f"(역사 평균 {gap['historical_avg']}개, 범위 {gap['historical_min']}~{gap['historical_max']}개)")

    if gap["is_below_avg"]:
        diff = round(gap["historical_avg"] - gap["event_count"], 1)
        print(f"  ⚠ 역사 평균보다 {diff}개 적습니다.")

    if not gap["gap_events"]:
        print("  → 누락 이벤트 없음 ✅")
        return

    print(f"  {'#':<3} | {'이벤트 유형':<22} | {'등장률':>5} | {'횟수':>7} | 우선순위")
    print("  " + "─" * 64)
    for i, g in enumerate(gap["gap_events"], 1):
        print(
            f"  {i:<3} | {g['event_type']:<22} | {g['historical_rate_pct']:>4} | "
            f"{g['count_out_of']:>7} | {g['priority_label']}"
        )


# ─── 메인 ─────────────────────────────────────────────────────────────────────

def main(
    source: str = SOURCE,
    new_tabs_path: str | None = None,
    new_tab_names: str | None = None,
) -> dict:
    if _paths_ap:
        _paths_ap.ensure_dirs()
    OUTPUT_DIR.mkdir(exist_ok=True)
    OUTPUT_JSON_DIR.mkdir(exist_ok=True)

    out_path = _paths_ap.event_pattern_analysis if _paths_ap else OUTPUT_JSON_DIR / "event_pattern_analysis.json"

    print("[이벤트 패턴 분석]")
    print(f"  소스: {source}")

    # ① 소스 xlsx 전체 날짜형 탭 스캔
    wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
    date_tabs = [n for n in wb.sheetnames if is_date_tab(n)]
    print(f"  날짜형 탭: {len(date_tabs)}개")

    all_tab_sections: dict[str, list] = {}
    for tab in date_tabs:
        try:
            sections = extract_event_sections(wb[tab])
            all_tab_sections[tab] = sections
        except Exception as e:
            print(f"  [{tab}] 읽기 오류: {e} — 건너뜀")
    wb.close()

    # ② 빈도 통계 집계
    stats = build_frequency_stats(all_tab_sections)
    frequency_stats: dict = stats["frequency_stats"]
    tab_count_stats: dict = stats["tab_count_stats"]

    _print_frequency_table(frequency_stats, len(date_tabs))

    # ③ 신규 탭 갭 분석 (선택)
    new_tabs_gap_analysis: dict[str, dict] = {}

    if new_tabs_path and new_tab_names:
        target_tabs = [t.strip() for t in new_tab_names.split(",") if t.strip()]
        print(f"\n[신규 탭 갭 분석] — {target_tabs}")

        wb2 = openpyxl.load_workbook(new_tabs_path, data_only=True, read_only=True)
        for tab in target_tabs:
            if tab not in wb2.sheetnames:
                print(f"  [{tab}] 탭 없음 — 건너뜀")
                continue
            try:
                sections = extract_event_sections(wb2[tab])
                gap = analyze_gaps(sections, frequency_stats, tab_count_stats)
                new_tabs_gap_analysis[tab] = gap
                _print_gap_table(tab, gap)
            except Exception as e:
                print(f"  [{tab}] 분석 오류: {e}")
        wb2.close()

    # ④ JSON 저장
    result = {
        "analyzed_at":         datetime.now().isoformat(timespec="seconds"),
        "source":              source,
        "total_source_tabs":   len(date_tabs),
        "tab_count_stats":     tab_count_stats,
        "event_type_frequency": frequency_stats,
        "new_tabs_gap_analysis": new_tabs_gap_analysis,
    }

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"\n저장: {out_path}")
    return result


if __name__ == "__main__":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

    src       = sys.argv[1] if len(sys.argv) > 1 else SOURCE
    new_path  = sys.argv[2] if len(sys.argv) > 2 else None
    new_names = sys.argv[3] if len(sys.argv) > 3 else None
    main(src, new_path, new_names)
