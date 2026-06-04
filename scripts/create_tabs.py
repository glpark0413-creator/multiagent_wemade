#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
이벤트 시트 탭 생성 스크립트 — NC_KR (나이트 크로우)
- 날짜/헤더 자동 갱신 (한국어 날짜 텍스트 + datetime 셀)
- datetime 셀 직접 갱신 (date_map 사용)
- output/json/event_names_config.json 이 있으면 이벤트 명칭 자동 치환
- 시즌·월 키워드 포함 셀 경고 출력
"""
import io
import json
import re
import sys
from pathlib import Path
from datetime import datetime, date
import openpyxl
from openpyxl.styles import PatternFill

import sys as _sys_ct
_sys_ct.path.insert(0, str(Path(__file__).resolve().parent))
from _project_config import load_project_paths as _load_paths_ct
_paths_ct = _load_paths_ct()

def _resolve_source_ct() -> str:
    """current_project.json 에서 source_xlsx 읽기. 없으면 빈 문자열 반환."""
    _base = Path(__file__).resolve().parent.parent
    _cpf  = _base / "output" / "json" / "current_project.json"
    if _cpf.exists():
        try:
            cfg = json.loads(_cpf.read_text(encoding="utf-8"))
            p = cfg.get("source_xlsx", "")
            if p and Path(p).exists():
                return p
        except Exception:
            pass
    return ""

SOURCE = _resolve_source_ct()

OUTPUT_DIR = Path("output")
OUTPUT_FILE_DIR = _paths_ct.file_dir if _paths_ct else Path("output") / "file"
OUTPUT_JSON_DIR = _paths_ct.work_dir if _paths_ct else Path("output") / "json"
OUTPUT_FILE = OUTPUT_FILE_DIR / "이벤트기획_260625_260702.xlsx"
EVENT_NAMES_CONFIG = OUTPUT_JSON_DIR / "event_names_config.json"

# ─── 변경 셀 하이라이트 색상 ────────────────────────────────────────────────
# 날짜(datetime) 변경: 연파랑
FILL_DATE   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
# 날짜 문자열·헤더 변경: 연초록
FILL_TEXT   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
# 보상 변경: 연주황 (이전 시트 대비 보상 아이템/수량 차이)
FILL_REWARD = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

# 시즌·월 키워드: 이 단어가 포함된 셀은 수동 검토 권장
SEASON_KEYWORDS = [
    "봄의", "여름의", "가을의", "겨울의",
    "황금의", "축제의", "기념",
    "1월의", "2월의", "3월의", "4월의", "5월의",
    "6월의", "7월의", "8월의", "9월의", "10월의", "11월의", "12월의",
    "얼리썸머", "초여름", "쿨 서머", "5월의",
]

# 각 탭별 설정 — FB_GL 날짜 치환 규칙
# 260625 (260611 참조, +14일 시프트): 2026-06-25 이벤트 주차 생성 (14일 주기 메인 탭)
# 260702 (260618 참조, +14일 시프트): 2026-07-02 이벤트 주차 생성 (7일 주기 포인트레이스/빙고 탭)
UPDATES = {
    "260625": {
        "source_tab": "260611",
        # datetime 셀용 직접 매핑 (MM/DD → MM/DD), +14일
        # 출석 이벤트 날짜 셀(C13~C26): 06/11~06/24 → 06/25~07/08
        "date_map": {
            "06/11": "06/25",
            "06/12": "06/26",
            "06/13": "06/27",
            "06/14": "06/28",
            "06/15": "06/29",
            "06/16": "06/30",
            "06/17": "07/01",
            "06/18": "07/02",
            "06/19": "07/03",
            "06/20": "07/04",
            "06/21": "07/05",
            "06/22": "07/06",
            "06/23": "07/07",
            "06/24": "07/08",
            "06/25": "07/09",
        },
        # 문자열 셀용 치환 규칙 (순서 중요 — 긴 패턴 먼저)
        "replacements": [
            # 헤더 (B3)
            ("06.11_ Event", "06.25_ Event"),
            # 출석 이벤트 기간 (수요일 종료)
            ("06/11(목) 09:00 ~ 06/24(수) 23:59", "06/25(목) 09:00 ~ 07/08(수) 23:59"),
            # 14일 이벤트 기간 (응모권·플레이미션·교환소·야구공찾기 등)
            ("06/11(목) 09:00 ~ 06/25(목) 08:59:59 (14일)", "06/25(목) 09:00 ~ 07/09(목) 08:59:59 (14일)"),
            # 장기 이벤트 기간 (승부예측 등 07/16 마감)
            ("06/11(목) 09:00 ~ 07/16(목) 08:59:59", "06/25(목) 09:00 ~ 07/30(목) 08:59:59"),
            # 나머지 06/11(목) 패턴 (위 규칙에서 처리 안 된 경우)
            ("06/11(목)", "06/25(목)"),
        ],
        "event_name_replacements": [],
    },
    "260702": {
        "source_tab": "260618",
        # datetime 셀용 직접 매핑 (MM/DD → MM/DD), +14일
        "date_map": {
            "06/18": "07/02",
            "06/19": "07/03",
            "06/20": "07/04",
            "06/21": "07/05",
            "06/22": "07/06",
            "06/23": "07/07",
            "06/24": "07/08",
            "06/25": "07/09",
        },
        # 문자열 셀용 치환 규칙 (순서 중요 — 긴 패턴 먼저)
        "replacements": [
            # 헤더 (B3)
            ("06.18_ Event", "07.02_ Event"),
            # 포인트 레이스 기간 (7일간 진행, 공백 없는 형식)
            ("06/18(목) 09:00 ~ 06/25(목) 08:59 (7일간 진행)", "07/02(목) 09:00 ~ 07/09(목) 08:59 (7일간 진행)"),
            # 룰렛·빙고 기간 (공백 포함 "06/18 (목)" 형식)
            ("06/18 (목) 09:00 ~ 06/25(목) 08:59:59", "07/02(목) 09:00 ~ 07/09(목) 08:59:59"),
            # 나머지 "06/18 (목)" (공백 있는 형식)
            ("06/18 (목)", "07/02(목)"),
            # 나머지 "06/18(목)" (공백 없는 형식)
            ("06/18(목)", "07/02(목)"),
            # 나머지 "06/25(목)" (위에서 처리 안 된 경우)
            ("06/25(목)", "07/09(목)"),
        ],
        "event_name_replacements": [],
    },
}


def load_event_names_config():
    """
    output/event_names_config.json 이 존재하면 로드해서 탭별 이벤트 명칭 치환 목록 반환.
    Claude가 장르·시즌 분석 후 미리 생성해두는 파일.

    반환 형식:
    {
      "260611": [("구이름", "새이름"), ...],
      "260618": []
    }
    """
    if not EVENT_NAMES_CONFIG.exists():
        return {}
    with open(EVENT_NAMES_CONFIG, encoding="utf-8-sig") as f:
        cfg = json.load(f)

    genre = cfg.get("genre", "")
    target_month = cfg.get("target_month", "")
    genre_phrases = cfg.get("genre_phrases", [])
    raw = cfg.get("event_name_replacements", {})

    if genre or target_month:
        print(f"  [config] 장르={genre}, 대상월={target_month}")
    if genre_phrases:
        print(f"  [config] 장르 키워드: {', '.join(genre_phrases[:8])}")

    return {tab: [tuple(r) for r in repls] for tab, repls in raw.items()}


def apply_replacements(ws, replacements, date_map=None, event_name_replacements=None):
    """
    워크시트 전체 셀에 치환 적용 + 날짜 변경 하이라이트.
    - datetime/date 객체: date_map으로 직접 날짜 갱신 → 연파랑 하이라이트
    - str (날짜 문자열·헤더): replacements + date_map → 연초록 하이라이트
    - str (이벤트 명칭): event_name_replacements → 하이라이트 없음 (보상 비교와 별개)
    """
    changed = []
    _event_repls = list(event_name_replacements or [])
    _text_repls  = list(replacements)

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue

            # ── datetime / date 객체: date_map으로 직접 매핑 → 연파랑 ──────────
            if isinstance(cell.value, (datetime, date)):
                if not date_map:
                    continue
                old_date = cell.value
                mmdd = old_date.strftime("%m/%d")
                if mmdd not in date_map:
                    continue
                new_mmdd = date_map[mmdd]
                month, day = map(int, new_mmdd.split("/"))
                year = old_date.year
                if isinstance(old_date, datetime):
                    new_date = datetime(year, month, day,
                                        old_date.hour, old_date.minute, old_date.second)
                else:
                    new_date = date(year, month, day)
                changed.append((cell.coordinate, str(old_date), str(new_date)))
                cell.value = new_date
                # 날짜 변경은 하이라이트 없음 (보상 변경만 표기)
                continue

            if not isinstance(cell.value, str):
                continue

            original_val = cell.value
            new_val      = cell.value

            # 1) 이벤트 명칭 치환 (하이라이트 없음 — 보상 비교로 대체)
            for old, new in _event_repls:
                new_val = new_val.replace(old, new)

            # 2) 날짜·헤더 등 일반 텍스트 치환
            for old, new in _text_repls:
                new_val = new_val.replace(old, new)

            # 3) date_map 을 문자열 셀에도 적용 — 단일 패스 regex
            if date_map:
                _slash_map = dict(date_map)
                _slash_pat = re.compile("|".join(re.escape(k) for k in sorted(_slash_map, key=len, reverse=True)))
                new_val = _slash_pat.sub(lambda m: _slash_map[m.group(0)], new_val)
                _dot_map = {k.replace("/", "."): v.replace("/", ".") for k, v in date_map.items()}
                _dot_pat = re.compile("|".join(re.escape(k) for k in sorted(_dot_map, key=len, reverse=True)))
                new_val = _dot_pat.sub(lambda m: _dot_map[m.group(0)], new_val)

            if new_val != original_val:
                changed.append((cell.coordinate, original_val, new_val))
                cell.value = new_val
                # 날짜·텍스트 변경은 하이라이트 없음 (보상 변경만 표기)

    return changed


def snapshot_ws(ws) -> dict:
    """워크시트 셀 값 스냅샷 반환 {(row, col): value}."""
    return {
        (cell.row, cell.column): cell.value
        for row in ws.iter_rows()
        for cell in row
    }


def _find_reward_col_pairs(ws) -> list:
    """
    워크시트에서 '보상 아이템' 헤더를 찾고, 같은 행에서 바로 오른쪽에 있는
    '보상 수량' 헤더와 쌍을 이룬다.

    이벤트마다 보상 열 위치가 다르기 때문에 헤더 텍스트로 동적 탐색.

    예) F12="보상 아이템", H12="보상 수량"
        E34="보상 아이템", G34="보상 수량"
        E38="보상 아이템", F38="보상 수량", G38="보상 아이템", I38="보상 수량"

    반환: [(header_row, item_col, qty_col), ...]
    """
    # 모든 "보상 아이템" / "보상 수량" 위치 수집
    item_positions: list[tuple[int, int]] = []
    qty_positions:  list[tuple[int, int]] = []

    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v == "보상 아이템":
                item_positions.append((cell.row, cell.column))
            elif v == "보상 수량":
                qty_positions.append((cell.row, cell.column))

    # 각 "보상 아이템"에 대해 같은 행에서 오른쪽에 가장 가까운 "보상 수량" 매핑
    pairs = []
    for h_row, item_col in item_positions:
        candidates = [(r, c) for r, c in qty_positions if r == h_row and c > item_col]
        if candidates:
            qty_col = min(candidates, key=lambda x: x[1])[1]
            pairs.append((h_row, item_col, qty_col))

    return pairs


def _find_prev_same_type_ws(wb, ref_tab: str, ref_pairs: list) -> tuple:
    """
    소스 파일에서 ref_tab과 동일한 보상 열 구조를 가진 가장 최근의 이전 탭을 찾는다.
    '동일 타입' 판단: (item_col, qty_col) 쌍이 2개 이상 겹치면 동일 타입으로 간주.
    (1개만 겹치면 우연의 일치 가능성이 높으므로 제외)

    Returns: (tab_name, worksheet) or (None, None)
    """
    import re as _re
    date_tabs = sorted(
        [s for s in wb.sheetnames if _re.fullmatch(r"\d{6}", s)],
        reverse=True,
    )

    ref_col_sigs = {(ic, qc) for _, ic, qc in ref_pairs}
    past_ref = False

    for tab in date_tabs:
        if tab == ref_tab:
            past_ref = True
            continue
        if not past_ref:
            continue

        ws = wb[tab]
        tab_pairs = _find_reward_col_pairs(ws)
        tab_col_sigs = {(ic, qc) for _, ic, qc in tab_pairs}

        overlap = ref_col_sigs & tab_col_sigs
        if len(overlap) >= 3:           # 3개 이상 겹쳐야 동일 타입
            return tab, ws

    return None, None


def _find_all_prev_same_type_ws(wb, ref_tab: str, ref_pairs: list) -> list:
    """
    ref_tab 이전의 동일 타입 탭을 모두 찾아 [(탭명, ws), ...] 최신→구버전 순으로 반환.

    타입 판별 우선순위:
      1. 일반 임계값(3): 열 쌍이 3개 이상 겹치면 동일 타입 (B타입 ↔ B타입)
      2. A타입 임계값(1): 참조 탭과 후보 탭 모두 ic ≤ 4 (C열) 열 쌍을 가지고,
         그 쌍이 1개 이상 겹치면 동일 A타입 (포인트레이스/룰렛/빙고 계열)
         → B타입(ic=5,6,7) 탭이 A타입 탭으로 오매칭되는 것을 방지
    """
    import re as _re
    date_tabs = sorted(
        [s for s in wb.sheetnames if _re.fullmatch(r"\d{6}", s)],
        reverse=True,
    )
    ref_col_sigs = {(ic, qc) for _, ic, qc in ref_pairs}
    # 참조 탭의 A타입 시그니처: ic ≤ 4 (C열 이하) 열 쌍
    ref_a_sigs = {(ic, qc) for ic, qc in ref_col_sigs if ic <= 4}

    past_ref = False
    result = []
    for tab in date_tabs:
        if tab == ref_tab:
            past_ref = True
            continue
        if not past_ref:
            continue
        ws = wb[tab]
        tab_pairs = _find_reward_col_pairs(ws)
        tab_col_sigs = {(ic, qc) for _, ic, qc in tab_pairs}
        overlap = ref_col_sigs & tab_col_sigs

        # 방법 1: 일반 임계값 (B타입 ↔ B타입)
        if len(overlap) >= 3:
            result.append((tab, ws))
            continue

        # 방법 2: A타입 매칭 (포인트레이스/룰렛/빙고 ↔ 동일 계열)
        # 참조 탭이 A타입 시그니처를 가지고, 후보 탭과 A타입 열 쌍이 겹치면 매칭
        if ref_a_sigs:
            tab_a_sigs = {(ic, qc) for ic, qc in tab_col_sigs if ic <= 4}
            if len(ref_a_sigs & tab_a_sigs) >= 1:
                result.append((tab, ws))

    return result


def apply_balanced_rewards(ws_new, history_list: list) -> int:
    """
    역사적 보상 패턴을 기반으로 새 탭 보상을 업데이트한다.

    history_list: [(탭명, worksheet), ...] 참조 탭(최신) + 이전 탭들 순서
      예) [("260611", ws_611), ("260514", ws_514)]

    동작 방식:
      1. 아이템 교체: 참조 탭과 이전 탭의 아이템이 다른 슬롯 →
         이전 탭 아이템으로 교체 (역사적 다양성 확보)
      2. 수량 소폭 조정 (±5% 이내):
         - 아이템이 교체된 슬롯 → 이전 탭의 수량 그대로 사용
         - 아이템이 같은 슬롯   → 참조 수량에서 이전 수량 방향으로 5% 이동

    반환: 변경된 셀 수 (아이템 변경 + 수량 변경 합계)
    """
    if len(history_list) < 2:
        return 0

    _SKIP = {"보상 아이템", "보상 수량", "확률", "획득 보상"}

    _ref_name, ws_ref = history_list[0]   # 참조 탭 (260611)
    _prv_name, ws_prv = history_list[1]   # 이전 동일 타입 탭 (260514)

    new_pairs = _find_reward_col_pairs(ws_new)
    prv_pairs = _find_reward_col_pairs(ws_prv)

    # 이전 탭 섹션 매핑: (ic,qc) → [header_row, ...]
    prv_hdr_map: dict[tuple, list[int]] = {}
    for h, ic, qc in prv_pairs:
        prv_hdr_map.setdefault((ic, qc), []).append(h)

    col_pair_counter: dict[tuple, int] = {}
    item_changed = 0
    qty_changed  = 0

    for new_h, ic, qc in new_pairs:
        pair_key = (ic, qc)
        sec_idx  = col_pair_counter.get(pair_key, 0)
        col_pair_counter[pair_key] = sec_idx + 1

        prv_matches = prv_hdr_map.get(pair_key, [])
        has_mapping = sec_idx < len(prv_matches)
        prv_h = prv_matches[sec_idx] if has_mapping else None

        offset = 1
        while True:
            cur_iv = ws_new.cell(new_h + offset, ic).value   # 현재 아이템 (= 참조 탭)
            if cur_iv is None:
                break
            if cur_iv in _SKIP:
                offset += 1
                continue

            cur_qv = ws_new.cell(new_h + offset, qc).value   # 현재 수량

            if not has_mapping:
                # ── 매핑 없는 섹션: 행 번호 기반 ±3% 결정적 조정 ──────────────
                # (이전 이력 없는 신규 섹션도 일관된 변화 적용)
                if cur_qv is not None and cur_qv not in _SKIP:
                    try:
                        ref_q = float(cur_qv)
                        sign  = 1 if (new_h + offset) % 2 == 0 else -1
                        step  = max(1, int(round(ref_q * 0.03)))
                        new_q = ref_q + sign * step
                        if isinstance(cur_qv, int) or (
                                isinstance(cur_qv, float) and cur_qv == int(cur_qv)):
                            new_q = int(round(new_q))
                        else:
                            new_q = round(new_q, 4)
                        new_q = max(1, new_q)   # 보상 수량 최솟값 1 보장
                        if new_q != cur_qv:
                            ws_new.cell(new_h + offset, qc).value = new_q
                            qty_changed += 1
                    except (ValueError, TypeError):
                        pass
                offset += 1
                continue

            prv_iv  = ws_prv.cell(prv_h + offset, ic).value  # 이전 탭 아이템
            prv_qv  = ws_prv.cell(prv_h + offset, qc).value  # 이전 탭 수량

            if prv_iv is None or prv_iv in _SKIP or isinstance(prv_iv, (datetime, date)):
                offset += 1
                continue

            # ── 아이템 교체 여부 결정 ─────────────────────────────────────────
            if cur_iv != prv_iv:
                # 아이템이 다름 → 이전 탭 아이템으로 교체
                ws_new.cell(new_h + offset, ic).value = prv_iv
                item_changed += 1

                # 수량도 이전 탭 수량 사용
                if prv_qv is not None and prv_qv not in _SKIP:
                    ws_new.cell(new_h + offset, qc).value = prv_qv
                    qty_changed += 1

            else:
                # 아이템이 같음 → 수량을 이전 방향으로 소폭 이동
                if cur_qv is not None and prv_qv is not None and prv_qv not in _SKIP:
                    try:
                        ref_q = float(cur_qv)
                        prv_q = float(prv_qv)
                        diff  = prv_q - ref_q
                        # 역사 데이터와 차이 없으면 → 행 번호 기반 결정적 소폭 조정
                        if abs(diff) < 1e-9:
                            sign  = 1 if (new_h + offset) % 2 == 0 else -1
                            step  = max(1, int(round(ref_q * 0.03)))
                            new_q = ref_q + sign * step
                        else:
                            # ref 기준 5% 이내로 소폭 이동
                            step  = ref_q * 0.05
                            move  = max(-step, min(step, diff * 0.5))
                            new_q = ref_q + move

                        if isinstance(cur_qv, int) or (
                                isinstance(cur_qv, float) and cur_qv == int(cur_qv)):
                            new_q = int(round(new_q))
                        else:
                            new_q = round(new_q, 4)
                        new_q = max(1, new_q)   # 보상 수량 최솟값 1 보장
                        if new_q != cur_qv:
                            ws_new.cell(new_h + offset, qc).value = new_q
                            qty_changed += 1
                    except (ValueError, TypeError):
                        pass

            offset += 1

    print(f"    아이템 교체: {item_changed}개, 수량 조정: {qty_changed}개")
    return item_changed + qty_changed


def highlight_reward_diffs(ws_new, ref_snapshot: dict) -> list:
    """
    보상 셀을 참조 탭 스냅샷과 비교하여 실제로 달라진 셀에 FILL_REWARD 적용.

    - ws_new       : 보상값이 이미 업데이트된 새 워크시트
    - ref_snapshot : 참조 탭(복사 원본) 스냅샷 {(row, col): value}
    - 반환: [(coord, ref_value, new_value), ...]
    """
    _SKIP = {"보상 아이템", "보상 수량", "확률", "획득 보상"}

    new_pairs = _find_reward_col_pairs(ws_new)
    reward_coords: set[tuple[int, int]] = set()

    for h_row, item_col, qty_col in new_pairs:
        for r in range(h_row + 1, ws_new.max_row + 1):
            iv = ws_new.cell(row=r, column=item_col).value
            qv = ws_new.cell(row=r, column=qty_col).value
            if iv is None:
                break
            if iv not in _SKIP:
                reward_coords.add((r, item_col))
            if qv is not None and qv not in _SKIP:
                reward_coords.add((r, qty_col))

    hits = []
    for (r, c) in sorted(reward_coords):
        cell    = ws_new.cell(row=r, column=c)
        ref_val = ref_snapshot.get((r, c))
        new_val = cell.value
        if ref_val == new_val:
            continue
        if ref_val is None or new_val is None:
            continue
        if isinstance(new_val, (datetime, date)) or isinstance(ref_val, (datetime, date)):
            continue
        cell.fill = FILL_REWARD
        hits.append((cell.coordinate, ref_val, new_val))

    return hits


def _safe_print(text: str) -> None:
    """cp949 인코딩 불가 문자를 '?'로 대체하여 출력."""
    print(text.encode("cp949", errors="replace").decode("cp949"))


def warn_season_keywords(ws, tab_name):
    """시즌·월 키워드가 남아있는 셀을 경고로 출력. 반환값은 (coord, val) 리스트."""
    hits = []
    for row in ws.iter_rows():
        for cell in row:
            if not cell.value or not isinstance(cell.value, str):
                continue
            for kw in SEASON_KEYWORDS:
                if kw in cell.value:
                    hits.append((cell.coordinate, cell.value[:100]))
                    break
    if hits:
        _safe_print(f"\n  [경고] [{tab_name}] 시즌·월 키워드 포함 셀 - 이벤트 명칭 확인 권장:")
        for coord, val in hits:
            _safe_print(f"       {coord}: '{val}'")
    return hits


def main():
    OUTPUT_DIR.mkdir(exist_ok=True)
    OUTPUT_FILE_DIR.mkdir(exist_ok=True)
    OUTPUT_JSON_DIR.mkdir(exist_ok=True)

    print("[작업 시작]")
    print(f"  소스: {SOURCE}")

    # event_names_config.json 로드 (Claude가 생성해둔 경우)
    event_name_cfg = load_event_names_config()
    if event_name_cfg:
        print(f"  이벤트 명칭 config 적용: {list(event_name_cfg.keys())} 탭")
        for tab_name, repls in event_name_cfg.items():
            if tab_name in UPDATES:
                UPDATES[tab_name]["event_name_replacements"] = repls
    else:
        print("  event_names_config.json 없음 - 이벤트 명칭은 날짜/패턴 치환만 적용")

    print("  파일 로드 중...")
    wb = openpyxl.load_workbook(SOURCE)

    # 필요한 소스 탭만 남기기
    sheets_needed = {cfg["source_tab"] for cfg in UPDATES.values()}
    for name in list(wb.sheetnames):
        if name not in sheets_needed:
            wb.remove(wb[name])
    print(f"  남은 시트: {wb.sheetnames}")

    all_changes = {}
    all_season_warnings = {}

    for new_tab, cfg in UPDATES.items():
        src = cfg["source_tab"]
        if src not in wb.sheetnames:
            print(f"\n  [SKIP] 소스 탭 '{src}' 없음 → {new_tab} 건너뜀")
            continue

        ws = wb[src]
        changes = apply_replacements(
            ws,
            cfg["replacements"],
            date_map=cfg.get("date_map"),
            event_name_replacements=cfg.get("event_name_replacements"),
        )
        ws.title = new_tab
        all_changes[new_tab] = changes

        print(f"\n  [{new_tab}] {src} 기반, {len(changes)}개 셀 갱신")
        for coord, old, new in changes:
            print(f"    {coord}: '{old}' -> '{new}'")

        season_hits = warn_season_keywords(ws, new_tab)
        all_season_warnings[new_tab] = season_hits

    # 탭 순서 정렬 (오름차순): 역순으로 각 탭을 앞으로 이동
    for tab in reversed(sorted(UPDATES.keys())):
        if tab in wb.sheetnames:
            wb.move_sheet(tab, offset=-wb.sheetnames.index(tab))

    wb.save(OUTPUT_FILE)

    print(f"\n완료!")
    print(f"  생성 파일: {OUTPUT_FILE}")
    print(f"  탭 순서: {wb.sheetnames}")

    # 경고가 남은 탭 안내
    remaining = {t: hits for t, hits in all_season_warnings.items() if hits}
    if remaining:
        print(
            "\n  ※ 위 경고 셀의 이벤트 명칭을 갱신하려면:\n"
            "     1) Claude에게 '이벤트 명칭 자동 갱신' 요청\n"
            "     2) 또는 UPDATES['{탭명}']['event_name_replacements'] 에 직접 추가 후 재실행"
        )
    else:
        print("  [OK] 시즌·월 키워드 경고 없음")


def run_with_config(
    source_path: str,
    output_path: str,
    updates: dict,
    event_name_cfg: dict | None = None,
) -> dict:
    """Streamlit/외부 직접 호출용. 결과 dict + xlsx 바이트 반환."""
    import copy

    updates = copy.deepcopy(updates)

    if event_name_cfg:
        for tab_name, repls in event_name_cfg.get("event_name_replacements", {}).items():
            if tab_name in updates:
                updates[tab_name]["event_name_replacements"] = [tuple(r) for r in repls]

    wb = openpyxl.load_workbook(source_path)

    sheets_needed = {cfg["source_tab"] for cfg in updates.values()}
    for name in list(wb.sheetnames):
        if name not in sheets_needed:
            wb.remove(wb[name])

    all_changes: dict = {}
    all_season_warnings: dict = {}

    for new_tab, cfg in updates.items():
        src = cfg["source_tab"]
        if src not in wb.sheetnames:
            raise ValueError(f"소스 탭 '{src}' 없음. 사용 가능: {wb.sheetnames}")
        ws = wb[src]
        ref_pairs = _find_reward_col_pairs(ws)
        ref_snap  = snapshot_ws(ws)
        changes = apply_replacements(
            ws,
            cfg["replacements"],
            date_map=cfg.get("date_map"),
            event_name_replacements=cfg.get("event_name_replacements"),
        )
        ws.title = new_tab
        # 역사 패턴 분석 → 균형 보상 적용 → 참조 대비 하이라이트
        prev_list = _find_all_prev_same_type_ws(wb, src, ref_pairs)
        if prev_list:
            history = [(src, ws)] + prev_list
            apply_balanced_rewards(ws, history)
            highlight_reward_diffs(ws, ref_snap)
        all_changes[new_tab] = changes
        all_season_warnings[new_tab] = warn_season_keywords(ws, new_tab)

    for tab in reversed(sorted(updates.keys())):
        if tab in wb.sheetnames:
            wb.move_sheet(tab, offset=-wb.sheetnames.index(tab))

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    buf = io.BytesIO()
    wb.save(buf)

    result = {
        "output_path": output_path,
        "xlsx_bytes": buf.getvalue(),
        "tabs": list(wb.sheetnames),
        "changes": all_changes,
        "season_warnings": all_season_warnings,
    }

    # last_run_result.json 저장 (save_learning.py 용)
    try:
        run_log = {
            "output_path": output_path,
            "tabs": result["tabs"],
            "changes": {
                tab: [list(c) for c in chgs]
                for tab, chgs in all_changes.items()
            },
            "season_warnings": {
                tab: [list(w) for w in warns]
                for tab, warns in all_season_warnings.items()
            },
        }
        log_path = OUTPUT_JSON_DIR / "last_run_result.json"
        log_path.write_text(
            json.dumps(run_log, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    except Exception:
        pass  # 로그 저장 실패는 메인 흐름에 영향 없음

    return result


def _auto_date_map(ref_tab: str, new_tab: str) -> dict:
    """YYMMDD 또는 MMDD 탭명으로부터 날짜 시프트 map 자동 생성."""
    from datetime import date as _date, timedelta

    def _parse(tab: str) -> "_date | None":
        tab = tab.strip()
        try:
            if len(tab) == 6:
                return _date(2000 + int(tab[:2]), int(tab[2:4]), int(tab[4:6]))
            if len(tab) == 4:
                return _date(_date.today().year, int(tab[:2]), int(tab[2:4]))
        except ValueError:
            pass
        return None

    ref_date = _parse(ref_tab)
    new_date = _parse(new_tab)
    if not ref_date or not new_date:
        return {}
    delta = new_date - ref_date
    if delta.days == 0:
        return {}

    date_map = {}
    for i in range(-7, 61):
        src = ref_date + timedelta(days=i)
        dst = src + delta
        s = f"{src.month:02d}/{src.day:02d}"
        d = f"{dst.month:02d}/{dst.day:02d}"
        if s != d:
            date_map[s] = d
    return date_map


def _best_ref_tab(new_tab: str, existing: list) -> str:
    """new_tab 직전에 가장 가까운 날짜의 기존 탭 반환."""
    from datetime import date as _date
    def _parse(t):
        try:
            return _date(2000 + int(t[:2]), int(t[2:4]), int(t[4:6]))
        except Exception:
            return None
    nd = _parse(new_tab)
    if nd is None or not existing:
        return existing[-1] if existing else new_tab
    # new_tab 날짜보다 이전인 탭 중 가장 가까운 것
    candidates = [(abs((_parse(e) - nd).days), e) for e in existing if _parse(e) and _parse(e) < nd]
    if candidates:
        return min(candidates)[1]
    return existing[0]


def _b_val_to_etype(val: str) -> str:
    """B열 raw 값을 이벤트 유형명으로 변환. 'parse_section_title + detect_event_type' 사용."""
    try:
        from generate_event_names import parse_section_title, detect_event_type
        title = parse_section_title(val)
        if title:
            return detect_event_type(title)
    except Exception:
        pass
    return "기타"


def _remove_event_sections(ws, event_types: list):
    """
    워크시트에서 지정한 이벤트 타입에 해당하는 행 구간을 삭제.
    B열 값을 parse_section_title + detect_event_type으로 유형 변환 후 매칭한다.
    섹션 경계는 '이벤트 제목 :' 패턴이 있는 행 기준으로 구분한다.
    """
    if not event_types:
        return 0

    etypes_set = set(event_types)
    max_row = ws.max_row

    # B열 스캔: '이벤트 제목 :' 패턴이 있는 행만 섹션 시작 후보로 수집
    # → parse_section_title이 None이 아닌 행 (탭 헤더 제외, 진행기간/내용 등 제외)
    section_rows = []  # (start_row, etype, raw_val)

    for row in range(1, max_row + 1):
        v = ws.cell(row=row, column=2).value
        if not v or not isinstance(v, str):
            continue
        raw = v.strip()
        etype = _b_val_to_etype(raw)
        if etype != "기타":
            section_rows.append((row, etype, raw))

    print(f"  [섹션 탐지] 총 {len(section_rows)}개 섹션: {[(r, e) for r, e, _ in section_rows]}")

    # 제거 대상 범위 계산 (역순 삭제)
    rows_to_delete = []
    for i, (start_row, etype, raw) in enumerate(section_rows):
        if etype not in etypes_set:
            continue
        end_row = section_rows[i + 1][0] - 1 if i + 1 < len(section_rows) else max_row
        rows_to_delete.append((start_row, end_row, etype))
        print(f"  [이벤트 제거 예정] '{etype}' ({raw[:30]}) : {start_row}~{end_row}행")

    # 역순 삭제 (행 번호 밀림 방지)
    deleted = 0
    for start_row, end_row, etype in reversed(rows_to_delete):
        count = end_row - start_row + 1
        ws.delete_rows(start_row, count)
        deleted += count
        print(f"  [이벤트 제거 완료] '{etype}': {count}행 삭제")

    return deleted


def _find_event_section_in_history(wb, source_tab: str, event_type: str):
    """
    소스 탭 이전의 이력 탭들에서 event_type 섹션 행들을 찾아 반환.
    B열 값을 parse_section_title + detect_event_type으로 유형 변환 후 매칭한다.
    반환: (ws, start_row, end_row, tab_name) 또는 None
    """
    import re as _re2
    # source_tab을 포함한 모든 탭 검색 (역순 — 최신부터)
    date_tabs = sorted([s for s in wb.sheetnames if _re2.fullmatch(r"\d{6}", s)], reverse=True)
    try:
        src_idx = date_tabs.index(source_tab)
        candidates = date_tabs[src_idx + 1:]  # source_tab 이전(오래된) 탭만
    except ValueError:
        candidates = date_tabs

    for tab in candidates:
        try:
            ws = wb[tab]
        except Exception:
            continue

        # 섹션 시작 행 수집 (이벤트 유형으로 변환)
        section_rows = []
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=2).value
            if not v or not isinstance(v, str):
                continue
            raw = v.strip()
            etype = _b_val_to_etype(raw)
            if etype != "기타":
                section_rows.append((row, etype))

        for i, (start_row, etype) in enumerate(section_rows):
            if etype == event_type:
                end_row = section_rows[i + 1][0] - 1 if i + 1 < len(section_rows) else ws.max_row
                print(f"  [이력 탐색] '{event_type}' → {tab}탭 {start_row}~{end_row}행 발견")
                return ws, start_row, end_row, tab

    print(f"  [이력 탐색] '{event_type}' → 모든 이력 탭에서 찾지 못함")
    return None


def _fix_coupon_end_date(ws, row_start: int, row_end: int, tab: str):
    """
    쿠폰이벤트 섹션의 진행 기간 종료일을 tab이 속한 달의 말일로 덮어쓴다.
    예) tab='260625' → 월=6 → 말일=30 → 종료일을 06/30으로 교체
    패턴: '∎ 진행 기간 : MM/DD(...) HH:MM ~ MM/DD(...) HH:MM:SS ...'
    """
    import re as _re
    import calendar as _cal

    # tab에서 연/월 파싱
    try:
        year  = 2000 + int(tab[:2])
        month = int(tab[2:4])
    except Exception:
        return

    last_day = _cal.monthrange(year, month)[1]
    new_end_mmdd = f"{month:02d}/{last_day:02d}"

    # 진행 기간 행 탐색 (B열 또는 C열에 '진행 기간' 포함)
    _period_pat = _re.compile(
        r'(.*~\s*)(\d{1,2}/\d{1,2})(.*)',  # ~ 이후 MM/DD 캡처
        _re.DOTALL
    )

    for row in range(row_start, row_end + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            v = cell.value
            if not isinstance(v, str):
                continue
            if "진행 기간" not in v:
                continue
            m = _period_pat.search(v)
            if not m:
                continue
            old_end_mmdd = f"{int(m.group(2).split('/')[0]):02d}/{int(m.group(2).split('/')[1]):02d}"
            if old_end_mmdd == new_end_mmdd:
                continue  # 이미 말일이면 스킵
            new_v = v[:m.start(2)] + new_end_mmdd + v[m.end(2):]
            cell.value = new_v
            print(f"  [쿠폰 말일 적용] row{row} col{col}: '{old_end_mmdd}' → '{new_end_mmdd}' (월 말일)")


def _add_event_sections(ws_new, wb_src, source_tab: str, event_types: list,
                        new_tab: str = ""):
    """
    이력 탭에서 event_types 섹션을 찾아 ws_new 끝에 추가.
    복사 후 from_tab → new_tab 날짜 시프트 맵으로 날짜/텍스트를 자동 갱신한다.
    """
    if not event_types:
        return 0

    from copy import copy as _copy2

    added = 0
    for etype in event_types:
        result = _find_event_section_in_history(wb_src, source_tab, etype)
        if result is None:
            print(f"  [이벤트 추가 스킵] '{etype}' 섹션을 이력에서 찾을 수 없음")
            continue

        ws_hist, start_row, end_row, from_tab = result
        dest_start = ws_new.max_row + 1

        # ── 셀 복사 (값 + 스타일) ────────────────────────────────────────
        for src_row in range(start_row, end_row + 1):
            dest_row = dest_start + (src_row - start_row)
            for col in range(1, ws_hist.max_column + 1):
                src_cell  = ws_hist.cell(row=src_row, column=col)
                dest_cell = ws_new.cell(row=dest_row, column=col)
                dest_cell.value = src_cell.value
                if src_cell.has_style:
                    dest_cell.font         = _copy2(src_cell.font)
                    dest_cell.fill         = _copy2(src_cell.fill)
                    dest_cell.border       = _copy2(src_cell.border)
                    dest_cell.alignment    = _copy2(src_cell.alignment)
                    dest_cell.number_format = src_cell.number_format

        count = end_row - start_row + 1
        print(f"  [이벤트 추가] '{etype}' 섹션: {from_tab}탭에서 {count}행 복사 → {dest_start}행~")

        # ── 날짜 시프트: from_tab → new_tab 날짜 맵 생성 후 복사된 행에 적용 ──
        date_target = new_tab if new_tab else source_tab
        date_map = _auto_date_map(from_tab, date_target)
        if date_map:
            _apply_date_map_to_rows(ws_new, dest_start, dest_start + count - 1, date_map)
            print(f"  [날짜 갱신] '{etype}': {from_tab} → {date_target} ({len(date_map)}개 맵핑 적용)")
        else:
            print(f"  [날짜 갱신 스킵] '{etype}': {from_tab} → {date_target} 날짜 맵 없음")

        # ── 이벤트 유형별 특수 날짜 규칙 적용 ───────────────────────────────
        if etype == "쿠폰이벤트":
            _fix_coupon_end_date(ws_new, dest_start, dest_start + count - 1, date_target)

        added += count

    return added


def _apply_date_map_to_rows(ws, row_start: int, row_end: int, date_map: dict):
    """
    ws의 row_start~row_end 행에만 날짜 시프트를 적용.
    - datetime/date 객체: date_map으로 직접 변환
    - str: date_map의 MM/DD 패턴을 찾아 치환
    """
    import re as _re
    from datetime import datetime as _dt, date as _d

    # 날짜 문자열 패턴 (MM/DD, M/DD, MM/D 등 포함)
    _date_pat = _re.compile(r'\b(\d{1,2})/(\d{1,2})\b')

    def _shift_str(s: str) -> str:
        def _repl(m):
            key = f"{int(m.group(1)):02d}/{int(m.group(2)):02d}"
            return m.group(0).replace(
                m.group(0),
                date_map[key] if key in date_map else m.group(0)
            )
        return _date_pat.sub(_repl, s)

    for row in range(row_start, row_end + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                continue
            if isinstance(cell.value, (_dt, _d)):
                old = cell.value
                mmdd = old.strftime("%m/%d")
                if mmdd in date_map:
                    nm, nd = map(int, date_map[mmdd].split("/"))
                    if isinstance(old, _dt):
                        cell.value = _dt(old.year, nm, nd, old.hour, old.minute, old.second)
                    else:
                        cell.value = _d(old.year, nm, nd)
            elif isinstance(cell.value, str):
                new_val = _shift_str(cell.value)
                if new_val != cell.value:
                    cell.value = new_val


def _cli_main(source: str, output: str, new_tabs: list, ref_tabs: list,
              events_to_remove: list = None, events_to_add: list = None):
    """서버 파이프라인에서 CLI 인수로 호출되는 진입점."""
    import re as _re, copy as _copy
    wb_tmp = openpyxl.load_workbook(source, read_only=True)
    existing_date_tabs = sorted([s for s in wb_tmp.sheetnames if _re.match(r'^\d{6}$', s)])
    all_sheetnames    = wb_tmp.sheetnames[:]
    wb_tmp.close()

    # ref_tabs가 지정되지 않으면 각 new_tab에 가장 적합한 기존 탭 자동 선택
    if not ref_tabs:
        ref_tabs = [_best_ref_tab(t, existing_date_tabs) for t in new_tabs]

    updates = {}
    for i, new_tab in enumerate(new_tabs):
        ref = ref_tabs[i] if i < len(ref_tabs) else (ref_tabs[-1] if ref_tabs else new_tab)
        updates[new_tab] = {
            "source_tab": ref,
            "replacements": [],
            "date_map": _auto_date_map(ref, new_tab),
        }

    print(f"[CLI 모드] 소스: {source}")
    for t, cfg in updates.items():
        print(f"  [{t}] 참조: {cfg['source_tab']}  date_map {len(cfg['date_map'])}개")

    # 같은 source_tab이 여러 new_tab에 쓰이면 중복 처리: 직접 workbook 조작
    wb = openpyxl.load_workbook(source)
    all_changes = {}
    all_season_warnings = {}
    # load_event_names_config() 는 이미 {tab: [(old,new),...]} 형태로 반환
    enr = load_event_names_config() or {}

    for new_tab, cfg in updates.items():
        src = cfg["source_tab"]
        if src not in wb.sheetnames:
            print(f"  [SKIP] 소스 탭 '{src}' 없음 → {new_tab} 건너뜀")
            continue
        # 소스 시트를 복사해서 새 이름으로 추가
        ws_src = wb[src]

        ws_new = wb.copy_worksheet(ws_src)
        ws_new.title = new_tab

        event_repls = enr.get(new_tab, [])
        changes = apply_replacements(
            ws_new,
            cfg.get("replacements", []),
            date_map=cfg.get("date_map"),
            event_name_replacements=event_repls,
        )

        # ── 보상 처리: 역사 패턴 분석 → 균형 보상 적용 → 참조 대비 하이라이트 ───
        ref_pairs  = _find_reward_col_pairs(ws_src)
        ref_snap   = snapshot_ws(ws_src)   # 참조 탭(260611) 스냅샷 — 하이라이트 기준
        prev_list  = _find_all_prev_same_type_ws(wb, src, ref_pairs)  # [(탭,ws), ...]

        if prev_list:
            # 참조 탭을 첫 번째로 포함 → 현재 수준을 기준점으로 가중 평균 계산
            history = [(src, ws_src)] + prev_list
            tab_names = [src] + [t for t, _ in prev_list]
            applied = apply_balanced_rewards(ws_new, history)
            print(f"  [{new_tab}] 보상 패턴 분석: {tab_names} → 균형 조정 {applied}개 행")
            # 적용 후 참조 탭(260611) 대비 달라진 보상 셀만 하이라이트
            reward_hits = highlight_reward_diffs(ws_new, ref_snap)
            if reward_hits:
                print(f"  [{new_tab}] 보상 변경 셀 {len(reward_hits)}개 하이라이트 (vs {src}):")
                for coord, old, new_v in reward_hits[:10]:
                    print(f"    {coord}: '{old}' → '{new_v}'")
                if len(reward_hits) > 10:
                    print(f"    ... 외 {len(reward_hits)-10}개")
        else:
            print(f"  [{new_tab}] 이전 동일 타입 탭 없음 → 보상 변경 생략")

        # ── 이벤트 구성 조정 ────────────────────────────────────────────────
        if events_to_remove:
            removed = _remove_event_sections(ws_new, events_to_remove)
            print(f"  [{new_tab}] 이벤트 섹션 제거: {removed}행")
        if events_to_add:
            added = _add_event_sections(ws_new, wb, src, events_to_add, new_tab=new_tab)
            print(f"  [{new_tab}] 이벤트 섹션 추가: {added}행")

        all_changes[new_tab] = changes
        all_season_warnings[new_tab] = warn_season_keywords(ws_new, new_tab)
        print(f"  [{new_tab}] {src} 복사 완료, {len(changes)}개 셀 갱신")

    # 불필요한 기존 탭 제거 (새로 만든 탭만 남김)
    for name in list(wb.sheetnames):
        if name not in updates:
            del wb[name]

    # 탭 순서 정렬
    for tab in reversed(sorted(updates.keys())):
        if tab in wb.sheetnames:
            wb.move_sheet(tab, offset=-wb.sheetnames.index(tab))

    Path(output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(f"완료! 생성 파일: {output}")
    print(f"탭 목록: {wb.sheetnames}")


if __name__ == "__main__":
    # Windows cp949 콘솔에서 utf-8 출력 강제
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

    # CLI 인수 있으면 동적 모드, 없으면 하드코딩 UPDATES 사용
    if len(sys.argv) >= 4:
        import argparse as _ap
        _parser = _ap.ArgumentParser(add_help=False)
        _parser.add_argument("source")
        _parser.add_argument("output")
        _parser.add_argument("new_tabs")
        _parser.add_argument("ref_tabs", nargs="?", default="")
        _parser.add_argument("--remove-events", default="")
        _parser.add_argument("--add-events", default="")
        _args, _ = _parser.parse_known_args()

        _src      = _args.source
        _out      = _args.output
        _new_tabs = [t.strip() for t in _args.new_tabs.split(",") if t.strip()]
        _ref_tabs = [t.strip() for t in _args.ref_tabs.split(",") if t.strip()]
        _remove   = [t.strip() for t in _args.remove_events.split(",") if t.strip()]
        _add      = [t.strip() for t in _args.add_events.split(",") if t.strip()]
        _cli_main(_src, _out, _new_tabs, _ref_tabs, events_to_remove=_remove, events_to_add=_add)
    else:
        main()
