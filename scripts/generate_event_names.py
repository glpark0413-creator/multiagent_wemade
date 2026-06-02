#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
이벤트 명칭 자동 생성기

historical_event_names.json (학습 데이터)와 schedule_patterns.json (시즌 컨텍스트)를
읽어서, 참조 탭의 현재 이벤트 제목을 대상 월에 맞게 자동 치환하고
event_names_config.json 의 event_name_replacements 를 채운다.

create_tabs.py 가 이 파일을 읽어 실제 셀 치환을 수행한다.

동작 흐름:
  1. 소스 xlsx 의 참조 탭에서 이벤트 섹션 제목 추출 (현재 이름)
  2. historical_event_names.json 에서 월별 시즌 키워드 로드
  3. 대상 탭 날짜로부터 목표 연월 결정
  4. 각 제목에서 이전 시즌 키워드 → 새 시즌 키워드 치환 쌍 생성
  5. 사용자가 입력한 genre_phrases 로 추가 제목 변형 시도
  6. event_names_config.json 에 결과 병합 저장

사용:
  python scripts/generate_event_names.py \\
      --source "path/to/source.xlsx" \\
      --new-tabs "260709,260716" \\
      --ref-tabs "260625,260702" \\
      --target-month "2026-07" \\
      --genre "야구" \\
      --phrases "7월의,여름,올스타,홈런"
"""
import argparse
import io
import json
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

# ─── 경로 설정 ────────────────────────────────────────────────────────────────
_BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _project_config import load_project_paths

_paths = load_project_paths()

# ─── 섹션 제목 추출 패턴 ─────────────────────────────────────────────────────
SECTION_RE       = re.compile(r"^\d+\.")
TITLE_RE         = re.compile(r"^이벤트\s*제목\s*:\s*(.+)$")
TAB_HEADER_RE    = re.compile(r"^\d{1,2}[./-]\d{1,2}[._\s]")   # "06.11_ Event" 등 제외

# 월/시즌 키워드 감지 패턴 — 제목에서 교체 대상이 되는 부분
# 우선순위 순서 (긴 패턴 먼저, 뒤따르는 "의"까지 포함해서 통째로 매치)
SEASON_PATTERNS = [
    # "N월 이달의" ("6월 이달의" → "7월 이달의")
    re.compile(r"\d+월\s*이달의"),
    # "N월의" (숫자+월의)
    re.compile(r"\d+월의"),
    # "N월 " (후행 공백 포함, "이달의"가 없는 형식, 예: "6월 축제")
    re.compile(r"(\d+월)\s(?=[가-힣])"),
    # 얼리썸머의?, 초여름의?, 쿨 서머의?, 한여름의?, 늦여름의?
    re.compile(r"(얼리썸머|초여름|쿨\s*서머|한여름|늦여름)(의)?"),
    # 봄/여름/가을/겨울 + "의" 선택적 (버그수정: "의" 없어도 매치)
    re.compile(r"(봄|여름|가을|겨울)(의)?(?=[\s의이!,.]|$)"),
    # 설날/크리스마스/추석/핼러윈
    re.compile(r"(설날|크리스마스|추석|핼러윈)"),
    # 전반기/후반기/시즌 — 야구 시즌 용어 확장
    re.compile(r"(전반기|후반기|포스트시즌|개막\s*시즌|정규\s*시즌|올스타|드래프트)"),
    # 주년
    re.compile(r"\d+주년"),
]

# 월 번호 → 대표 한국어 표현 (수동 설정)
MONTH_LABEL_KR: dict[int, str] = {
    1:  "1월의",
    2:  "2월의",
    3:  "봄의",
    4:  "봄의",
    5:  "5월의",
    6:  "초여름",
    7:  "7월의",
    8:  "여름의",
    9:  "가을의",
    10: "10월의",
    11: "11월의",
    12: "크리스마스",
}

# 月별 대체 후보 우선순위 (season_keywords_by_month 에 없을 때 fallback)
MONTH_SEASON_FALLBACK: dict[int, list[str]] = {
    1:  ["1월의", "신년", "새해"],
    2:  ["2월의", "봄의"],
    3:  ["봄의", "3월의"],
    4:  ["봄의", "4월의"],
    5:  ["5월의", "황금연휴"],
    6:  ["얼리썸머", "초여름", "6월의"],
    7:  ["7월의", "여름의", "한여름"],
    8:  ["여름의", "한여름", "8월의"],
    9:  ["가을의", "추석", "9월의"],
    10: ["10월의", "가을의", "핼러윈"],
    11: ["11월의", "가을의"],
    12: ["크리스마스", "연말", "12월의"],
}


# ─── 유틸 함수 ────────────────────────────────────────────────────────────────

def parse_section_title(val: str) -> str | None:
    if TAB_HEADER_RE.match(val):   # "06.11_ Event" 형식 탭 헤더 → 제외
        return None
    if SECTION_RE.match(val):
        return val
    m = TITLE_RE.match(val)
    return m.group(1).strip() if m else None


def tab_to_date(tab: str) -> datetime | None:
    if not re.fullmatch(r"\d{6}", tab):
        return None
    try:
        return datetime.strptime("20" + tab, "%Y%m%d")
    except ValueError:
        return None


def extract_section_titles(wb, tab_name: str) -> list[str]:
    """워크시트 B열에서 이벤트 섹션 제목 추출."""
    if tab_name not in wb.sheetnames:
        return []
    ws = wb[tab_name]
    titles = []
    for row in ws.iter_rows():
        for cell in row:
            if getattr(cell, "column", None) != 2:
                continue
            val = getattr(cell, "value", None)
            if not val:
                continue
            t = parse_section_title(str(val).strip())
            if t is not None:
                titles.append(t)
    return titles


def detect_season_keyword(title: str) -> tuple[str, re.Pattern] | None:
    """제목에서 첫 번째로 매치되는 시즌 키워드 전체 문자열과 패턴 반환."""
    for pat in SEASON_PATTERNS:
        m = pat.search(title)
        if m:
            return m.group(0), pat
    return None


def _has_trailing_eui(old_kw: str) -> bool:
    """매치된 키워드가 이미 '의'로 끝나는지 확인."""
    return old_kw.endswith("의")


def pick_best_season_kw(
    target_month:  int,
    learned_kws:   list[str],
    genre_phrases: list[str],
    old_kw:        str = "",
) -> str:
    """
    대상 월에 가장 적합한 시즌 키워드를 선택한다.
    old_kw 가 "의"로 끝나면 new_kw 도 "의"로 끝나도록 맞춘다.
    우선순위: genre_phrases 에서 시즌 패턴 매치 > learned_kws > fallback
    """
    needs_eui = _has_trailing_eui(old_kw)

    # "N월 이달의" 패턴 처리 — 숫자만 교체하면 됨
    num_month_re = re.compile(r"^(\d+)(월\s*이달의)$")
    m = num_month_re.match(old_kw)
    if m:
        return f"{target_month}{m.group(2)}"

    # genre_phrases 에서 시즌 패턴 매치
    for phrase in genre_phrases:
        for pat in SEASON_PATTERNS:
            if pat.search(phrase):
                kw = phrase
                if needs_eui and not kw.endswith("의"):
                    kw += "의"
                elif not needs_eui and kw.endswith("의"):
                    kw = kw[:-1]
                return kw

    # 학습된 키워드에서 대상 월과 가장 관련된 것
    for kw in learned_kws:
        if re.match(rf"^{target_month}월", kw):
            if needs_eui and not kw.endswith("의"):
                kw += "의"
            return kw

    # fallback
    fb = MONTH_SEASON_FALLBACK.get(target_month, [])
    base = fb[0] if fb else MONTH_LABEL_KR.get(target_month, f"{target_month}월의")
    if needs_eui and not base.endswith("의"):
        base += "의"
    elif not needs_eui and base.endswith("의"):
        base = base[:-1]
    return base


# ─── 이벤트 유형 분류 ────────────────────────────────────────────────────────

# 이벤트 유형 키워드 맵 (유형명 → 감지 키워드 리스트)
EVENT_TYPE_MAP: dict[str, list[str]] = {
    "출석이벤트":    ["출석"],
    "응모권이벤트":  ["응모권"],
    "미션이벤트":    ["플레이 미션", "미션 이벤트"],
    "교환소이벤트":  ["교환소"],
    "오더경쟁이벤트":["오더 경쟁", "오더경쟁"],
    "PvP이벤트":    ["pvp", "핫타임"],
    "포인트레이스":  ["포인트 레이스", "포인트레이스"],
    "룰렛이벤트":    ["룰렛"],
    "빙고이벤트":    ["빙고"],
    "쿠폰이벤트":    ["쿠폰"],
    "야구공찾기":    ["야구공 찾기", "야구공찾기"],
    "승부예측":      ["승부 예측", "승부예측"],
}

# 절대 변경하지 않는 고정 이벤트 유형 (A타입)
FIXED_EVENT_TYPES: set[str] = {"포인트레이스", "룰렛이벤트", "빙고이벤트"}

# 이벤트 유형별 고정 후미 패턴 (새 제목 생성 시 보존할 부분)
EVENT_SUFFIX_PATTERNS: dict[str, re.Pattern] = {
    "출석이벤트":    re.compile(r"(\d+일\s*출석\s*이벤트\S*)"),
    "응모권이벤트":  re.compile(r"(응모권\s*이벤트\S*)"),
    "미션이벤트":    re.compile(r"(플레이\s*미션\s*이벤트\S*)"),
    "교환소이벤트":  re.compile(r"(교환소\s*이벤트\S*)"),
    "오더경쟁이벤트":re.compile(r"(오더\s*경쟁\s*이벤트\S*)"),
    "야구공찾기":    re.compile(r"(야구공\s*찾기\s*이벤트\S*)"),
    "PvP이벤트":    re.compile(r"(pvp.*이벤트\S*)", re.IGNORECASE),
    "응모권이벤트":  re.compile(r"(응모권\s*이벤트\S*)"),
}


def detect_event_type(title: str) -> str:
    """이벤트 제목에서 이벤트 유형 감지."""
    tl = title.lower().replace(" ", "")
    for etype, keywords in EVENT_TYPE_MAP.items():
        if any(kw.replace(" ", "").lower() in tl for kw in keywords):
            return etype
    return "기타"


def analyze_title_stability(wb) -> tuple[set, set]:
    """
    이벤트 유형 기반 STABLE / CHANGEABLE 분류.

    - 같은 이벤트 유형이 탭마다 다른 제목 → CHANGEABLE
    - 고정 이벤트 유형(포인트레이스/룰렛/빙고) → 항상 STABLE
    - 단 하나의 탭에만 있거나 항상 동일 → STABLE

    Returns:
        stable_titles    — 변경하지 않을 제목 집합
        changeable_types — 변경 대상 이벤트 유형 집합
    """
    date_tabs = sorted([s for s in wb.sheetnames if re.fullmatch(r"\d{6}", s)])
    if not date_tabs:
        return set(), set()

    # event_type → set of unique titles seen
    type_to_titles: dict[str, set] = {}
    title_to_type:  dict[str, str] = {}

    for tab in date_tabs:
        for title in extract_section_titles(wb, tab):
            etype = detect_event_type(title)
            # 숫자 접두사 제거 후 비교 ("6. PvP ..." == "PvP ..." 동일 취급)
            normalized = re.sub(r"^\d+\.\s*", "", title).strip()
            type_to_titles.setdefault(etype, set()).add(normalized)
            title_to_type[title] = etype

    stable_titles:    set[str] = set()
    changeable_types: set[str] = set()

    for etype, titles in type_to_titles.items():
        if etype in FIXED_EVENT_TYPES:
            stable_titles.update(titles)        # 고정 이벤트 → 모두 STABLE
        elif len(titles) == 1:
            stable_titles.update(titles)        # 변형 없음 → STABLE
        else:
            changeable_types.add(etype)         # 변형 있음 → CHANGEABLE

    print(f"  고정(STABLE) 이벤트 유형: {sorted(FIXED_EVENT_TYPES & set(type_to_titles))}")
    print(f"  변경(CHANGEABLE) 이벤트 유형: {sorted(changeable_types)}")
    return stable_titles, changeable_types


def _generate_creative_title(
    old_title:     str,
    event_type:    str,
    target_m:      int,
    genre_phrases: list[str],
    learned_kws:   list[str],
) -> str | None:
    """
    시즌 키워드가 없는 CHANGEABLE 제목을 위해 새 제목 생성.
    이벤트 유형의 핵심 후미(출석 이벤트 / 응모권 이벤트 등)를 보존하고
    앞에 시즌 키워드를 붙인다.

    예) "1. 벚꽃 엔딩 14일 출석 이벤트!" → "1. 7월의 14일 출석 이벤트!"
    예) "2. 스프링 피날레 챌린지 응모권 이벤트!" → "2. 7월의 응모권 이벤트!"
    """
    season_kw = pick_best_season_kw(target_m, learned_kws, genre_phrases, "의")
    if not season_kw:
        return None

    # 번호 접두사 분리
    num_m = re.match(r"^(\d+\.\s*)(.*)", old_title)
    num_prefix = num_m.group(1) if num_m else ""
    body       = num_m.group(2) if num_m else old_title

    # 이벤트 유형 후미 패턴으로 핵심 부분 추출
    pat = EVENT_SUFFIX_PATTERNS.get(event_type)
    if pat:
        m = pat.search(body)
        if m:
            return f"{num_prefix}{season_kw} {m.group(1)}"

    # fallback: 기존 제목 앞에 시즌 키워드 삽입
    return f"{num_prefix}{season_kw} {body}"


def generate_new_title(
    old_title:       str,
    old_month:       int,
    target_month:    int,
    learned_kws:     list[str],
    genre_phrases:   list[str],
    stable_titles:   set | None = None,
    changeable_types: set | None = None,   # ← 기존 changeable_bases 대체
) -> str | None:
    """
    이벤트 제목을 대상 월에 맞게 교체한다.

    규칙:
    - STABLE 제목 → 무조건 유지
    - CHANGEABLE + 시즌 키워드 감지 → 키워드 교체
    - CHANGEABLE + 시즌 키워드 없음 + 월 다름 → _generate_creative_title로 새 제목 생성
    - 그 외 → 유지
    """
    # ── STABLE 제목 → 항상 유지 ──────────────────────────────────────────────
    if stable_titles and old_title in stable_titles:
        return None

    # ── 이벤트 유형 확인 ─────────────────────────────────────────────────────
    etype = detect_event_type(old_title)
    is_changeable = (changeable_types is None) or (etype in changeable_types)

    if not is_changeable:
        return None   # CHANGEABLE 아닌 이벤트 유형 → 유지

    detected = detect_season_keyword(old_title)

    # ── Case 1: 시즌 키워드 감지됨 → 교체 ──────────────────────────────────
    if detected:
        old_kw, _pat = detected
        new_kw = pick_best_season_kw(target_month, learned_kws, genre_phrases, old_kw)
        if old_kw == new_kw:
            return None
        new_title = old_title.replace(old_kw, new_kw, 1)
        return new_title if new_title != old_title else None

    # ── Case 2: 시즌 키워드 없음 + CHANGEABLE + 월 다름 → 창의적 제목 생성 ──
    if old_month == target_month:
        return None   # 같은 달이면 변경 불필요

    return _generate_creative_title(
        old_title, etype, target_month, genre_phrases, learned_kws
    )


# ─── 메인 ─────────────────────────────────────────────────────────────────────

def generate(
    source_xlsx:   str,
    new_tabs:      list[str],
    ref_tabs:      list[str],
    target_month:  str,        # "YYYY-MM"
    genre:         str = "",
    genre_phrases: list[str] | None = None,
    work_dir:      Path | None = None,
) -> dict:
    """
    학습 데이터를 기반으로 이벤트 제목 치환 쌍을 생성하고
    event_names_config.json 에 저장한다.

    반환: { new_tab: [(old, new), ...], ... }
    """
    genre_phrases = genre_phrases or []

    if work_dir is None:
        work_dir = _paths.work_dir if _paths else _BASE_DIR / "output" / "event-planner" / "work"
    work_dir = Path(work_dir)
    work_dir.mkdir(parents=True, exist_ok=True)

    # ── 학습 데이터 로드 ────────────────────────────────────────────────────
    names_data   = {}
    sched_data   = {}
    names_file   = work_dir / "historical_event_names.json"
    sched_file   = work_dir / "schedule_patterns.json"

    if names_file.exists():
        names_data = json.loads(names_file.read_text(encoding="utf-8"))
    if sched_file.exists():
        sched_data = json.loads(sched_file.read_text(encoding="utf-8"))

    # 월별 시즌 키워드 (학습 결과 우선, 없으면 fallback)
    season_kws_by_month: dict[str, list[str]] = names_data.get("season_keywords_by_month", {})

    # schedule_patterns 의 month_specific 에서도 추출
    for month_key, info in sched_data.get("month_specific", {}).items():
        notes = info.get("notes", "")
        if notes and month_key not in season_kws_by_month:
            # notes 에서 키워드 추출
            kws = re.findall(r"[가-힣]{2,6}(?:의|적|형)?", notes)
            if kws:
                season_kws_by_month[month_key] = kws

    # 대상 월 번호 파싱
    try:
        target_year, target_m = map(int, target_month.split("-"))
    except ValueError:
        target_year = datetime.now().year
        target_m    = datetime.now().month

    target_month_key = f"{target_m:02d}"
    learned_kws_for_target = season_kws_by_month.get(target_month_key, [])

    # ── 소스 xlsx 열기 ──────────────────────────────────────────────────────
    if not Path(source_xlsx).exists():
        print(f"[오류] 소스 파일 없음: {source_xlsx}", file=sys.stderr)
        return {}

    wb = openpyxl.load_workbook(source_xlsx, data_only=True)

    # ── 전체 탭 분석: STABLE / CHANGEABLE 분류 (이벤트 유형 기반) ────────────
    print(f"\n[제목 패턴 분석 — 이벤트 유형 기반]")
    stable_titles, changeable_types = analyze_title_stability(wb)
    print(f"  STABLE  제목 수: {len(stable_titles)}개")
    print(f"  CHANGEABLE 유형: {sorted(changeable_types)}")

    all_replacements: dict[str, list] = {}
    stats: dict[str, int] = {}

    for i, new_tab in enumerate(new_tabs):
        ref_tab = ref_tabs[i] if i < len(ref_tabs) else (ref_tabs[-1] if ref_tabs else "")
        if not ref_tab:
            print(f"  [{new_tab}] 참조 탭 없음 → 건너뜀")
            continue

        # 참조 탭 날짜 → 원본 月 파악
        ref_dt = tab_to_date(ref_tab)
        old_month = ref_dt.month if ref_dt else target_m

        # 참조 탭에서 현재 이벤트 제목 추출
        current_titles = extract_section_titles(wb, ref_tab)
        if not current_titles:
            print(f"  [{new_tab}] 참조 탭 '{ref_tab}' 에서 섹션 제목을 찾을 수 없음")
            all_replacements[new_tab] = []
            continue

        replacements: list[tuple[str, str]] = []
        changed = 0

        for old_title in current_titles:
            new_title = generate_new_title(
                old_title, old_month, target_m,
                learned_kws_for_target, genre_phrases,
                stable_titles=stable_titles,
                changeable_types=changeable_types,
            )
            if new_title:
                replacements.append((old_title, new_title))
                changed += 1
                print(f"  [{new_tab}] 변경: '{old_title}' → '{new_title}'")
            else:
                print(f"  [{new_tab}] 유지: '{old_title}'")

        all_replacements[new_tab] = replacements
        stats[new_tab] = changed

    wb.close()

    # ── event_names_config.json 에 병합 저장 ──────────────────────────────
    config_path = work_dir / "event_names_config.json"
    if config_path.exists():
        try:
            config = json.loads(config_path.read_text(encoding="utf-8"))
        except Exception:
            config = {}
    else:
        config = {}

    config["genre"]         = genre or config.get("genre", "")
    config["target_month"]  = target_month
    config["genre_phrases"] = genre_phrases

    # 기존 replacements 와 병합 (덮어쓰기)
    existing = config.get("event_name_replacements", {})
    existing.update(all_replacements)
    config["event_name_replacements"] = existing
    config["generated_at"] = datetime.now().isoformat(timespec="seconds")

    config_path.write_text(
        json.dumps(config, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"\n[저장] event_names_config.json → {config_path}")
    total_changed = sum(stats.values())
    print(f"  탭 {len(new_tabs)}개, 총 {total_changed}개 제목 치환 쌍 생성")

    return all_replacements


# ─── CLI ─────────────────────────────────────────────────────────────────────

def main():
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

    parser = argparse.ArgumentParser(description="이벤트 명칭 자동 생성")
    parser.add_argument("--source",        required=True, help="소스 xlsx 경로")
    parser.add_argument("--new-tabs",      required=True, help="생성 탭명 (쉼표 구분)")
    parser.add_argument("--ref-tabs",      default="",    help="참조 탭명 (쉼표 구분)")
    parser.add_argument("--target-month",  default=None,  help="대상 연월 YYYY-MM")
    parser.add_argument("--genre",         default="",    help="장르명")
    parser.add_argument("--phrases",       default="",    help="장르 키워드 (쉼표 구분)")
    parser.add_argument("--work-dir",      default=None,  help="작업 디렉터리")
    args = parser.parse_args()

    new_tabs = [t.strip() for t in args.new_tabs.split(",") if t.strip()]
    ref_tabs = [t.strip() for t in args.ref_tabs.split(",") if t.strip()]
    phrases  = [p.strip() for p in args.phrases.split(",") if p.strip()]

    # 대상 월 자동 결정 (new_tabs 첫 탭의 날짜)
    if args.target_month:
        target_month = args.target_month
    elif new_tabs:
        dt = tab_to_date(new_tabs[0])
        target_month = dt.strftime("%Y-%m") if dt else datetime.now().strftime("%Y-%m")
    else:
        target_month = datetime.now().strftime("%Y-%m")

    work_dir = Path(args.work_dir) if args.work_dir else None

    print(f"[이벤트 명칭 생성]")
    print(f"  소스: {args.source}")
    print(f"  대상 탭: {new_tabs}  참조 탭: {ref_tabs}")
    print(f"  대상 월: {target_month}  장르: {args.genre}")

    generate(
        source_xlsx   = args.source,
        new_tabs      = new_tabs,
        ref_tabs      = ref_tabs,
        target_month  = target_month,
        genre         = args.genre,
        genre_phrases = phrases,
        work_dir      = work_dir,
    )


if __name__ == "__main__":
    main()
