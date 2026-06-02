#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Google Sheets URL → 프로젝트별 학습 데이터 크롤러

동작:
  1. Google Sheets URL에서 spreadsheet_id 파싱
  2. xlsx 다운로드 — 공개 시트는 API 키 없이, 비공개 시트는 서비스 계정 사용
  3. 이벤트 명칭·보상 패턴·이벤트 빈도 Python으로 직접 분석 (토큰 최소화)
  4. output/json/projects/{project_id}/ 에 프로젝트별 학습 데이터 저장

다운로드 방식 (자동 결정):
  ① 공개 시트 (링크 있는 모든 사용자 → 뷰어):
     requests 라이브러리로 Export URL 직접 다운로드 — API 키 불필요
     pip install requests

  ② 비공개 시트 (조직 내부 / 특정 사용자 공유):
     서비스 계정으로 Drive API 다운로드
     pip install google-api-python-client google-auth
     credentials/service_account.json 필요 (docs/google_auth_setup.md 참조)

사용:
  # 기본 크롤링 (공개 시트 → API 키 없이 실행)
  python scripts/crawl_gdrive_project.py "https://docs.google.com/spreadsheets/d/..."

  # project_id 직접 지정
  python scripts/crawl_gdrive_project.py "https://docs.google.com/spreadsheets/d/..." --project-id NC_KR

  # 이미 다운로드된 xlsx 재사용 (다운로드 건너뜀)
  python scripts/crawl_gdrive_project.py "https://docs.google.com/spreadsheets/d/..." --skip-download "Readdocs/projects/NC_KR_xxxx.xlsx"

  # 등록된 프로젝트 목록
  python scripts/crawl_gdrive_project.py --list

출력 (output/json/projects/{project_id}/):
  - agent_learning.json        : 누적 학습 데이터 (세션 시작 시 로드용)
  - historical_event_names.json: 탭별 이벤트 섹션 제목
  - reward_by_event.json       : 이벤트 유형별 보상 패턴
  - event_pattern_analysis.json: 이벤트 유형 빈도 분석
"""
import io
import json
import re
import sys
import argparse
import shutil
from collections import defaultdict
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("[오류] openpyxl 필요: pip install openpyxl")
    sys.exit(1)

# requests — 공개 시트 다운로드에 사용 (선택적)
try:
    import requests as _requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

# Google API — 비공개 시트 다운로드에 사용 (선택적)
try:
    from googleapiclient.discovery import build as _gapi_build
    from google.oauth2 import service_account as _sa
    from googleapiclient.http import MediaIoBaseDownload as _MIBD
    HAS_GOOGLE_API = True
except ImportError:
    HAS_GOOGLE_API = False

# ─── 경로 상수 ────────────────────────────────────────────────────────────────
BASE_DIR         = Path(__file__).resolve().parent.parent
CREDENTIALS_PATH = BASE_DIR / "credentials" / "service_account.json"
PROJECTS_DIR     = BASE_DIR / "output" / "projects"
DOWNLOAD_DIR     = BASE_DIR / "Readdocs" / "projects"

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets.readonly",
    "https://www.googleapis.com/auth/drive.readonly",
]

# ─── 패턴 상수 ────────────────────────────────────────────────────────────────
SECTION_PATTERN  = re.compile(r"^\d+\.")
DATE_TAB_PATTERN = re.compile(r"^\d{6}$")

SEASON_RE = re.compile(
    r"(봄의|여름의|가을의|겨울의|새해|신년|설날|추석|크리스마스|핼러윈|"
    r"각성|전설|영웅|봉인|결전|\d+주년|\d+월의|\d+월 )"
)

# 공통 이벤트 유형 키워드 (MMORPG + 스포츠 게임 혼용)
EVENT_TYPE_KEYWORDS = [
    ("출석",          "출석_이벤트"),
    ("로그인",        "출석_이벤트"),
    ("미션",          "미션_이벤트"),
    ("플레이",        "미션_이벤트"),
    ("챌린지",        "미션_이벤트"),
    ("던전",          "던전_이벤트"),
    ("레이드",        "레이드_이벤트"),
    ("패스",          "패스"),
    ("교환 상점",     "교환상점_이벤트"),
    ("교환소",        "교환상점_이벤트"),
    ("보물 상자",     "보물상자_이벤트"),
    ("주사위",        "주사위_이벤트"),
    ("제작",          "제작_이벤트"),
    ("우편",          "우편지급_이벤트"),
    ("할인",          "할인_이벤트"),
    ("성장 가이드",   "성장가이드_이벤트"),
    ("지령",          "지령_이벤트"),
    # 스포츠/야구 게임 유형
    ("응모권",        "응모권_이벤트"),
    ("룰렛",          "응모권_이벤트"),
    ("포인트 레이스", "포인트레이스_이벤트"),
    ("포인트레이스",  "포인트레이스_이벤트"),
    ("빙고",          "빙고_이벤트"),
    ("예측",          "예측_이벤트"),
    ("찾기",          "탐색_이벤트"),
]

REWARD_TYPE_MAP = [
    ("다이아",   ["다이아"]),
    ("골드",     ["골드"]),
    ("팩",       ["팩"]),
    ("박스",     ["박스"]),
    ("코인",     ["코인"]),
    ("쿠폰",     ["쿠폰"]),
    ("선수",     ["선수"]),
    ("뽑기",     ["뽑기"]),
    ("주머니",   ["주머니"]),
    ("상자",     ["상자"]),
    ("정수",     ["정수"]),
    ("파편",     ["파편"]),
]

SKIP_PREFIXES = ("∎", "※", "·", "①", "②", "③", "④", "◆", "▶", "■", "○", "●", "—", "-")

# ─── 헤더 기반 보상 컬럼 탐지 ─────────────────────────────────────────────
# 앞으로 [보상 아이템] / [보상 수량] 컬럼명을 표준으로 사용.
# 현재 파일(보상 아이템 / 수량)도 하위 호환 지원.
REWARD_ITEM_HEADERS: frozenset = frozenset({"[보상 아이템]", "보상 아이템"})
REWARD_QTY_HEADERS:  frozenset = frozenset({"[보상 수량]",  "수량"})

QTY_PATTERNS = [
    re.compile(r'(\d[\d,]*)\s*(개|회|장|세트|EA|번)(?!\s*[%팩])', re.UNICODE),
    re.compile(r'(다이아|골드|코인|포인트)\s*(\d[\d,]+)(?!\s*[%팩])', re.UNICODE),
    re.compile(r'^(\d[\d,]+)$'),
]

RATE_THRESHOLDS = {"required": 0.80, "recommended": 0.50, "optional": 0.30}


# ─── 유틸 함수 ───────────────────────────────────────────────────────────────

def is_date_tab(name: str) -> bool:
    return bool(DATE_TAB_PATTERN.match(str(name)))


def normalize_event_type(title: str) -> str:
    """이벤트 섹션 제목 → 표준 유형명."""
    for keyword, etype in EVENT_TYPE_KEYWORDS:
        if keyword.lower() in title.lower():
            return etype
    return "기타"


def parse_quantity(text: str) -> dict | None:
    """수량 패턴 파싱. 신뢰도(high/medium/low) 포함."""
    if not text:
        return None
    text = str(text).strip()
    m = QTY_PATTERNS[0].search(text)
    if m:
        return {"value": int(m.group(1).replace(',', '')), "unit": m.group(2), "confidence": "high"}
    m = QTY_PATTERNS[1].search(text)
    if m:
        return {"value": int(m.group(2).replace(',', '')), "unit": m.group(1), "confidence": "medium"}
    m = QTY_PATTERNS[2].match(text)
    if m:
        try:
            return {"value": int(m.group(1).replace(',', '')), "unit": "", "confidence": "low"}
        except ValueError:
            pass
    return None


def has_reward_keyword(text: str) -> bool:
    """보상 관련 키워드 포함 여부 판단."""
    if not text:
        return False
    if any(text.startswith(p) for p in SKIP_PREFIXES):
        return False
    return any(kw in text for _, keywords in REWARD_TYPE_MAP for kw in keywords)


def safe_cell_value(cell) -> str | None:
    """read_only 모드의 EmptyCell 등 방어적 값 추출."""
    val = getattr(cell, "value", None)
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


# ─── URL / ID 파싱 ───────────────────────────────────────────────────────────

def parse_spreadsheet_id(url: str) -> str:
    """Google Sheets URL에서 spreadsheet_id 추출."""
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
    if not m:
        raise ValueError(
            f"URL에서 spreadsheet_id 추출 실패: {url}\n"
            "예시: https://docs.google.com/spreadsheets/d/SPREADSHEET_ID/edit"
        )
    return m.group(1)


def extract_project_id_from_title(title: str) -> str:
    """
    스프레드시트 제목에서 project_id 추출.
    예: "[NC_KR] 라이브 이벤트 문서" → "NC_KR"
        "[FB_GL] 2026 라이브 이벤트"  → "FB_GL"
        "라이브 이벤트 문서"          → "라이브_이벤트"  (fallback)
    """
    m = re.search(r"\[([^\]]+)\]", title)
    if m:
        return re.sub(r"[^\w]", "_", m.group(1)).strip("_")
    # fallback: 제목 앞 2단어 조합
    words = re.sub(r"[^\w\s]", "", title).split()
    return "_".join(words[:2])[:24] if words else "project"


# ─── 다운로드 (공개 Export URL) ───────────────────────────────────────────────

def try_public_download(spreadsheet_id: str, dest_path: Path) -> bool:
    """
    API 키 없이 공개 Export URL로 xlsx 다운로드 시도.

    공개 시트(링크 있는 모든 사용자 → 뷰어)라면 인증 없이 다운로드 가능.
    비공개 시트는 Google 로그인 페이지로 리다이렉트되므로 False 반환.

    Returns:
        True  — 다운로드 성공, dest_path 에 xlsx 저장됨
        False — 비공개 시트이거나 오류 (requests 미설치 포함)
    """
    if not HAS_REQUESTS:
        return False

    export_url = (
        f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"
        f"/export?format=xlsx"
    )

    try:
        session = _requests.Session()
        session.headers.update({
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            )
        })

        resp = session.get(export_url, allow_redirects=True, timeout=120, stream=True)

        # 로그인 페이지로 리다이렉트 → 비공개 시트
        if "accounts.google.com" in resp.url:
            return False

        content_type = resp.headers.get("Content-Type", "")
        if resp.status_code != 200 or "spreadsheetml" not in content_type:
            return False

        dest_path.parent.mkdir(parents=True, exist_ok=True)
        total = int(resp.headers.get("Content-Length", 0))
        downloaded = 0
        with open(dest_path, "wb") as f:
            for chunk in resp.iter_content(chunk_size=65536):
                if chunk:
                    f.write(chunk)
                    downloaded += len(chunk)
                    if total:
                        pct = int(downloaded / total * 100)
                        print(f"\r  다운로드 중... {pct}%", end="", flush=True)
        print()  # 줄바꿈
        return True

    except Exception as e:
        print(f"\n  공개 다운로드 오류: {e}")
        return False


def get_title_from_html(spreadsheet_id: str) -> str:
    """
    공개 HTML 페이지 <title> 태그에서 스프레드시트 제목 추출.
    예: "<title>[NC_KR] 라이브 이벤트 문서 - Google Sheets</title>"
         → "[NC_KR] 라이브 이벤트 문서"
    실패 시 빈 문자열 반환.
    """
    if not HAS_REQUESTS:
        return ""
    try:
        url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/"
        resp = _requests.get(url, timeout=15, headers={
            "User-Agent": "Mozilla/5.0 (compatible; Googlebot/2.1)"
        })
        m = re.search(
            r"<title[^>]*>(.+?)(?:\s*[-|–]\s*Google (?:Sheets|스프레드시트))?</title>",
            resp.text,
            re.IGNORECASE,
        )
        if m:
            return m.group(1).strip()
    except Exception:
        pass
    return ""


def get_title_from_xlsx(xlsx_path: Path) -> str:
    """
    xlsx 파일 속성(core.xml)에서 제목 추출.
    Google Sheets 내보내기 파일은 대부분 비어 있으므로 보조 수단으로만 사용.
    """
    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
        title = getattr(wb.properties, "title", "") or ""
        wb.close()
        return title.strip()
    except Exception:
        return ""


# ─── 다운로드 (서비스 계정 — 비공개 시트 fallback) ───────────────────────────

def get_credentials():
    """서비스 계정 JSON에서 인증 정보 생성."""
    if not HAS_GOOGLE_API:
        raise ImportError(
            "google-api-python-client 미설치.\n"
            "비공개 시트 접근 시 필요: pip install google-api-python-client google-auth"
        )
    if not CREDENTIALS_PATH.exists():
        raise FileNotFoundError(
            f"서비스 계정 키 파일 없음: {CREDENTIALS_PATH}\n"
            "해결: credentials/service_account.json 을 추가해주세요. (docs/google_auth_setup.md 참조)\n"
            "또는 시트를 공개로 설정해주세요: 공유 → '링크가 있는 모든 사용자' → 뷰어"
        )
    return _sa.Credentials.from_service_account_file(
        str(CREDENTIALS_PATH), scopes=SCOPES
    )


def get_spreadsheet_title_api(sheets_svc, spreadsheet_id: str) -> str:
    """Sheets API로 스프레드시트 제목 조회."""
    result = sheets_svc.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        fields="properties/title"
    ).execute()
    return result["properties"]["title"]


def download_as_xlsx_api(drive_svc, spreadsheet_id: str, dest_path: Path) -> Path:
    """Drive API export로 Google Sheets → xlsx 다운로드."""
    request = drive_svc.files().export_media(
        fileId=spreadsheet_id,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    buf = io.BytesIO()
    downloader = _MIBD(buf, request)
    done = False
    while not done:
        status, done = downloader.next_chunk()
        if status:
            pct = int(status.progress() * 100)
            print(f"\r  다운로드 중... {pct}%", end="", flush=True)
    print()  # 줄바꿈
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    with open(dest_path, "wb") as f:
        f.write(buf.getvalue())
    return dest_path


# ─── 분석 함수 (순수 Python — LLM 토큰 미사용) ──────────────────────────────

def extract_event_names(wb) -> dict:
    """
    모든 날짜형 탭의 B열에서 이벤트 섹션 제목(숫자. 패턴) 추출.
    extract_event_names.py 로직 인라인 버전.
    """
    result = {}
    date_tabs = [n for n in wb.sheetnames if is_date_tab(n)]

    for tab in date_tabs:
        ws = wb[tab]
        sections = []
        for row in ws.iter_rows():
            for cell in row:
                if getattr(cell, "column", None) != 2:
                    continue
                val = safe_cell_value(cell)
                if val and SECTION_PATTERN.match(val):
                    sections.append({
                        "title": val,
                        "event_type": normalize_event_type(val),
                        "season_keywords": SEASON_RE.findall(val),
                        "cell": getattr(cell, "coordinate", "?"),
                        "row": getattr(cell, "row", None),
                    })
        if sections:
            result[tab] = sections

    return result


def _crawl_scan_by_header(rows_data: dict, start_row: int, end_row: int) -> list:
    """
    [보상 아이템] / [보상 수량] 헤더 기반 보상 수집.
    헤더가 없으면 빈 리스트 반환 → 키워드 폴백 사용.
    """
    result = []
    item_col: int | None = None
    qty_col:  int | None = None
    header_found = False

    for row_num in range(start_row, end_row + 1):
        cells = rows_data.get(row_num, {})
        if not cells:
            continue

        new_item_col: int | None = None
        new_qty_col:  int | None = None
        for col, (coord, val) in cells.items():
            v = val.strip()
            if v in REWARD_ITEM_HEADERS:
                new_item_col = col
            if v in REWARD_QTY_HEADERS:
                new_qty_col = col

        if new_item_col is not None:
            item_col     = new_item_col
            qty_col      = new_qty_col
            header_found = True
            continue

        if item_col is None:
            continue

        item_cell = cells.get(item_col)
        if not item_cell:
            continue
        item_coord, item_name = item_cell
        item_name = item_name.strip()
        if not item_name:
            continue

        nearest = None
        if qty_col is not None and qty_col in cells:
            qty_coord, qty_raw = cells[qty_col]
            # openpyxl이 숫자 셀을 '1000000.0' 형식으로 반환하는 경우 정수로 변환
            try:
                fval = float(qty_raw)
                if fval > 0 and fval == int(fval):
                    qty_raw = str(int(fval))
            except (ValueError, TypeError):
                pass
            qty_val = parse_quantity(qty_raw)
            if qty_val:
                # 헤더 기반 수량 컬럼은 명확히 지정된 값 → 항상 고신뢰도
                qty_val["confidence"] = "high"
                nearest = {
                    "cell": qty_coord, "value": qty_raw,
                    "quantity": qty_val, "dist": abs(qty_col - item_col),
                }

        result.append({
            "cell":             item_coord,
            "reward_name":      item_name,
            "quantity_in_cell": None,
            "nearest_quantity": nearest,
            "has_quantity":     nearest is not None,
            "scan_mode":        "header",
        })

    return result if header_found else []


def _crawl_scan_by_keyword(rows_data: dict, start_row: int, end_row: int) -> list:
    """키워드 기반 보상 수집 (헤더 없는 섹션의 폴백)."""
    result = []
    for rn in range(start_row, end_row + 1):
        cells = rows_data.get(rn)
        if not cells:
            continue
        for col, (coord, val) in cells.items():
            if not has_reward_keyword(val):
                continue
            nearby: list = []
            for oc, (oc_coord, ov) in cells.items():
                if oc == col:
                    continue
                qty = parse_quantity(ov)
                if qty and qty["confidence"] in ("high", "medium"):
                    nearby.append({
                        "cell": oc_coord, "value": ov,
                        "quantity": qty, "dist": abs(oc - col)
                    })
            nearby.sort(key=lambda x: x["dist"])
            self_qty = parse_quantity(val) if re.search(r'\d', val) else None
            result.append({
                "cell":             coord,
                "reward_name":      val,
                "quantity_in_cell": self_qty,
                "nearest_quantity": nearby[0] if nearby else None,
                "has_quantity":     bool(self_qty or nearby),
                "scan_mode":        "keyword",
            })
    return result


def scan_rewards(wb) -> dict:
    """
    이벤트 섹션별 보상 패턴 스캔 + 이벤트 유형별 통계.
    scan_rewards_by_event.py 로직 인라인 버전.
    """
    date_tabs = [n for n in wb.sheetnames if is_date_tab(n)]
    all_tab_sections: dict = {}

    for tab in date_tabs:
        ws = wb[tab]

        # 전체 행 데이터 수집
        rows_data: dict[int, dict] = {}
        for row in ws.iter_rows():
            row_num = None
            row_dict: dict = {}
            for cell in row:
                rn = getattr(cell, "row", None)
                if row_num is None:
                    row_num = rn
                val = safe_cell_value(cell)
                if val:
                    col = getattr(cell, "column", 0)
                    coord = getattr(cell, "coordinate", f"?{rn}")
                    row_dict[col] = (coord, val)
            if row_dict and row_num is not None:
                rows_data[row_num] = row_dict

        if not rows_data:
            all_tab_sections[tab] = []
            continue

        # 섹션 경계 탐색 (B열 = column 2)
        section_starts: list[tuple] = []
        for row_num in sorted(rows_data.keys()):
            b_cell = rows_data[row_num].get(2)
            if b_cell and SECTION_PATTERN.match(b_cell[1]):
                section_starts.append((row_num, b_cell[0], b_cell[1]))

        max_row = max(rows_data.keys())
        sections = []

        for i, (start_row, title_coord, title) in enumerate(section_starts):
            end_row = section_starts[i + 1][0] - 1 if (i + 1) < len(section_starts) else max_row

            # ── [보상 아이템]/[보상 수량] 헤더 기반 우선, 없으면 키워드 폴백 ──
            reward_rows = _crawl_scan_by_header(rows_data, start_row + 1, end_row)
            if not reward_rows:
                reward_rows = _crawl_scan_by_keyword(rows_data, start_row + 1, end_row)

            sections.append({
                "title": title,
                "event_type": normalize_event_type(title),
                "start_row": start_row,
                "end_row": end_row,
                "reward_count": len(reward_rows),
                "reward_rows": reward_rows,
            })

        all_tab_sections[tab] = sections

    # 이벤트 유형별 패턴 통계
    type_data: dict = {}
    for tab, sections in all_tab_sections.items():
        for sec in sections:
            etype = sec["event_type"]
            if etype not in type_data:
                type_data[etype] = {
                    "seen_tabs": [], "reward_name_freq": {}, "qty_samples": {},
                    "compositions": [],
                }
            type_data[etype]["seen_tabs"].append(tab)
            composition = []
            for rr in sec["reward_rows"]:
                rname = rr["reward_name"]
                type_data[etype]["reward_name_freq"][rname] = \
                    type_data[etype]["reward_name_freq"].get(rname, 0) + 1
                qty = rr.get("quantity_in_cell") or (
                    rr["nearest_quantity"]["quantity"] if rr.get("nearest_quantity") else None
                )
                rtype = next(
                    (rt for rt, kws in REWARD_TYPE_MAP if any(kw in rname for kw in kws)),
                    "기타"
                )
                if qty and qty.get("confidence") in ("high", "medium"):
                    type_data[etype]["qty_samples"].setdefault(rtype, []).append(qty["value"])
                composition.append({"reward_name": rname, "reward_type": rtype, "quantity": qty})
            if composition:
                type_data[etype]["compositions"].append({"tab": tab, "rewards": composition})

    event_type_patterns: dict = {}
    for etype, data in type_data.items():
        qty_stats = {
            rtype: {
                "min": min(s), "max": max(s),
                "avg": round(sum(s) / len(s)), "samples": len(s)
            }
            for rtype, s in data["qty_samples"].items()
        }
        top_names = sorted(data["reward_name_freq"].items(), key=lambda x: -x[1])[:5]
        event_type_patterns[etype] = {
            "seen_count":      len(set(data["seen_tabs"])),
            "seen_tabs":       sorted(set(data["seen_tabs"])),
            "top_reward_names": [n for n, _ in top_names],
            "quantity_stats":  qty_stats,
            "recent_compositions": data["compositions"][-3:],
        }

    return {
        "scanned_at":        datetime.now().isoformat(timespec="seconds"),
        "event_type_patterns": event_type_patterns,
        "per_tab_sections":  all_tab_sections,
    }


def analyze_frequency(wb) -> dict:
    """
    이벤트 유형별 등장 빈도·우선순위 분석.
    analyze_event_patterns.py 로직 인라인 버전.
    """
    date_tabs = [n for n in wb.sheetnames if is_date_tab(n)]
    total_tabs = len(date_tabs)

    type_tabs: dict   = defaultdict(list)
    type_titles: dict = defaultdict(list)
    events_per_tab: list[int] = []

    for tab in sorted(date_tabs):
        ws = wb[tab]
        sections = []
        for row in ws.iter_rows():
            for cell in row:
                if getattr(cell, "column", None) != 2:
                    continue
                val = safe_cell_value(cell)
                if val and SECTION_PATTERN.match(val):
                    sections.append({
                        "title": val,
                        "event_type": normalize_event_type(val),
                    })

        events_per_tab.append(len(sections))
        seen: set = set()
        for sec in sections:
            etype = sec["event_type"]
            if etype not in seen:
                type_tabs[etype].append(tab)
                seen.add(etype)
            if sec["title"] not in type_titles[etype] and len(type_titles[etype]) < 5:
                type_titles[etype].append(sec["title"])

    frequency_stats: dict = {}
    for etype in type_tabs:
        tabs = type_tabs[etype]
        rate = len(tabs) / total_tabs if total_tabs > 0 else 0
        priority = "rare"
        for pname in ("required", "recommended", "optional"):
            if rate >= RATE_THRESHOLDS[pname]:
                priority = pname
                break
        frequency_stats[etype] = {
            "count":         len(tabs),
            "total_tabs":    total_tabs,
            "rate":          round(rate, 3),
            "rate_pct":      f"{rate * 100:.0f}%",
            "priority":      priority,
            "seen_tabs":     sorted(tabs),
            "title_examples": type_titles.get(etype, [])[-3:],
        }

    tab_count_stats: dict = {}
    if events_per_tab:
        tab_count_stats = {
            "avg": round(sum(events_per_tab) / len(events_per_tab), 1),
            "min": min(events_per_tab),
            "max": max(events_per_tab),
            "recent": events_per_tab[-5:],
        }

    return {
        "analyzed_at":       datetime.now().isoformat(timespec="seconds"),
        "total_source_tabs": total_tabs,
        "tab_count_stats":   tab_count_stats,
        "event_type_frequency": frequency_stats,
    }


# ─── 저장 ────────────────────────────────────────────────────────────────────

def save_project_data(
    project_id: str,
    spreadsheet_title: str,
    source_url: str,
    xlsx_path: Path,
    event_names: dict,
    rewards: dict,
    frequency: dict,
) -> Path:
    """
    분석 결과를 output/json/projects/{project_id}/ 에 저장.
    기존 agent_learning.json 과 호환되는 구조로 저장.
    """
    project_dir  = PROJECTS_DIR / project_id
    learning_dir = project_dir / "learning"
    learning_dir.mkdir(parents=True, exist_ok=True)
    # work/file 디렉터리도 미리 생성
    (project_dir / "work").mkdir(parents=True, exist_ok=True)
    (project_dir / "file").mkdir(parents=True, exist_ok=True)

    # ① historical_event_names.json — extract_event_names.py 호환 형식
    tabs_list = [
        {
            "tab":    tab,
            "date":   f"20{tab[:2]}-{tab[2:4]}-{tab[4:6]}",
            "month":  f"20{tab[:2]}-{tab[2:4]}",
            "event_sections": sections,
        }
        for tab, sections in sorted(event_names.items())
    ]
    (learning_dir / "historical_event_names.json").write_text(
        json.dumps({
            "source":       str(xlsx_path),
            "extracted_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "total_tabs":   len(tabs_list),
            "tabs":         tabs_list,
            "recent_tabs":  tabs_list[-4:],
        }, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

    # ② reward_by_event.json — scan_rewards_by_event.py 호환 형식
    (learning_dir / "reward_by_event.json").write_text(
        json.dumps({**rewards, "source": str(xlsx_path)}, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

    # ③ event_pattern_analysis.json — analyze_event_patterns.py 호환 형식
    (learning_dir / "event_pattern_analysis.json").write_text(
        json.dumps({**frequency, "source": str(xlsx_path)}, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

    # ④ agent_learning.json — CLAUDE.md 세션 로드용 메인 파일
    learning_path = learning_dir / "agent_learning.json"
    existing: dict = {}
    if learning_path.exists():
        try:
            existing = json.loads(learning_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            pass

    learning = {
        "project_id":        project_id,
        "spreadsheet_title": spreadsheet_title,
        "source_url":        source_url,
        "source_xlsx":       str(xlsx_path),
        "last_crawled":      datetime.now().isoformat(timespec="seconds"),
        "version":           existing.get("version", 0) + 1,
        "accumulated_learnings": {
            # 장르 키워드 — 기존 누적 유지
            "genre_keywords": existing.get("accumulated_learnings", {}).get("genre_keywords", {}),
            # 이벤트 명칭 패턴 (탭명 → 제목 목록)
            "event_name_patterns": {
                tab: [s["title"] for s in secs]
                for tab, secs in event_names.items()
            },
            # 보상 패턴 — 기존 누적 유지
            "reward_patterns": existing.get("accumulated_learnings", {}).get("reward_patterns", {}),
            # 보상 치환 이력 — 기존 누적 유지
            "reward_replacement_patterns": existing.get("accumulated_learnings", {}).get("reward_replacement_patterns", {}),
            # 이벤트 유형별 보상 구성 (새로 덮어쓰기)
            "event_reward_patterns": rewards.get("event_type_patterns", {}),
            # 이벤트 유형 빈도 (새로 덮어쓰기)
            "event_frequency_patterns": frequency.get("event_type_frequency", {}),
        },
        "tab_count_stats": frequency.get("tab_count_stats", {}),
    }

    learning_path.write_text(
        json.dumps(learning, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    return learning_path


# ─── 프로젝트 목록 ────────────────────────────────────────────────────────────

def list_projects() -> list[dict]:
    """등록된 프로젝트 목록 반환. 새 구조(learning/) + 레거시(직접) 모두 검색."""
    results: dict = {}

    def _scan(projects_dir: Path, learning_subdir: bool) -> None:
        if not projects_dir.exists():
            return
        for d in sorted(projects_dir.iterdir()):
            if not d.is_dir():
                continue
            lf = (d / "learning" / "agent_learning.json") if learning_subdir else (d / "agent_learning.json")
            if not lf.exists():
                continue
            pid = d.name
            if pid in results:
                continue  # 새 구조 우선
            try:
                data = json.loads(lf.read_text(encoding="utf-8"))
                acc = data.get("accumulated_learnings", {})
                results[pid] = {
                    "project_id":    pid,
                    "title":         data.get("spreadsheet_title", "?"),
                    "last_crawled":  data.get("last_crawled", "?"),
                    "tab_count":     len(acc.get("event_name_patterns", {})),
                    "event_types":   len(acc.get("event_frequency_patterns", {})),
                    "reward_types":  len(acc.get("event_reward_patterns", {})),
                }
            except (json.JSONDecodeError, KeyError):
                pass

    # 새 구조 우선 스캔
    _scan(PROJECTS_DIR, learning_subdir=True)
    # 레거시 폴백 (output/json/projects/)
    _legacy = BASE_DIR / "output" / "json" / "projects"
    _scan(_legacy, learning_subdir=False)

    return list(results.values())


def print_project_list(projects: list[dict]) -> None:
    if not projects:
        print("등록된 프로젝트 없음.")
        print("실행: python scripts/crawl_gdrive_project.py <Google Sheets URL>")
        return
    print(f"등록된 프로젝트 ({len(projects)}개):")
    print(f"  {'ID':<16} {'제목':<32} {'탭':>4} {'유형':>5} {'마지막 크롤링'}")
    print("  " + "─" * 72)
    for p in projects:
        print(
            f"  {p['project_id']:<16} {p['title']:<32} "
            f"{p['tab_count']:>4}개 {p['event_types']:>4}종  {p['last_crawled'][:10]}"
        )


# ─── 메인 ────────────────────────────────────────────────────────────────────

def main():
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

    parser = argparse.ArgumentParser(
        description="Google Sheets URL → 프로젝트별 학습 데이터 크롤러"
    )
    parser.add_argument("url", nargs="?", help="Google Sheets URL")
    parser.add_argument("--project-id", help="project_id 직접 지정 (미지정 시 제목에서 자동 추출)")
    parser.add_argument("--list", action="store_true", help="등록된 프로젝트 목록 출력")
    parser.add_argument(
        "--skip-download",
        metavar="XLSX_PATH",
        help="이미 다운로드된 xlsx 경로 지정 (다운로드 건너뜀)"
    )
    args = parser.parse_args()

    # ── 목록 출력 모드 ──────────────────────────────────────────────────────
    if args.list:
        print_project_list(list_projects())
        return

    if not args.url:
        parser.print_help()
        sys.exit(1)

    try:
        # ── Step 1: URL 파싱 ─────────────────────────────────────────────
        print("[1/4] URL 파싱...")
        spreadsheet_id = parse_spreadsheet_id(args.url)
        print(f"  → spreadsheet_id: {spreadsheet_id}")

        # ── Step 2: xlsx 취득 ────────────────────────────────────────────
        title = ""
        temp_path = DOWNLOAD_DIR / f"temp_{spreadsheet_id[:8]}.xlsx"

        if args.skip_download:
            # 수동 지정 파일 사용
            xlsx_path = Path(args.skip_download).resolve()
            if not xlsx_path.exists():
                raise FileNotFoundError(f"지정된 xlsx 파일 없음: {xlsx_path}")
            print(f"[2/4] 다운로드 건너뜀 → {xlsx_path}")
            title = get_title_from_html(spreadsheet_id) or get_title_from_xlsx(xlsx_path)

        else:
            # ① 공개 Export URL 시도 (API 키 불필요)
            if HAS_REQUESTS:
                print("[2/4] xlsx 다운로드 시도 (공개 Export URL — API 키 불필요)...")
                ok = try_public_download(spreadsheet_id, temp_path)
            else:
                print("[2/4] requests 미설치 — 서비스 계정으로 바로 시도합니다.")
                print("       (공개 시트 지원 설치: pip install requests)")
                ok = False

            if ok:
                # 공개 다운로드 성공
                print("  → 공개 다운로드 성공")
                title = get_title_from_html(spreadsheet_id)
                if title:
                    print(f"  → 제목 (HTML): {title}")
                else:
                    title = get_title_from_xlsx(temp_path)
                    if title:
                        print(f"  → 제목 (xlsx 속성): {title}")

            else:
                # ② 서비스 계정 fallback
                if ok is False and HAS_REQUESTS:
                    print("  → 비공개 시트 또는 다운로드 실패.")
                print("  서비스 계정으로 재시도...")

                if not HAS_GOOGLE_API:
                    raise ImportError(
                        "google-api-python-client 미설치.\n"
                        "  pip install google-api-python-client google-auth\n"
                        "또는 시트를 공개로 설정하면 API 키 없이 실행 가능:\n"
                        "  공유 → '링크가 있는 모든 사용자' → 뷰어"
                    )

                creds      = get_credentials()
                sheets_svc = _gapi_build("sheets", "v4", credentials=creds)
                drive_svc  = _gapi_build("drive",  "v3", credentials=creds)

                title = get_spreadsheet_title_api(sheets_svc, spreadsheet_id)
                print(f"  → 제목 (API): {title}")

                download_as_xlsx_api(drive_svc, spreadsheet_id, temp_path)

            # 파일명 확정: temp → {project_id}_{id[:8]}.xlsx
            project_id_temp = args.project_id or extract_project_id_from_title(title or spreadsheet_id[:8])
            xlsx_filename   = f"{project_id_temp}_{spreadsheet_id[:8]}.xlsx"
            xlsx_path       = DOWNLOAD_DIR / xlsx_filename

            if temp_path.resolve() != xlsx_path.resolve():
                xlsx_path.parent.mkdir(parents=True, exist_ok=True)
                shutil.move(str(temp_path), str(xlsx_path))

            size_kb = xlsx_path.stat().st_size // 1024
            print(f"  → 저장: {xlsx_path}  ({size_kb} KB)")

        # ── project_id / title 확정 ───────────────────────────────────────
        if not title:
            title = get_title_from_xlsx(xlsx_path) or spreadsheet_id[:16]
        project_id = args.project_id or extract_project_id_from_title(title)
        print(f"  → project_id: {project_id}")
        print(f"  → 제목: {title}")

        # ── Step 3: 데이터 분석 ──────────────────────────────────────────
        print("[3/4] 데이터 분석 중...")
        wb         = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
        date_tabs  = [n for n in wb.sheetnames if is_date_tab(n)]
        total_tabs = len(date_tabs)
        preview    = ", ".join(date_tabs[:5]) + ("..." if total_tabs > 5 else "")
        print(f"  → 날짜형 탭: {total_tabs}개  ({preview})")

        print("  [3a] 이벤트 명칭 추출...")
        event_names = extract_event_names(wb)
        print(f"       {len(event_names)}개 탭에서 이벤트 섹션 추출 완료")

        print("  [3b] 보상 패턴 스캔...")
        rewards = scan_rewards(wb)
        types_found = len(rewards.get("event_type_patterns", {}))
        print(f"       {types_found}종 이벤트 유형 패턴 수집 완료")

        print("  [3c] 이벤트 빈도 분석...")
        frequency = analyze_frequency(wb)
        freq_types = len(frequency.get("event_type_frequency", {}))
        print(f"       {freq_types}종 이벤트 유형 빈도 분석 완료")

        wb.close()

        # ── Step 4: 저장 ─────────────────────────────────────────────────
        print("[4/4] 학습 데이터 저장...")
        learning_path = save_project_data(
            project_id, title, args.url, xlsx_path,
            event_names, rewards, frequency
        )

        # ── 현재 프로젝트 설정 저장 (다른 스크립트가 참조) ──────────────────
        current_proj_path = BASE_DIR / "output" / "json" / "current_project.json"
        current_proj_path.parent.mkdir(parents=True, exist_ok=True)
        current_proj_path.write_text(
            json.dumps({
                "project_id":  project_id,
                "source_xlsx": str(xlsx_path),
                "title":       title,
                "updated_at":  datetime.now().isoformat(timespec="seconds"),
            }, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(f"  → current_project.json 갱신: {project_id}")

        # ── 결과 출력 ─────────────────────────────────────────────────────
        print(f"\n✅ 크롤링 완료!")
        print(f"  project_id    : {project_id}")
        print(f"  스프레드시트  : {title}")
        print(f"  분석 탭 수    : {len(event_names)}개")
        print(f"  이벤트 유형   : {freq_types}종")
        print(f"  보상 유형     : {types_found}종")
        print(f"  저장 경로     : {learning_path.parent}/  (learning/)")
        print(f"    ├ agent_learning.json")
        print(f"    ├ historical_event_names.json")
        print(f"    ├ reward_by_event.json")
        print(f"    └ event_pattern_analysis.json")
        print(f"\n다음 세션에서 이 프로젝트를 사용하려면:")
        print(f"  python scripts/load_project_learning.py --project-id {project_id} --summary")

    except FileNotFoundError as e:
        print(f"\n[에스컬레이션] {e}")
        sys.exit(1)
    except ImportError as e:
        print(f"\n[설치 필요] {e}")
        sys.exit(1)
    except Exception as e:
        import traceback
        print(f"\n[오류] {type(e).__name__}: {e}")
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
