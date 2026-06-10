#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
멀티 에이전트 PM — Flask 백엔드 서버

사용자의 자연어 요청 → PM(Claude) 분석 → 에이전트 자동 위임 → 파이프라인 실행
"""
import json
import subprocess
import sys
import webbrowser
import threading
import time
import uuid
import queue as _queue_module
from pathlib import Path
from datetime import datetime

from flask import (Flask, jsonify, request, send_file, render_template,
                   send_from_directory, Response, stream_with_context)

# ── AI 클라이언트 설정 ───────────────────────────────────────────────────────
# 우선순위: 1) Claude CLI (subprocess)  2) Claude API  3) Ollama  4) none
import os
import shutil
import subprocess
import urllib.request

HAS_AI     = False
_AI        = None
_ANTHROPIC = None
AI_MODE    = "none"   # "claude_cli" | "anthropic" | "ollama" | "none"

OLLAMA_HOST  = os.environ.get("OLLAMA_HOST", "http://localhost:11434")
OLLAMA_MODEL = os.environ.get("OLLAMA_MODEL", "")

def _detect_ollama_model(host: str) -> str:
    """실행 중인 Ollama에서 첫 번째 모델명 반환. 없으면 빈 문자열."""
    try:
        with urllib.request.urlopen(f"{host}/api/tags", timeout=2) as r:
            data = json.loads(r.read())
            models = data.get("models", [])
            return models[0]["name"] if models else ""
    except Exception:
        return ""

# Windows에서 npm global bin의 claude.cmd 경로 자동 탐색
def _find_claude_cmd() -> list | None:
    """
    claude 실행 커맨드 목록 반환.
    Windows: npm global bin의 claude.cmd 우선, 없으면 PATH에서 탐색.
    """
    import platform
    candidates = []
    if platform.system() == "Windows":
        # npm global prefix 경로 탐색
        npm_global = os.path.expandvars(r"%APPDATA%\npm")
        cmd_path = os.path.join(npm_global, "claude.cmd")
        if os.path.exists(cmd_path):
            candidates.append(["cmd", "/c", cmd_path])
        # node_modules 직접 경로
        exe_path = os.path.join(npm_global, "node_modules",
                                "@anthropic-ai", "claude-code", "bin", "claude.exe")
        if os.path.exists(exe_path):
            candidates.append([exe_path])
    # 공통: PATH에서 탐색
    found = shutil.which("claude") or shutil.which("claude.cmd")
    if found:
        candidates.append(["cmd", "/c", found] if found.endswith(".cmd") else [found])
    return candidates[0] if candidates else None

_CLAUDE_CMD = _find_claude_cmd()

def _call_claude_cli(prompt: str, timeout: int = 60) -> str:
    """
    claude CLI를 subprocess로 호출하여 응답 텍스트 반환.
    Windows의 .cmd 래퍼도 자동으로 처리.
    """
    if not _CLAUDE_CMD:
        raise RuntimeError("claude CLI를 찾을 수 없습니다.")
    result = subprocess.run(
        _CLAUDE_CMD + ["-p", prompt],
        capture_output=True, text=True, encoding="utf-8",
        timeout=timeout,
    )
    # stderr에 warning만 있어도 returncode=0이면 성공으로 처리
    if result.returncode != 0 and not result.stdout.strip():
        raise RuntimeError(f"claude CLI 오류: {result.stderr.strip()[:200]}")
    return result.stdout.strip()

# 1) Claude CLI 자동 감지 (최우선)
if _CLAUDE_CMD:
    try:
        _test = subprocess.run(
            _CLAUDE_CMD + ["-p", "숫자 1만 출력"],
            capture_output=True, text=True, encoding="utf-8", timeout=20,
        )
        if _test.returncode == 0 and _test.stdout.strip():
            HAS_AI  = True
            AI_MODE = "claude_cli"
            print(f"[PM] Claude CLI 연결 완료  ({_CLAUDE_CMD[0]})")
        else:
            print(f"[PM] Claude CLI 응답 없음: {_test.stderr.strip()[:100]}")
    except Exception as _ce:
        print(f"[PM] Claude CLI 테스트 실패: {_ce}")

# 2) Claude API fallback
if not HAS_AI:
    try:
        import anthropic
        _ANTHROPIC = anthropic.Anthropic()
        HAS_AI  = True
        AI_MODE = "anthropic"
        print("[PM] Claude API 연결")
    except Exception:
        pass

# 3) Ollama fallback
if not HAS_AI:
    try:
        from openai import OpenAI as _OpenAI
        detected = OLLAMA_MODEL or _detect_ollama_model(OLLAMA_HOST)
        if detected:
            OLLAMA_MODEL = detected
            _AI = _OpenAI(base_url=f"{OLLAMA_HOST.rstrip('/')}/v1", api_key="ollama")
            HAS_AI  = True
            AI_MODE = "ollama"
            print(f"[PM] Ollama 연결: {OLLAMA_HOST}  모델: {OLLAMA_MODEL}")
    except Exception as e:
        print(f"[PM] Ollama 감지 실패: {e}")

if not HAS_AI:
    print("[PM] AI 미설정 — 수동 제어 모드")

# ── 경로 설정 ─────────────────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).resolve().parent
SCRIPTS_DIR = BASE_DIR / "scripts"
OUTPUT_DIR  = BASE_DIR / "output"

# 스크립트들이 기대하는 ProjectPaths 구조와 동일하게 맞춤
# _project_config.ProjectPaths("event-planner") → output/projects/event-planner/
EP_PROJECT_ID = "event-planner"
EP_WORK       = OUTPUT_DIR / "projects" / EP_PROJECT_ID / "work"
EP_FILE       = OUTPUT_DIR / "projects" / EP_PROJECT_ID / "file"

GL_DIR      = OUTPUT_DIR / "game-localizer"
HANDOFF_DIR  = OUTPUT_DIR / "handoff"
JSON_DIR     = OUTPUT_DIR / "json"
EP_CONFIG    = OUTPUT_DIR / "config" / "agent_config.json"
PROJECT_CFG  = OUTPUT_DIR / "config" / "projects.json"

# ── 프로젝트 기본 설정 ────────────────────────────────────────────────────────
_DEFAULT_PROJECTS = {
    "FB/GL": {
        "display_name": "FB Global",
        "genre":  "야구",
        "market": "글로벌",
        "configured": True,
        "description": "야구 게임 (글로벌 마켓)"
    },
    "FB/JP": {
        "display_name": "FB Japan",
        "genre":  "",
        "market": "일본",
        "configured": False,
        "description": "FB 일본 마켓"
    },
    "NC/KR": {
        "display_name": "NC Korea",
        "genre":  "",
        "market": "한국",
        "configured": False,
        "description": "NC 한국 마켓"
    },
    "NC/GL": {
        "display_name": "NC Global",
        "genre":  "",
        "market": "글로벌",
        "configured": False,
        "description": "NC 글로벌 마켓"
    },
}

def _load_projects() -> dict:
    """프로젝트 설정 로드 (없으면 기본값 반환)."""
    try:
        if PROJECT_CFG.exists():
            saved = json.loads(PROJECT_CFG.read_text(encoding="utf-8"))
            # 기본값에 저장된 값 병합 (새 프로젝트 추가 시 누락 방지)
            merged = dict(_DEFAULT_PROJECTS)
            for k, v in saved.items():
                merged[k] = {**merged.get(k, {}), **v}
            return merged
    except Exception:
        pass
    return dict(_DEFAULT_PROJECTS)

def _save_projects(projects: dict):
    PROJECT_CFG.parent.mkdir(parents=True, exist_ok=True)
    PROJECT_CFG.write_text(json.dumps(projects, ensure_ascii=False, indent=2), encoding="utf-8")

sys.path.insert(0, str(SCRIPTS_DIR))

# ── 자기학습 매니저 ──────────────────────────────────────────────────────────
try:
    from learning_manager import LearningManager
    _LEARNING = LearningManager()
    print("[PM] 자기학습 매니저 초기화 완료 (세션 수:", _LEARNING.data["session_count"], ")")
except Exception as _le:
    print(f"[PM] 학습 매니저 초기화 실패: {_le}")
    _LEARNING = None

app = Flask(__name__, template_folder="templates", static_folder="static")

# Google Drive 다운로드 캐시 디렉토리
GDRIVE_CACHE = OUTPUT_DIR / "gdrive_cache"
GDRIVE_CACHE.mkdir(parents=True, exist_ok=True)

def resolve_source(url_or_path: str) -> tuple[str, bool]:
    """
    로컬 경로 / Google Sheets URL / Google Drive URL → 로컬 파일 경로 반환.
    Returns: (local_path, was_downloaded)
    """
    from gdrive_utils import is_google_url, resolve_to_local_file
    if is_google_url(url_or_path):
        local = resolve_to_local_file(url_or_path, dest_dir=GDRIVE_CACHE)
        return str(local), True
    return url_or_path, False


# ── 유틸 ─────────────────────────────────────────────────────────────────────
def run_script(cmd: list, cwd=None) -> tuple[bool, str]:
    result = subprocess.run(
        cmd, capture_output=True, text=True,
        encoding="utf-8", errors="replace",
        cwd=str(cwd or BASE_DIR)
    )
    output = result.stdout or result.stderr or ""
    return result.returncode == 0, output


def ensure_dirs():
    for d in [EP_WORK, EP_FILE, GL_DIR, HANDOFF_DIR, JSON_DIR]:
        d.mkdir(parents=True, exist_ok=True)

def write_current_project(source_xlsx: str = ""):
    """current_project.json 갱신 — 스크립트들이 올바른 경로를 사용하도록."""
    JSON_DIR.mkdir(parents=True, exist_ok=True)
    data = {
        "project_id": EP_PROJECT_ID,
        "source_xlsx": source_xlsx,
        "title": "event-planner",
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }
    (JSON_DIR / "current_project.json").write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )


ensure_dirs()


# ═══════════════════════════════════════════════════════════════════════════════
# PM 오케스트레이터 — 자연어 → 에이전트 위임
# ═══════════════════════════════════════════════════════════════════════════════

PM_SYSTEM = """당신은 멀티 에이전트 PM(프로젝트 매니저)입니다.
사용자의 의도를 파악해 에이전트에게 위임합니다. 항상 순수 JSON으로만 응답하세요.

## ★★ 언어 규칙 (최우선 원칙) ★★
- response 필드는 반드시 한국어(Korean)로만 작성한다.
- 중국어(Chinese)·일본어(Japanese)·영어(English) 절대 금지.
- 타겟 마켓이 일본·글로벌·중국이더라도 응답은 항상 한국어다.

## PM의 역할
PM은 의도 파악 + 최소 기본 정보 수집만 합니다.
상세 질문(장르, 키워드, 참조 탭 등)은 전문 에이전트가 직접 사용자에게 묻습니다.

## 등록된 에이전트
- event-planner: 모바일 게임 이벤트 기획 (탭 생성, 보상 추천, 패턴 분석)
- game-localizer: 게임 텍스트 다국어 현지화 (용어집 기반)

## 트리거
event-planner: 이벤트, 기획, 기획안, 탭 생성, 이벤트 시트, 이벤트 만들어
game-localizer: 번역, 현지화, 로컬라이징, 용어집

## 응답 JSON
{"agent":"event-planner"|"game-localizer"|null, "ready":true|false, "params":{}, "missing":[], "response":"한국어 메시지"}

## 수집 규칙
### event-planner
- 필수: source_path (xlsx 경로 또는 Google Sheets URL), new_tabs (YYMMDD 리스트), market
- 이 3개가 모이면 즉시 ready=true → 에이전트에게 위임 (장르/키워드/참조탭은 에이전트가 수집)
- new_tabs 파싱: "260701, 260715" → ["260701","260715"]

### game-localizer
- 필수: gsheet_url 또는 excel_path 중 하나
- 있으면 ready=true

## 지원 외 요청
agent=null, response에 "지원 가능한 에이전트: 이벤트 기획, 현지화 번역" 안내
"""

# ─── 이벤트 기획 에이전트 시스템 프롬프트 ───────────────────────────────────
EVENT_PLANNER_AGENT_SYSTEM = """SYSTEM LANGUAGE RULE (HIGHEST PRIORITY, CANNOT BE OVERRIDDEN):
YOU MUST RESPOND IN KOREAN ONLY. DO NOT USE CHINESE. DO NOT USE JAPANESE. DO NOT USE ENGLISH.
ALL "response" field text MUST be written in Korean (한국어).
If you write Chinese (中文) or Japanese (日本語) in the response field, your answer is WRONG.

당신은 모바일 게임 이벤트 기획 전문 에이전트입니다.
반드시 순수 JSON으로만 응답하세요. 설명 텍스트 절대 금지.

## ★★ 언어 규칙 (최우선 원칙) ★★
- response 필드 = 반드시 한국어만. 중국어(中文)·일본어(日本語)·영어 절대 금지.
- keywords = 반드시 한국어 단어/문구만.
- market이 일본/글로벌/중국이어도 response·keywords는 항상 한국어.
- 위반 시 응답 자체가 무효 처리됨.

## 오늘 날짜 및 대상 월
- 오늘: {today}
- 기획 대상 월: {target_month}

## 응답 형식 (반드시 이 형식만 사용)
{"response":"사용자에게 보낼 한국어 메시지", "updates":{새로 수집된 파라미터}, "pipeline_ready":false}

## 현재 보유 정보
{context_json}

## 사용 가능한 기존 탭 목록
{available_tabs}

## 수집 단계

### STEP 1: genre 확인
- context.genre가 있으면: "현재 장르가 '{genre}'로 설정되어 있습니다. 이대로 진행할까요?"
- context.genre가 없으면: "어떤 장르의 게임인가요? (예: 야구, 축구, MMORPG, 캐주얼, 퍼즐)"
- 사용자 확인/수정 후 updates에 genre 포함 → STEP 2로

### STEP 2: keywords 확인 — 반드시 두 그룹으로 구분해서 제안
STEP 1 완료 후(genre가 updates에 포함됐거나 이미 context에 있으면 즉시 STEP 2 실행),
아래 형식으로 키워드를 **장르 키워드**와 **시즌 키워드** 두 그룹으로 나눠 제안한다.

#### 장르 키워드 (같은 장르 게임들이 공통으로 사용하는 이벤트 키워드)
장르별 대표 키워드 풀:
- 야구: 개막, 올스타, 포스트시즌, 드래프트, 한국시리즈, 역대급 선수, 레전드, 명예의 전당, 순위 경쟁, 우승 도전, 클러치, 끝내기, 홈런, 만루, 완봉, 퍼펙트게임, 세이브, 삼진, 도루, 번트
- 축구: 이적 시장, 챔피언스리그, 리그 개막, 월드컵, 국가대표, 전술, 유망주, 이적료, FA, 감독, 주전 경쟁, 스쿼드, 포메이션, 골든부트, 베스트 11
- MMORPG: 신규 클래스, 업데이트, 레이드, 보스, 던전, 장비 강화, 길드전, 서버 이전, 시즌 패스, 영웅 성장, 전설 아이템
- 캐주얼/퍼즐: 신규 스테이지, 업데이트, 협동 이벤트, 친구 초대, 랭킹전, 한정 스킨, 콜라보
- 스포츠 공통: 시즌 개막, 결승전, 우승, 베스트 선수, 특별 스쿼드, 한정 강화

#### 시즌 키워드 (오늘 날짜 {today} / 기획 대상 월 {target_month} 기준)
월별 시즌 가이드:
- 1월: 신년, 새해 맞이, 새 출발, 겨울 대전, 설날 준비, 동계 이벤트
- 2월: 발렌타인, 겨울 마무리, 봄 예고, 설날(음력 설 해당 시), 화이트데이 예고
- 3월: 봄 개막, 개막전, 새 시즌 시작, 벚꽃, 봄의 기운, 새봄 챌린지
- 4월: 봄 절정, 황금연휴 예고, 벚꽃 만개, 4월의 기적, 봄 대전
- 5월: 황금연휴(골든위크), 어린이날, 가정의 달, 봄 마무리, 초여름 예고, 전반기 중반
- 6월: 초여름, 전반기 결산, 여름 예고, 얼리썸머, 장마 전, 6월의 열기
- 7월: 여름 성수기, 여름 대축제, 한여름의 열기, 전반기 마무리, 올스타 시즌, 장마, 피서
- 8월: 한여름 절정, 여름 마무리, 휴가 시즌, 후반기 개막, 8월의 열정
- 9월: 가을 개막, 추석, 한가위, 가을 대전, 포스트시즌 진입, 시즌 막바지
- 10월: 포스트시즌, 핼러윈, 가을 절정, 시즌 결산, 10월의 드라마
- 11월: 시즌 종료, 겨울 예고, 연말 준비, 드래프트 시즌, 이적 시장 개막, FA 시장
- 12월: 크리스마스, 연말 결산, 올해의 선수, 새해 예고, 겨울 대축제, 연말 이벤트

대상 월 {target_month}에 해당하는 시즌 키워드를 우선 도출할 것.

#### STEP 2 response 출력 형식 (반드시 이 형식으로)
response 내용을 아래처럼 두 그룹으로 명확히 구분해서 작성한다:

"📌 [{genre}] 장르 키워드 (같은 장르 게임 공통)\n
1. 키워드A\n2. 키워드B\n3. 키워드C\n...\n\n
🗓️ 시즌 키워드 ({target_month} 기준)\n
1. 키워드X\n2. 키워드Y\n3. 키워드Z\n...\n\n
사용하실 키워드를 선택하거나, 추가·수정해 주세요. 또는 두 그룹 모두 사용하려면 '모두 사용'이라고 해주세요."

- 장르 키워드: 최소 7개, 시즌 키워드: 최소 5개 제안
- context.keywords가 이미 있으면: "현재 키워드: {keywords}\n위와 같이 저장되어 있습니다. 그대로 사용할까요? 또는 아래 추천 키워드에서 추가·교체할 수 있습니다." 후 두 그룹 추천 제시
- 사용자 확인/수정 후 updates에 keywords(선택된 전체 목록) 포함 → **반드시 같은 응답에서 STEP 3 질문을 함께 작성할 것**
- ★ keywords를 updates에 넣는 즉시, response에 STEP 3 질문(참조 탭)을 포함해야 한다. 별도 턴 없이 한 번에.

### STEP 3: ref_tabs 확인
- STEP 2 완료 후에만 진행
- context.ref_tabs가 있어도 반드시 사용자에게 확인
- 사용 가능 탭 목록: {available_tabs}
- response: "각 탭의 참조 탭을 알려주세요.\n생성할 탭: {new_tabs}\n사용 가능: {available_tabs}\n예: 260709→260625, 260716→260702"
  (context.ref_tabs가 있으면: "현재 참조 탭: {ref_tabs}. 변경하려면 새로 입력하고, 그대로면 '확인'이라고 해주세요.")
- 사용자가 확인하거나 새로 입력하면 updates에 ref_tabs 포함

## ★ 완료 조건
genre(비어있지 않음) AND keywords(길이>0) AND ref_tabs(길이>0) 이면:
{"response":"모든 정보가 준비됐습니다! 이벤트 기획 파이프라인을 시작합니다.", "updates":{수집된 파라미터 모두 포함}, "pipeline_ready":true}

## 절대 금지
- "수집되었습니다", "알겠습니다" 등 확인만 하고 멈추는 것
- pipeline_ready:false 로 응답한 후 아무 것도 안 하는 것
- context에 값이 이미 있다고 해서 질문을 건너뛰는 것
- ref_tabs를 받은 직후 pipeline_ready:false 로 응답하는 것 (반드시 true)
- response나 keywords를 한국어 이외의 언어로 작성하는 것 (절대 금지)
- STEP 2에서 장르 키워드와 시즌 키워드를 구분하지 않고 섞어서 제안하는 것
"""

_jobs: dict[str, _queue_module.Queue] = {}


@app.route("/api/pm/chat", methods=["POST"])
def pm_chat():
    data     = request.json
    user_msg = data.get("message", "").strip()
    history  = data.get("history", [])

    if not user_msg:
        return jsonify(ok=False, message="메시지가 비어 있습니다."), 400

    job_id = str(uuid.uuid4())[:8]
    q = _queue_module.Queue()
    _jobs[job_id] = q

    def run():
        try:
            _pm_process(user_msg, history, q)
        except Exception as e:
            q.put({"type": "error", "message": f"PM 처리 오류: {e}"})
            q.put({"type": "done", "status": "error"})

    threading.Thread(target=run, daemon=True).start()
    return jsonify(job_id=job_id)


@app.route("/api/pm/stream/<job_id>")
def pm_stream(job_id):
    q = _jobs.get(job_id)
    if not q:
        return jsonify(error="job not found"), 404

    def generate():
        deadline = time.time() + 600
        while time.time() < deadline:
            try:
                msg = q.get(timeout=30)
                yield f"data: {json.dumps(msg, ensure_ascii=False)}\n\n"
                if msg.get("type") == "done":
                    _jobs.pop(job_id, None)
                    break
            except _queue_module.Empty:
                yield 'data: {"type":"heartbeat"}\n\n'

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


def _pm_rule_route(user_msg: str, history: list) -> dict | None:
    """
    claude_cli 모드용 규칙 기반 PM 라우팅.
    LLM 없이 키워드/패턴으로 의도를 파악하고 에이전트를 결정한다.
    라우팅 불가 시 None 반환 → 기본 안내 메시지 출력.
    """
    import re as _re

    # 전체 대화 이력 + 현재 메시지에서 파라미터 누적 수집
    all_text = " ".join(
        m["content"] for m in history if m["role"] == "user"
    ) + " " + user_msg

    # ── 에이전트 분류 ──────────────────────────────────────────────────────────
    EVENT_KW  = ["이벤트 기획", "기획안", "이벤트 생성", "이벤트 시트",
                 "탭 생성", "이벤트 만들", "이벤트 문서", "xlsx 만들", "이벤트 작성"]
    LOCAL_KW  = ["번역", "현지화", "로컬라이징", "localization", "용어집"]

    is_event = any(kw in all_text for kw in EVENT_KW)
    is_local = any(kw in all_text for kw in LOCAL_KW)

    # Google Sheets URL
    url_m = _re.search(r'https://docs\.google\.com/spreadsheets/[^\s\]>]+', all_text)
    sheets_url = url_m.group(0).rstrip(".,)") if url_m else None

    # xlsx 파일 경로
    xlsx_m = _re.search(r'[\w/\\: ]+\.xlsx', all_text)
    xlsx_path = xlsx_m.group(0).strip() if xlsx_m else None

    source = sheets_url or xlsx_path

    # YYMMDD 날짜 탭 목록 (6자리 YYMMDD 또는 4자리 MMDD 모두 인식)
    _year_prefix = str(__import__('datetime').datetime.now().year)[2:]
    _tabs_raw = list(dict.fromkeys(
        _re.findall(r'\b(2[0-9]{5})\b', all_text) +           # 6자리: 260625
        [_year_prefix + t for t in _re.findall(r'\b((?:0[1-9]|1[0-2])(?:0[1-9]|[12]\d|3[01]))\b', all_text)]  # 4자리: 0625 → 260625
    ))
    tabs = list(dict.fromkeys(_tabs_raw))  # 중복 제거

    # 마켓
    market = None
    for m_kw in ["한국", "일본", "글로벌", "중국"]:
        if m_kw in all_text:
            market = m_kw
            break

    # ── 이벤트 기획 분기 ───────────────────────────────────────────────────────
    if is_event or source or (tabs and not is_local):
        params: dict = {}
        missing: list = []

        if source:
            params["source_path"] = source
        else:
            # 소스 경로 누락 → 에이전트 핸드오프 전에 PM이 직접 질문
            return {
                "agent": "event-planner",
                "ready": False,
                "block": True,          # block=True → _pm_process가 핸드오프하지 않고 메시지만 출력
                "params": {},
                "missing": ["source_path"],
                "response": (
                    "[PM] 이벤트 기획 요청을 확인했습니다.\n\n"
                    "📂 **참조할 이벤트 문서 경로가 필요합니다.**\n\n"
                    "아래 중 하나를 알려주세요:\n"
                    "- Google Sheets URL (예: `https://docs.google.com/spreadsheets/d/...`)\n"
                    "- 로컬 xlsx 파일 경로 (예: `C:/Users/.../이벤트.xlsx`)\n\n"
                    f"{'📅 탭: ' + ', '.join(tabs) if tabs else '생성할 탭 날짜도 함께 알려주세요. (예: 260625, 260702)'}"
                ),
            }

        if tabs:
            params["new_tabs"] = tabs
        else:
            missing.append("new_tabs (생성할 탭 날짜, 예: 260625)")

        if market:
            params["market"] = market
        else:
            missing.append("market (한국/일본/글로벌)")

        ready = len(missing) == 0

        if ready:
            resp = f"이벤트 기획을 시작합니다.\n- 소스: {source}\n- 탭: {', '.join(tabs)}\n- 마켓: {market}"
        else:
            resp = "이벤트 기획 에이전트로 연결합니다. 다음 정보를 에이전트가 추가로 확인합니다:\n" + \
                   "\n".join(f"- {m}" for m in missing)

        return {"agent": "event-planner", "ready": ready, "params": params,
                "missing": missing, "response": resp}

    # ── 현지화 번역 분기 ───────────────────────────────────────────────────────
    if is_local:
        params = {}
        if sheets_url:
            params["gsheet_url"] = sheets_url
        elif xlsx_path:
            params["excel_path"] = xlsx_path
        ready = bool(params)
        resp = "현지화 번역 에이전트로 연결합니다." if ready else \
               "번역할 Google Sheets URL 또는 Excel 파일 경로를 알려주세요."
        return {"agent": "game-localizer", "ready": ready, "params": params, "response": resp}

    return None   # 분류 불가


def _pm_process(user_msg: str, history: list, q: _queue_module.Queue):
    """PM 분석 → 에이전트 위임"""

    if not HAS_AI:
        # claude_cli 모드는 HAS_AI=True이므로 여기는 none 모드만 도달
        q.put({"type": "message", "content":
               "AI가 설정되지 않았습니다. 서버를 재시작하거나 claude CLI를 설치해주세요."})
        q.put({"type": "done", "status": "no_ai"})
        return

    # ── claude_cli 모드: 규칙 기반 라우팅 (LLM은 키워드 생성에만 사용) ────────
    if AI_MODE == "claude_cli":
        result = _pm_rule_route(user_msg, history)
        if result is None:
            # 라우팅 불가 → 지원 에이전트 안내
            q.put({"type": "message", "content":
                   "[PM] 안녕하세요. 다음 작업을 도와드릴 수 있습니다.\n\n"
                   "**이벤트 기획**: Google Sheets URL 또는 xlsx 파일 경로와 함께 "
                   "생성할 탭 날짜를 알려주세요.\n"
                   "예) `아래 URL 기반으로 260625, 260702 탭 생성해줘`\n\n"
                   "**현지화 번역**: 번역할 파일 경로나 Google Sheets URL을 알려주세요."})
            q.put({"type": "done", "status": "chat"})
            return

        if result.get("response"):
            q.put({"type": "message", "content": result["response"]})

        # block=True: 필수 정보 누락 → 핸드오프 없이 PM이 사용자에게 질문하고 대기
        if result.get("block"):
            q.put({"type": "done", "status": "chat"})
            return

        if result.get("ready") and result.get("agent") == "event-planner":
            q.put({"type": "handoff", "agent": "event-planner", "params": result["params"]})
            q.put({"type": "done", "status": "handoff"})
        elif result.get("ready") and result.get("agent") == "game-localizer":
            _run_localizer_pipeline(result["params"], q)
        else:
            # 소스는 있지만 탭/마켓 등 부가 정보 누락 → 에이전트에 handoff해서 수집
            q.put({"type": "handoff", "agent": result["agent"],
                   "params": result.get("params", {})})
            q.put({"type": "done", "status": "handoff"})
        return

    # ── Ollama / Anthropic 모드: LLM 기반 라우팅 ─────────────────────────────
    messages = history + [{"role": "user", "content": user_msg}]

    if AI_MODE == "ollama":
        resp = _AI.chat.completions.create(
            model=OLLAMA_MODEL,
            messages=[{"role": "system", "content": PM_SYSTEM}] + messages,
            temperature=0.1,
        )
        raw = resp.choices[0].message.content.strip()
    else:
        # anthropic
        resp = _ANTHROPIC.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1024,
            system=PM_SYSTEM,
            messages=messages,
        )
        raw = resp.content[0].text.strip()

    # JSON 파싱 — 첫 번째 완전한 JSON 오브젝트 추출
    pm_result = None
    cleaned_pm = raw
    if "```" in cleaned_pm:
        part = cleaned_pm.split("```")[1]
        cleaned_pm = part[4:] if part.startswith("json") else part
    cleaned_pm = cleaned_pm.strip()

    _decoder = json.JSONDecoder()
    for _i, _ch in enumerate(cleaned_pm):
        if _ch == '{':
            try:
                pm_result, _ = _decoder.raw_decode(cleaned_pm, _i)
                break
            except json.JSONDecodeError:
                continue

    if pm_result is None:
        q.put({"type": "message", "content": raw})
        q.put({"type": "done", "status": "chat"})
        return

    result = pm_result
    agent   = result.get("agent")
    ready   = result.get("ready", False)
    params  = result.get("params", {})
    pm_resp = result.get("response", "")

    if pm_resp:
        q.put({"type": "message", "content": pm_resp})

    if ready and agent == "event-planner":
        # PM은 기본 정보만 수집 후 에이전트에게 위임
        q.put({"type": "handoff", "agent": "event-planner", "params": params})
        q.put({"type": "done", "status": "handoff"})
    elif ready and agent == "game-localizer":
        _run_localizer_pipeline(params, q)
    else:
        q.put({"type": "done", "status": "chat"})


# ══════════════════════════════════════════════════════════════════════════════
# 에이전트 대화 엔드포인트 (PM 위임 후 에이전트가 직접 사용자와 대화)
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/agent/chat", methods=["POST"])
def agent_chat():
    data     = request.json
    agent_id = data.get("agent", "")
    user_msg = data.get("message", "").strip()
    history  = data.get("history", [])
    context  = data.get("context", {})

    if not user_msg:
        return jsonify(ok=False, message="메시지가 비어 있습니다."), 400

    job_id = str(uuid.uuid4())[:8]
    q = _queue_module.Queue()
    _jobs[job_id] = q

    def run():
        try:
            if agent_id == "event-planner":
                _event_planner_agent(user_msg, history, context, q)
            else:
                q.put({"type": "error", "message": f"알 수 없는 에이전트: {agent_id}"})
                q.put({"type": "done", "status": "error"})
        except Exception as e:
            q.put({"type": "error", "message": f"에이전트 오류: {e}"})
            q.put({"type": "done", "status": "error"})

    threading.Thread(target=run, daemon=True).start()
    return jsonify(job_id=job_id)


# ═══════════════════════════════════════════════════════════════════════════════
# 에이전트 설정 저장/불러오기 (장르 학습 등)
# ═══════════════════════════════════════════════════════════════════════════════

def _load_agent_config() -> dict:
    try:
        if EP_CONFIG.exists():
            return json.loads(EP_CONFIG.read_text(encoding="utf-8"))
    except Exception:
        pass
    return {}

def _save_agent_config(data: dict):
    try:
        EP_CONFIG.parent.mkdir(parents=True, exist_ok=True)
        existing = _load_agent_config()
        existing.update(data)
        EP_CONFIG.write_text(
            json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    except Exception:
        pass


# ═══════════════════════════════════════════════════════════════════════════════
# 이벤트 기획 에이전트 — 상태 머신 헬퍼 (LLM 의존 없음, 완전 오프라인)
# ═══════════════════════════════════════════════════════════════════════════════

import re as _re_agent

_GENRE_DETECT = {
    "야구":   ["야구", "baseball", "kbo", "프로야구"],
    "축구":   ["축구", "soccer", "football", "k리그"],
    "농구":   ["농구", "basketball", "kbl"],
    "MMORPG": ["mmorpg", "rpg", "역할", "레이드", "던전"],
    "캐주얼": ["캐주얼", "퍼즐", "puzzle", "casual"],
    "스포츠": ["스포츠", "sport"],
}

_GENRE_KW_POOL = {
    "야구":   ["올스타", "전반기 결산", "순위 경쟁", "우승 도전", "끝내기 홈런",
               "퍼펙트게임", "레전드", "드래프트", "명예의 전당", "한국시리즈",
               "홈런왕", "타격왕", "방어율왕", "신인왕", "MVP"],
    "축구":   ["이적 시장", "챔피언스리그", "리그 개막", "국가대표", "유망주",
               "골든부트", "베스트 11", "FA", "주전 경쟁", "포메이션",
               "스쿼드", "이적료", "월드컵", "드리블왕", "감독"],
    "MMORPG": ["신규 클래스", "업데이트", "레이드", "보스", "던전",
               "장비 강화", "길드전", "시즌 패스", "영웅 성장", "전설 아이템",
               "서버 이전", "한정 코스튬", "공성전", "월드보스", "신규 직업"],
    "캐주얼": ["신규 스테이지", "업데이트", "협동 이벤트", "친구 초대", "랭킹전",
               "한정 스킨", "콜라보", "시즌 챌린지", "출석 보너스", "미션"],
    "스포츠": ["시즌 개막", "결승전", "우승", "베스트 선수", "특별 스쿼드",
               "한정 강화", "챔피언십", "올스타", "시즌 결산", "신규 선수"],
}

_SEASON_KW_POOL = {
    "1":  ["신년", "새해 맞이", "새 출발", "겨울 대전", "설날 준비"],
    "2":  ["발렌타인", "겨울 마무리", "봄 예고", "화이트데이 예고", "설날"],
    "3":  ["봄 개막", "개막전", "새 시즌 시작", "벚꽃", "새봄 챌린지"],
    "4":  ["봄 절정", "황금연휴 예고", "벚꽃 만개", "4월의 기적", "봄 대전"],
    "5":  ["황금연휴", "어린이날", "가정의 달", "봄 마무리", "전반기 중반"],
    "6":  ["초여름", "전반기 결산", "여름 예고", "얼리썸머", "6월의 열기"],
    "7":  ["여름 대축제", "한여름의 열기", "전반기 마무리", "올스타 시즌", "피서 특별 이벤트"],
    "8":  ["한여름 절정", "여름 마무리", "휴가 시즌", "후반기 개막", "8월의 열정"],
    "9":  ["가을 개막", "추석", "한가위", "포스트시즌 진입", "시즌 막바지"],
    "10": ["포스트시즌", "핼러윈", "가을 절정", "시즌 결산", "10월의 드라마"],
    "11": ["시즌 종료", "겨울 예고", "연말 준비", "드래프트 시즌", "FA 시장"],
    "12": ["크리스마스", "연말 결산", "올해의 선수", "새해 예고", "겨울 대축제"],
}

def _detect_genre(text: str) -> str:
    tl = text.lower().replace(" ", "")
    for genre, kws in _GENRE_DETECT.items():
        if any(kw.replace(" ", "") in tl for kw in kws):
            return genre
    return ""

def _build_keyword_suggestion_fallback(genre: str, target_month: str) -> tuple:
    """키워드 제안 — 하드코딩 풀 폴백 (LLM 실패 시 사용)"""
    genre_kws = _GENRE_KW_POOL.get(genre, _GENRE_KW_POOL.get("스포츠", []))
    try:
        month_str = str(int(target_month.split("-")[1]))
    except Exception:
        month_str = str(datetime.now().month)
    season_kws = _SEASON_KW_POOL.get(month_str, [])
    return _format_keyword_msg(genre, target_month, genre_kws, season_kws)

def _format_keyword_msg(genre: str, target_month: str, genre_kws: list, season_kws: list) -> tuple:
    all_kws = genre_kws + season_kws
    lines = [f"**{genre}** 장르 이벤트에 활용할 키워드를 제안합니다.\n"]
    lines.append(f"📌 **[{genre}] 장르 키워드**")
    for i, kw in enumerate(genre_kws, 1):
        lines.append(f"{i}. {kw}")
    lines.append(f"\n🗓️ **시즌 키워드** ({target_month} 기준)")
    for i, kw in enumerate(season_kws, 1):
        lines.append(f"{i}. {kw}")
    lines.append("\n사용할 키워드를 선택하거나 **'모두 사용'** 이라고 해주세요.\n번호 선택도 가능합니다. (예: 1,3,5,11,12)")
    return all_kws, "\n".join(lines)

def _llm_generate_keywords(genre: str, target_month: str) -> tuple:
    """LLM으로 키워드 생성. 실패 시 하드코딩 풀로 폴백."""
    if not HAS_AI:
        return _build_keyword_suggestion_fallback(genre, target_month)

    prompt = (
        f"모바일 {genre} 게임 이벤트 기획용 한국어 키워드를 JSON으로만 반환하라.\n"
        f"genre_keywords: {genre} 장르 이벤트에 쓰이는 한국어 키워드 15개\n"
        f"season_keywords: {target_month} 시즌에 어울리는 한국어 키워드 7개\n"
        f"규칙: 반드시 한국어만. 중국어·일본어·영어 금지. JSON만 출력.\n"
        f'출력 형식: {{"genre_keywords":["키워드1","키워드2",...], "season_keywords":["키워드1",...]}}'
    )

    try:
        if AI_MODE == "claude_cli":
            raw = _call_claude_cli(prompt, timeout=60)
        elif AI_MODE == "ollama":
            resp = _AI.chat.completions.create(
                model=OLLAMA_MODEL,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.3,
                extra_body={"format": "json"},   # ← Ollama JSON 강제 모드
            )
            raw = resp.choices[0].message.content.strip()
        else:
            resp = _ANTHROPIC.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=512,
                messages=[{"role": "user", "content": prompt}],
            )
            raw = resp.content[0].text.strip()

        # JSON 파싱
        if "```" in raw:
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        data = json.loads(raw.strip())
        genre_kws  = data.get("genre_keywords", [])
        season_kws = data.get("season_keywords", [])

        # 한국어 검증 (중국어·일본어 포함 키워드 제거)
        def _is_korean_only(s: str) -> bool:
            return not bool(_re_agent.search(r'[一-鿿぀-ヿ]', s))
        genre_kws  = [k for k in genre_kws  if isinstance(k, str) and _is_korean_only(k)]
        season_kws = [k for k in season_kws if isinstance(k, str) and _is_korean_only(k)]

        if len(genre_kws) < 3:   # 너무 적으면 폴백
            raise ValueError("insufficient keywords")

        return _format_keyword_msg(genre, target_month, genre_kws, season_kws)

    except Exception:
        # LLM 실패 → 하드코딩 풀 폴백
        return _build_keyword_suggestion_fallback(genre, target_month)

def _parse_keyword_selection(user_msg: str, suggested: list) -> list:
    if any(w in user_msg for w in ["모두", "전부", "다 ", "다사용", "전체"]):
        return suggested
    nums = _re_agent.findall(r'\b(\d+)\b', user_msg)
    if nums:
        selected = []
        for n in nums:
            idx = int(n) - 1
            if 0 <= idx < len(suggested):
                selected.append(suggested[idx])
        if selected:
            return selected
    # 키워드 이름 직접 포함
    selected = [kw for kw in suggested if kw in user_msg]
    return selected if selected else suggested

def _parse_ref_tabs(user_msg: str, new_tabs: list, available_tabs: list) -> list:
    # 방법 1: 화살표 패턴
    pairs = _re_agent.findall(r'(\d{6})\s*[→\->]\s*(\d{6})', user_msg)
    if pairs:
        return [p[1] for p in pairs[:len(new_tabs)]]
    # 방법 2: 한국어 패턴 "260625는 260618을 참조"
    kr = _re_agent.findall(r'(\d{6})[은는이가]?\s*(\d{6})[을를]?\s*참조', user_msg)
    if kr:
        return [p[1] for p in kr[:len(new_tabs)]]
    # 방법 3: new_tab 직후 6자리 숫자
    result = []
    for nt in new_tabs:
        idx = user_msg.find(nt)
        if idx >= 0:
            m = _re_agent.search(r'\b(\d{6})\b', user_msg[idx + len(nt):])
            if m:
                result.append(m.group(1))
    if len(result) == len(new_tabs):
        return result
    # 방법 4: new_tabs 제외한 모든 6자리 숫자
    all_nums = [n for n in _re_agent.findall(r'\b(\d{6})\b', user_msg) if n not in new_tabs]
    if all_nums:
        while len(all_nums) < len(new_tabs):
            all_nums.append(all_nums[-1])
        return all_nums[:len(new_tabs)]
    return []

def _auto_ref_tabs(new_tabs: list, available_tabs: list) -> list:
    """new_tabs 직전에 가장 가까운 탭을 자동 매핑"""
    result = []
    for nt in new_tabs:
        best = available_tabs[0] if available_tabs else nt
        try:
            nd = datetime(2000 + int(nt[:2]), int(nt[2:4]), int(nt[4:6]))
            candidates = []
            for at in available_tabs:
                try:
                    ad = datetime(2000 + int(at[:2]), int(at[2:4]), int(at[4:6]))
                    if ad < nd:
                        candidates.append((ad, at))
                except Exception:
                    pass
            if candidates:
                best = sorted(candidates, reverse=True)[0][1]
        except Exception:
            pass
        result.append(best)
    return result


def _get_tab_event_types(wb, tab: str) -> set:
    """워크시트 B열에서 이벤트 타입 집합 추출."""
    from generate_event_names import detect_event_type, parse_section_title
    import re as _re2
    ws = wb[tab]
    types = set()
    for row in ws.iter_rows():
        for cell in row:
            if getattr(cell, "column", None) != 2:
                continue
            val = getattr(cell, "value", None)
            if not val:
                continue
            t = parse_section_title(str(val).strip())
            if t:
                etype = detect_event_type(t)
                if etype != "기타":
                    types.add(etype)
    return types


def _analyze_event_composition(source: str, ref_tabs: list, available_tabs: list) -> dict | None:
    """
    동일 타입 이력 탭들의 이벤트 구성 패턴을 분석하여 선택적 이벤트 추천 메시지 반환.

    반환:
      None → 분석 불가 (이력 부족)
      {
        "message": str,          # 사용자에게 보여줄 추천 메시지
        "analysis": {            # context에 저장할 분석 데이터
            "mandatory": [...],
            "optional_present": [...],
            "optional_absent": [...],
        }
      }
    """
    if not source or not Path(source).exists():
        return None

    try:
        import openpyxl as _xl
        import re as _re2
        from collections import Counter as _Counter

        wb = _xl.load_workbook(source, read_only=True, data_only=True)
        date_tabs = sorted([s for s in wb.sheetnames if _re2.fullmatch(r"\d{6}", s)])

        if not ref_tabs:
            return None

        # 참조 탭의 이벤트 타입 파악
        ref_tab = ref_tabs[0]
        if ref_tab not in wb.sheetnames:
            return None
        ref_events = _get_tab_event_types(wb, ref_tab)

        # 참조 탭과 같은 타입의 이력 탭 탐색
        # (A타입: 포인트레이스/룰렛/빙고 / B타입: 출석/응모권/미션 등)
        A_TYPE_MARKER = {"포인트레이스", "룰렛이벤트", "빙고이벤트"}
        is_a_type = bool(ref_events & A_TYPE_MARKER)

        same_type_tabs = []
        for tab in date_tabs:
            if tab == ref_tab:
                continue
            try:
                etypes = _get_tab_event_types(wb, tab)
                if not etypes:
                    continue
                tab_is_a = bool(etypes & A_TYPE_MARKER)
                if tab_is_a == is_a_type:
                    same_type_tabs.append((tab, etypes))
            except Exception:
                continue

        wb.close()

        if len(same_type_tabs) < 2:
            # 이력 탭이 2개 미만 → 패턴 분석 불가
            return None

        total = len(same_type_tabs) + 1  # 참조 탭 포함

        # 전체 등장 이벤트 빈도 계산 (참조 탭 포함)
        all_events_seen: list = list(ref_events)
        for _, etypes in same_type_tabs:
            all_events_seen.extend(etypes)
        freq = _Counter(all_events_seen)

        # 분류
        mandatory       = []   # 항상 등장 (100%)
        optional_present = []  # 가끔 등장 + 현재 참조 탭에 있음
        optional_absent  = []  # 과거에 있었지만 현재 참조 탭에 없음

        for etype, cnt in sorted(freq.items(), key=lambda x: -x[1]):
            rate = cnt / total
            if rate >= 1.0:
                mandatory.append({"type": etype, "rate": rate, "count": cnt})
            elif etype in ref_events:
                optional_present.append({"type": etype, "rate": rate, "count": cnt})
            else:
                optional_absent.append({"type": etype, "rate": rate, "count": cnt})

        # 추천 대상 없으면 None
        if not optional_present and not optional_absent:
            return None

        # 추천 메시지 생성
        history_tabs_str = ", ".join(t for t, _ in same_type_tabs[:5])
        lines = [
            f"**이벤트 구성 패턴 분석** (분석 탭: {history_tabs_str})\n",
            f"📋 **필수 이벤트** (매 주기 포함): "
            + ", ".join(e["type"] for e in mandatory) if mandatory else "",
        ]

        if optional_present:
            lines.append("\n⚠️ **선택적 이벤트** (현재 포함됨, 가끔 빠짐)")
            for e in optional_present:
                bar = "█" * int(e["rate"] * 5) + "░" * (5 - int(e["rate"] * 5))
                lines.append(f"  - **{e['type']}** [{bar}] {e['count']}/{total}회 ({e['rate']:.0%})")

        if optional_absent:
            lines.append("\n💡 **추가 가능한 이벤트** (과거에 포함된 적 있음, 현재 없음)")
            for e in optional_absent:
                bar = "█" * int(e["rate"] * 5) + "░" * (5 - int(e["rate"] * 5))
                lines.append(f"  - **{e['type']}** [{bar}] {e['count']}/{total}회 ({e['rate']:.0%})")
            lines.append("\n위 이벤트를 이번 시트에 추가하시겠어요?")
            lines.append("추가하려면 이벤트명을 입력하거나 **'그대로 진행'**을 말씀해 주세요.")
        else:
            lines.append("\n이대로 파이프라인을 시작할까요? (확인 또는 '그대로 진행')")

        message = "\n".join(l for l in lines if l)
        return {
            "message": message,
            "analysis": {
                "mandatory":        mandatory,
                "optional_present": optional_present,
                "optional_absent":  optional_absent,
            }
        }

    except Exception as e:
        print(f"[이벤트 구성 분석 오류] {e}")
        return None


def _event_planner_agent(user_msg: str, history: list, context: dict, q: _queue_module.Queue):
    """이벤트 기획 에이전트 — 서버 상태 머신 기반 (LLM 불필요, 완전 오프라인)"""

    # ── 소스 파일에서 available_tabs 추출 ─────────────────────────────────
    available_tabs: list = []
    source = context.get("source_path", "")

    # Google Sheets URL → 로컬 캐시 경로로 변환
    if source:
        try:
            from gdrive_utils import is_google_url
            if is_google_url(source):
                source_local, _ = resolve_source(source)
            else:
                source_local = source
            # 변환된 로컬 경로를 context에도 저장 (파이프라인에서 재사용)
            if Path(source_local).exists():
                context["source_path"] = source_local
                source = source_local
        except Exception:
            pass

    if source and Path(source).exists():
        try:
            import openpyxl as _xl
            wb = _xl.load_workbook(source, read_only=True)
            available_tabs = sorted(
                [s for s in wb.sheetnames if _re_agent.match(r'^\d{6}$', s)],
                reverse=True
            )[:15]   # 최근 15개
            wb.close()
        except Exception:
            pass

    new_tabs = context.get("new_tabs", [])
    try:
        target_month = datetime.strptime("20" + new_tabs[0], "%Y%m%d").strftime("%Y-%m") if new_tabs else datetime.now().strftime("%Y-%m")
    except Exception:
        target_month = datetime.now().strftime("%Y-%m")

    step = context.get("_agent_step", "start")

    def _done(msg: str = ""):
        if msg:
            q.put({"type": "message", "content": msg})
        q.put({"type": "context_update", "context": context})
        q.put({"type": "done", "status": "chat"})

    # ── START: 에이전트 시작 ───────────────────────────────────────────────
    if user_msg == "__agent_start__" or step == "start":
        # ① 프로젝트 설정에서 장르·마켓 자동 주입
        project_id = context.get("project_id", "")
        if project_id:
            projects   = _load_projects()
            proj_cfg   = projects.get(project_id, {})
            if proj_cfg.get("genre") and not context.get("genre"):
                context["genre"] = proj_cfg["genre"]
            if proj_cfg.get("market") and not context.get("market"):
                context["market"] = proj_cfg["market"]

        # ② 저장된 장르 불러오기 (프로젝트 설정 → 학습 값 순)
        _cfg = _load_agent_config()
        genre = context.get("genre") or _cfg.get("default_genre", "")

        if genre:
            # 장르 확정 → 장르 질문 생략, 바로 키워드 단계
            context["genre"] = genre
            context["_agent_step"] = "wait_keywords"
            proj_label = f" [{project_id}]" if project_id else ""
            q.put({"type": "message", "content": f"**{genre}** 장르{proj_label}로 진행합니다. 키워드를 생성하는 중..."})
            q.put({"type": "context_update", "context": context})
            all_kws, kw_msg = _llm_generate_keywords(genre, target_month)
            context["_suggested_keywords"] = all_kws
            _done(kw_msg)
        else:
            context["_agent_step"] = "wait_genre"
            proj_label = f" [{project_id}]" if project_id else ""
            _done(f"어떤 장르의 게임인가요?{proj_label}\n(예: 야구, 축구, MMORPG, 캐주얼, 퍼즐)")
        return

    # ── STEP 1: 장르 수집 (학습된 장르 없을 때만 실행) ─────────────────────
    if step == "wait_genre":
        genre = _detect_genre(user_msg) or context.get("genre", "")
        if not genre:
            _done("장르를 인식하지 못했습니다. 다시 알려주세요.\n(예: 야구, 축구, MMORPG, 캐주얼, 퍼즐)")
            return
        context["genre"] = genre
        context["_agent_step"] = "wait_keywords"
        # 장르 학습 저장 (다음 요청부터 질문 생략)
        _save_agent_config({"default_genre": genre})
        q.put({"type": "message", "content": f"**{genre}** 장르로 설정됐습니다. 키워드를 생성하는 중..."})
        q.put({"type": "context_update", "context": context})
        all_kws, kw_msg = _llm_generate_keywords(genre, target_month)
        context["_suggested_keywords"] = all_kws
        _done(kw_msg)
        return

    # ── STEP 2: 키워드 수집 ────────────────────────────────────────────────
    if step == "wait_keywords":
        suggested = context.get("_suggested_keywords", [])
        keywords = _parse_keyword_selection(user_msg, suggested)
        context["keywords"] = keywords
        kw_preview = ", ".join(keywords[:5]) + ("..." if len(keywords) > 5 else "")

        # 참조 가능한 탭이 없으면 소스 경로를 먼저 물어본다
        if not available_tabs:
            context["_agent_step"] = "wait_source_path"
            q.put({"type": "message", "content": (
                f"**{len(keywords)}개** 키워드가 저장됐습니다 ({kw_preview})\n\n"
                "참조할 수 있는 이전 탭이 없습니다.\n"
                "복사 원본으로 사용할 **소스 파일 경로** 또는 **Google Sheets URL**을 입력해주세요.\n\n"
                "예시:\n"
                "- `C:\\Users\\...\\파일명.xlsx`\n"
                "- `https://docs.google.com/spreadsheets/d/...`"
            )})
            q.put({"type": "context_update", "context": context})
            q.put({"type": "done", "status": "chat"})
            return

        context["_agent_step"] = "wait_ref_tabs"
        msg = f"**{len(keywords)}개** 키워드가 저장됐습니다 ({kw_preview})\n\n아래에서 각 탭의 **참조 탭**을 클릭해서 선택해주세요."
        q.put({"type": "message", "content": msg})
        # 클릭형 탭 선택 UI 이벤트 — 프론트엔드에서 시각적 선택기로 렌더링
        q.put({"type": "tab_selector",
               "new_tabs": new_tabs,
               "available_tabs": available_tabs[:15]})
        q.put({"type": "context_update", "context": context})
        q.put({"type": "done", "status": "chat"})
        return

    # ── STEP 2-B: 소스 경로 수집 (available_tabs 없을 때 진입) ───────────────
    if step == "wait_source_path":
        raw = user_msg.strip().strip('"').strip("'")
        # Google Sheets URL 또는 로컬 경로 판별
        try:
            from gdrive_utils import is_google_url
            if is_google_url(raw):
                source_local, _ = resolve_source(raw)
            else:
                source_local = raw
        except Exception:
            source_local = raw

        if not Path(source_local).exists():
            _done(
                f"파일을 찾을 수 없습니다: `{source_local}`\n\n"
                "올바른 경로 또는 Google Sheets URL을 다시 입력해주세요."
            )
            return

        # 소스 갱신 + available_tabs 재추출
        context["source_path"] = source_local
        source = source_local
        try:
            import openpyxl as _xl2
            wb2 = _xl2.load_workbook(source_local, read_only=True)
            available_tabs = sorted(
                [s for s in wb2.sheetnames if _re_agent.match(r'^\d{6}$', s)],
                reverse=True
            )[:15]
            wb2.close()
        except Exception:
            available_tabs = []

        if not available_tabs:
            _done(
                f"소스 파일에서 날짜 형식 탭(YYMMDD)을 찾을 수 없습니다.\n"
                "탭이 포함된 다른 파일 경로를 입력해주세요."
            )
            return

        context["_agent_step"] = "wait_ref_tabs"
        q.put({"type": "message", "content": (
            f"소스 파일을 확인했습니다. 사용 가능한 탭: **{', '.join(available_tabs[:5])}**{'...' if len(available_tabs) > 5 else ''}\n\n"
            "아래에서 각 탭의 **참조 탭**을 클릭해서 선택해주세요."
        )})
        q.put({"type": "tab_selector",
               "new_tabs": new_tabs,
               "available_tabs": available_tabs})
        q.put({"type": "context_update", "context": context})
        q.put({"type": "done", "status": "chat"})
        return

    # ── STEP 3: 참조 탭 수집 → 파이프라인 실행 ─────────────────────────────
    if step == "wait_ref_tabs":
        ref_tabs = _parse_ref_tabs(user_msg, new_tabs, available_tabs)

        # "작업해줘" 등 트리거 + 파싱 실패 → 자동 매핑
        _triggers = ("작업", "시작", "진행", "실행", "만들어", "생성", "해줘", "해봐", "확인", "맞아", "정확해", "ㅇㅇ")
        if not ref_tabs and any(w in user_msg for w in _triggers) and available_tabs:
            ref_tabs = _auto_ref_tabs(new_tabs, available_tabs)

        if not ref_tabs:
            avail_str = ", ".join(available_tabs[:10]) if available_tabs else "확인 불가"
            _done(
                f"참조 탭을 인식하지 못했습니다. 다시 알려주세요.\n\n"
                f"📅 생성할 탭: {', '.join(new_tabs)}\n"
                f"📂 사용 가능한 탭: {avail_str}\n\n"
                f"예시: {new_tabs[0] if new_tabs else 'YYMMDD'}→{available_tabs[0] if available_tabs else 'YYMMDD'}"
            )
            return

        context["ref_tabs"] = ref_tabs
        context["_agent_step"] = "wait_event_composition"
        q.put({"type": "context_update", "context": context})

        # ── 이벤트 구성 패턴 분석 후 추천 메시지 생성 ──────────────────────
        composition_msg = _analyze_event_composition(source, ref_tabs, available_tabs)
        if composition_msg:
            context["_composition_analysis"] = composition_msg["analysis"]
            _done(composition_msg["message"])
        else:
            # 분석 데이터 부족 → 바로 파이프라인 실행
            context["_agent_step"] = "done"
            q.put({"type": "context_update", "context": context})
            q.put({"type": "message", "content": "모든 정보가 준비됐습니다! 이벤트 기획 파이프라인을 시작합니다."})
            _run_event_pipeline(context, q)
        return

    # ── STEP 4: 이벤트 구성 추천 확인 → 파이프라인 실행 ───────────────────
    if step == "wait_event_composition":
        context["_agent_step"] = "done"
        q.put({"type": "context_update", "context": context})

        # 사용자 답변 파싱 → 이벤트 구성 지시 도출
        analysis = context.get("_composition_analysis", {})
        optional_present = [e["type"] for e in analysis.get("optional_present", [])]
        optional_absent  = [e["type"] for e in analysis.get("optional_absent", [])]

        events_to_remove = []
        events_to_add    = []

        # ── 섹션별 분리 파싱 ─────────────────────────────────────────────
        # 사용자 메시지를 줄별로 나눠 각 카테고리 컨텍스트를 분리한다.
        # "1. 필수 이벤트 : 모두 진행"의 "모두 진행"이
        # "선택적 이벤트" 처리에 오염되지 않도록 한다.
        def _extract_section_text(msg: str, markers: list) -> str:
            """markers 중 하나가 포함된 줄 이후의 텍스트(다음 번호 항목 전까지) 반환."""
            import re as _re2
            lines = msg.splitlines()
            for i, line in enumerate(lines):
                if any(m in line for m in markers):
                    # 해당 줄 + 다음 번호 항목 전까지
                    section_lines = [line]
                    for j in range(i + 1, len(lines)):
                        if _re2.match(r'^\s*\d+[.):\s]', lines[j]):
                            break
                        section_lines.append(lines[j])
                    return " ".join(section_lines).lower()
            return ""

        selective_text = _extract_section_text(user_msg, ["선택적 이벤트", "선택 이벤트", "2."])
        absent_text    = _extract_section_text(user_msg, ["추가 가능", "추가이벤트", "추가 이벤트", "3."])

        # 섹션 분리 실패 시 (구조 없는 자유형 입력) → 전체 메시지 사용
        if not selective_text:
            selective_text = user_msg.lower()
        if not absent_text:
            absent_text = user_msg.lower()

        _ALL_KEYWORDS = ["모두", "전부", "다 ", "전체", "진행", "적용", "추가"]

        # ── 선택적 이벤트 (optional_present) 처리 ────────────────────────
        if optional_present:
            # 선택적 섹션 텍스트 기준으로 keep_all 판단
            keep_all = any(k in selective_text for k in ["모두 진행", "전부 진행", "그대로", "모두 유지", "전체 진행"])
            if not keep_all:
                mentioned = [e for e in optional_present
                             if e.replace("이벤트", "").lower() in selective_text
                             or e.lower() in selective_text]
                if mentioned:
                    events_to_remove = [e for e in optional_present if e not in mentioned]
                # else: 아무것도 언급 안 했으면 그대로 유지

        # ── 추가 가능 이벤트 처리 ─────────────────────────────────────────
        # optional_absent에 있는 이벤트 + 사용자가 명시한 이벤트 모두 포함
        _ALL_ETYPES = [
            "쿠폰이벤트", "승부예측", "야구공찾기", "PvP이벤트",
            "출석이벤트", "응모권이벤트", "미션이벤트", "교환소이벤트",
            "오더경쟁이벤트", "포인트레이스", "룰렛이벤트", "빙고이벤트",
        ]
        absent_opt_list = list(optional_absent)  # 이미 [str, ...] 형태

        add_all = (
            any(k in absent_text for k in [
                "모두 적용", "전부 추가", "모두 추가", "다 추가", "다 적용",
                "다 넣", "전부 넣", "모두 넣", "다 포함", "전부 포함",
            ])
            or ("모두" in absent_text and any(k in absent_text for k in ["진행", "적용", "추가", "넣"]))
            or ("전부" in absent_text and any(k in absent_text for k in ["진행", "적용", "추가", "넣"]))
            or ("다" in absent_text and any(k in absent_text for k in ["넣어", "추가해", "적용해", "포함"]))
        )
        if add_all and absent_opt_list:
            events_to_add = list(absent_opt_list)
        else:
            # optional_absent 필터링
            from_absent = [e for e in absent_opt_list
                           if e.replace("이벤트", "").lower() in absent_text
                           or e.lower() in absent_text]
            # 사용자가 명시했지만 optional_absent에 없는 이벤트도 추가 (분석 누락 보완)
            explicitly_named = [e for e in _ALL_ETYPES
                                if e not in absent_opt_list
                                and (e.replace("이벤트", "").lower() in absent_text
                                     or e.lower() in absent_text)]
            events_to_add = from_absent + [e for e in explicitly_named if e not in from_absent]

        context["_events_to_remove"] = events_to_remove
        context["_events_to_add"]    = events_to_add
        q.put({"type": "context_update", "context": context})

        summary_parts = []
        if events_to_remove:
            summary_parts.append(f"제거: {', '.join(events_to_remove)}")
        if events_to_add:
            summary_parts.append(f"추가: {', '.join(events_to_add)}")
        summary = " / ".join(summary_parts) if summary_parts else "변경 없음"

        q.put({"type": "message", "content": f"확인했습니다! ({summary}) 이벤트 기획 파이프라인을 시작합니다."})
        _run_event_pipeline(context, q)
        return

    # ── 완료 후 추가 입력 ──────────────────────────────────────────────────
    _done("파이프라인이 이미 실행됐습니다. 새 이벤트 기획을 원하시면 PM에게 다시 요청해 주세요.")


# ── 이벤트 기획 전체 파이프라인 ───────────────────────────────────────────────

def _run_event_pipeline(params: dict, q: _queue_module.Queue):
    source          = params.get("source_path", "")
    new_tabs        = params.get("new_tabs", [])
    ref_tabs        = params.get("ref_tabs", [])
    market          = params.get("market", "글로벌")
    genre           = params.get("genre", "")
    keywords        = params.get("keywords", [])
    output_filename = params.get("output_filename",
                                 f"이벤트기획_{datetime.now().strftime('%Y%m%d')}.xlsx")

    # Google Sheets / Drive URL → 로컬 xlsx 다운로드
    from gdrive_utils import is_google_url
    if source and is_google_url(source):
        try:
            q.put({"type": "step", "name": "Drive 파일 다운로드"})
            source, _ = resolve_source(source)
            q.put({"type": "step_ok", "message": f"Drive 파일 다운로드 완료: {Path(source).name}"})
        except Exception as e:
            q.put({"type": "error", "message": f"Drive 파일 다운로드 실패: {e}"})
            q.put({"type": "done", "status": "error"})
            return

    if not source or not Path(source).exists():
        q.put({"type": "error", "message": f"소스 파일을 찾을 수 없습니다: {source}"})
        q.put({"type": "done", "status": "error"})
        if _LEARNING:
            try:
                _LEARNING.record_session({"agent": "event-planner", "genre": params.get("genre",""), "market": params.get("market",""), "new_tabs": new_tabs, "ref_tabs": ref_tabs, "keywords": params.get("keywords",[]), "success": False, "error_message": f"소스 파일 없음: {source}", "duration_seconds": None, "reward_stats": {}, "output_path": "", "user_feedback": None})
            except Exception:
                pass
        return
    if not new_tabs:
        q.put({"type": "error", "message": "생성할 탭명이 없습니다."})
        q.put({"type": "done", "status": "error"})
        return

    # 탭명 정규화: 4자리 MMDD → 6자리 YYMMDD
    def _norm_tab(t: str) -> str:
        t = t.strip()
        if len(t) == 4 and t.isdigit():
            return str(datetime.now().year)[2:] + t
        return t

    new_tabs = [_norm_tab(t) for t in new_tabs]
    if ref_tabs:
        ref_tabs = [_norm_tab(t) for t in ref_tabs]

    output_path = EP_FILE / output_filename

    def step(name: str):
        q.put({"type": "step", "name": name})

    def ok_step(msg: str):
        q.put({"type": "step_ok", "message": msg})

    def fail_step(msg: str):
        q.put({"type": "step_fail", "message": msg})

    # current_project.json 먼저 기록 → 스크립트들이 올바른 경로 사용
    write_current_project(source_xlsx=source)
    EP_WORK.mkdir(parents=True, exist_ok=True)
    EP_FILE.mkdir(parents=True, exist_ok=True)

    # 사전 파일 준비
    config = {
        "genre": genre,
        "target_month": datetime.now().strftime("%Y-%m"),
        "genre_phrases": keywords,
        "event_name_replacements": {tab: [] for tab in new_tabs},
    }
    (EP_WORK / "event_names_config.json").write_text(
        json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (HANDOFF_DIR / "event-planner_input.json").write_text(
        json.dumps({
            "source_path": source, "market": market, "genre": genre,
            "new_tab_names": new_tabs, "ref_tab_names": ref_tabs,
            "output_path": str(output_path), "genre_phrases": keywords,
        }, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    # ── 1단계: 이벤트 제목 패턴 생성 ───────────────────────────────────────
    step("이벤트 제목 패턴 생성")
    try:
        _dt0 = datetime.strptime("20" + new_tabs[0], "%Y%m%d")
        _target_month = _dt0.strftime("%Y-%m")
    except Exception:
        _target_month = datetime.now().strftime("%Y-%m")
    gen_cmd = [
        sys.executable, str(SCRIPTS_DIR / "generate_event_names.py"),
        "--source",       source,
        "--new-tabs",     ",".join(new_tabs),
        "--ref-tabs",     ",".join(ref_tabs) if ref_tabs else ",".join(new_tabs),
        "--target-month", _target_month,
        "--genre",        genre,
        "--phrases",      ",".join(keywords) if keywords else "",
        "--work-dir",     str(EP_WORK),
    ]
    ok_gen, log_gen = run_script(gen_cmd)
    if ok_gen:
        ok_step(f"제목 패턴 생성 완료\n{log_gen[:300]}")
    else:
        ok_step(f"제목 패턴 생성 스킵 (날짜만 치환)\n{log_gen[:200]}")

    # ── 2단계: 탭 생성 ──────────────────────────────────────────────────────
    step("탭 생성")
    events_to_remove = params.get("_events_to_remove", [])
    events_to_add    = params.get("_events_to_add", [])
    cmd = [sys.executable, str(SCRIPTS_DIR / "create_tabs.py"),
           source, str(output_path), ",".join(new_tabs)]
    if ref_tabs:
        cmd.append(",".join(ref_tabs))
    if events_to_remove:
        cmd += ["--remove-events", ",".join(events_to_remove)]
    if events_to_add:
        cmd += ["--add-events", ",".join(events_to_add)]
    ok, log = run_script(cmd)
    if not ok:
        fail_step(f"탭 생성 실패:\n{log[:400]}")
        q.put({"type": "done", "status": "error"})
        return
    ok_step(f"탭 생성 완료 ({', '.join(new_tabs)})")

    # ── 2단계: 소스 탭 보상 스캔 ────────────────────────────────────────────
    step("소스 보상 스캔")
    ok, log = run_script([sys.executable, str(SCRIPTS_DIR / "scan_rewards_by_event.py"),
                          source, str(EP_WORK / "reward_by_event.json")])
    if not ok:
        fail_step(f"소스 보상 스캔 실패:\n{log[:400]}")
        q.put({"type": "done", "status": "error"})
        return
    ok_step("소스 탭 보상 이력 스캔 완료")

    # ── 3단계: 신규 탭 보상 스캔 ────────────────────────────────────────────
    step("신규 탭 보상 스캔")
    ok, log = run_script([sys.executable, str(SCRIPTS_DIR / "scan_rewards_by_event.py"),
                          str(output_path), str(EP_WORK / "reward_new_tabs.json")])
    if not ok:
        fail_step(f"신규 탭 보상 스캔 실패:\n{log[:400]}")
        q.put({"type": "done", "status": "error"})
        return
    ok_step("신규 탭 보상 구조 스캔 완료")

    # ── 3.5단계: 히스토리 vs 신규 탭 보상 비교 ──────────────────────────────
    step("보상 히스토리 비교")
    cmp_out = str(EP_WORK / "reward_comparison.json")
    ok_cmp, log_cmp = run_script([
        sys.executable, str(SCRIPTS_DIR / "compare_rewards.py"),
        "--source-scan", str(EP_WORK / "reward_by_event.json"),
        "--new-scan",    str(EP_WORK / "reward_new_tabs.json"),
        "--out",         cmp_out,
    ])
    if ok_cmp and Path(cmp_out).exists():
        try:
            cmp_data = json.loads(Path(cmp_out).read_text(encoding="utf-8"))
            summary  = cmp_data.get("summary", {})
            lines = []
            for tab, counts in summary.items():
                parts = []
                if counts.get("유지"):  parts.append(f"유지 {counts['유지']}개")
                if counts.get("신규"):  parts.append(f"신규 ★{counts['신규']}개")
                if counts.get("증가"):  parts.append(f"증가 ↑{counts['증가']}개")
                if counts.get("감소"):  parts.append(f"감소 ↓{counts['감소']}개")
                if counts.get("제거"):  parts.append(f"제거 ✗{counts['제거']}개")
                lines.append(f"{tab}: {', '.join(parts)}")
            ok_step("보상 비교 완료\n" + "\n".join(lines))
            q.put({"type": "reward_comparison", "data": cmp_data})
        except Exception:
            ok_step("보상 비교 완료 (상세 파싱 실패)")
    else:
        ok_step("보상 비교 스킵 (블로킹 안 함)")

    # ── 4단계: 보상 추천 ────────────────────────────────────────────────────
    step("보상 추천 (Kendall-tau)")
    ok, log = run_script([sys.executable, str(SCRIPTS_DIR / "recommend_rewards.py"),
                          "--per-event"])
    if not ok:
        fail_step(f"보상 추천 실패:\n{log[:400]}")
        q.put({"type": "done", "status": "error"})
        return
    run_script([sys.executable, str(SCRIPTS_DIR / "_prep_sequential_review.py")])
    ok_step("보상 추천 및 리뷰 큐 준비 완료")

    # ── 5단계: 패턴 갭 분석 ─────────────────────────────────────────────────
    step("이벤트 패턴 갭 분석")
    ok, log = run_script([sys.executable, str(SCRIPTS_DIR / "analyze_event_patterns.py"),
                          source, str(output_path), ",".join(new_tabs)])
    ok_step("패턴 갭 분석 완료" if ok else "패턴 분석 스킵 (블로킹 안 함)")

    # ── 6단계: 학습 저장 (다음 세션에서 보상 추천 정확도 향상) ──────────────
    run_script([sys.executable, str(SCRIPTS_DIR / "save_learning.py")])

    # ── 통계 집계 ────────────────────────────────────────────────────────────
    stats = {}
    qfile = EP_WORK / "reward_review_queue.json"
    if qfile.exists():
        try:
            qd = json.loads(qfile.read_text(encoding="utf-8"))
            stats = {"total": qd.get("total_sections", 0),
                     "changes": qd.get("sections_with_changes", 0)}
        except Exception:
            pass

    q.put({
        "type": "done",
        "status": "success",
        "output_path": str(output_path),
        "output_name": output_path.name,
        "download_url": f"/api/event/download?path={output_path}",
        "stats": stats,
        "summary": (
            f"이벤트 기획 완료!\n\n"
            f"• 출력 파일: {output_path.name}\n"
            f"• 보상 검토 필요: {stats.get('changes', 0)}개 섹션\n\n"
            "보상을 적용하려면 **이벤트 기획 → ② 보상 추천** 탭에서 검토 후 변경을 적용하세요.\n"
            "Google Sheets 업로드는 **④ 최종 출력** 탭에서 실행할 수 있습니다."
        ),
    })

    # ── 자기학습: 세션 결과 기록 ──────────────────────────────────────────
    if _LEARNING:
        try:
            _pipeline_end = datetime.now()
            _reward_stats = {}
            if Path(cmp_out).exists() if 'cmp_out' in dir() else False:
                try:
                    _cmp = json.loads(Path(cmp_out).read_text(encoding="utf-8"))
                    _reward_stats = _cmp.get("summary", {})
                except Exception:
                    pass
            _LEARNING.record_session({
                "agent": "event-planner",
                "genre": genre,
                "market": market,
                "new_tabs": new_tabs,
                "ref_tabs": ref_tabs,
                "keywords": keywords,
                "success": True,
                "error_message": None,
                "duration_seconds": None,
                "reward_stats": _reward_stats,
                "output_path": str(output_path),
                "user_feedback": None
            })
            # 미적용 개선 제안이 있으면 프론트엔드에 알림
            pending = [i for i in _LEARNING.data.get("improvements", []) if not i.get("applied")]
            if pending:
                q.put({"type": "message", "content": f"💡 학습된 개선 제안 {len(pending)}건이 있습니다: {pending[-1]['insight']}"})
        except Exception as _lerr:
            print(f"[학습] 기록 실패: {_lerr}")


# ── 현지화 번역 파이프라인 ────────────────────────────────────────────────────

def _run_localizer_pipeline(params: dict, q: _queue_module.Queue):
    gsheet_url = params.get("gsheet_url", "")
    excel_path = params.get("excel_path", "")

    if not gsheet_url and not excel_path:
        q.put({"type": "message", "content":
               "용어집 소스가 필요합니다.\n"
               "Google Sheets URL 또는 Excel 파일 경로를 알려주세요."})
        q.put({"type": "done", "status": "chat"})
        return

    def step(name: str):
        q.put({"type": "step", "name": name})

    def ok_step(msg: str):
        q.put({"type": "step_ok", "message": msg})

    def fail_step(msg: str):
        q.put({"type": "step_fail", "message": msg})

    # glossary_config.json 저장
    (GL_DIR / "glossary_config.json").write_text(
        json.dumps({"gsheet_url": gsheet_url, "excel_path": excel_path},
                   ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (HANDOFF_DIR / "game-localizer_input.json").write_text(
        json.dumps(params, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    if gsheet_url:
        step("Google Sheets 용어집 다운로드")
        ok, log = run_script([sys.executable, str(SCRIPTS_DIR / "fetch_gsheet.py"),
                               gsheet_url])
        if ok:
            ok_step("Google Sheets 용어집 다운로드 완료")
        else:
            fail_step(f"Google Sheets 접근 실패 (Excel 단독 모드로 계속)\n{log[:200]}")

    if excel_path and Path(excel_path).exists():
        step("Excel 용어집 읽기")
        ok, log = run_script([sys.executable, str(SCRIPTS_DIR / "read_excel.py"),
                               excel_path])
        if ok:
            ok_step("Excel 용어집 읽기 완료")
        else:
            fail_step(f"Excel 읽기 실패:\n{log[:200]}")
            q.put({"type": "done", "status": "error"})
            return

    step("용어집 병합")
    ok, log = run_script([sys.executable, str(SCRIPTS_DIR / "merge_glossary.py")])
    if not ok:
        fail_step(f"용어집 병합 실패:\n{log[:400]}")
        q.put({"type": "done", "status": "error"})
        return
    ok_step("용어집 병합 완료")

    stats = {}
    cache_file = GL_DIR / "glossary_cache.json"
    if cache_file.exists():
        try:
            cache = json.loads(cache_file.read_text(encoding="utf-8"))
            stats = {"term_count": len(cache.get("terms", {}))}
        except Exception:
            pass

    q.put({
        "type": "done",
        "status": "success",
        "stats": stats,
        "summary": (
            f"용어집 로드 완료!\n\n"
            f"• 등록 용어: {stats.get('term_count', 0)}개\n\n"
            "**현지화 번역 → ② 번역** 탭에서 텍스트를 입력하고 번역을 진행하세요."
        ),
    })


# ── PM 상태 확인 ──────────────────────────────────────────────────────────────
@app.route("/api/pm/status")
def pm_status():
    return jsonify(has_ai=HAS_AI, ai_mode=AI_MODE,
                   ollama_model=OLLAMA_MODEL if AI_MODE == "ollama" else None,
                   claude_cli=shutil.which("claude") is not None,
                   active_jobs=len(_jobs))


# ── 메인 페이지 ───────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")


# ═══════════════════════════════════════════════════════════════════════════════
# 이벤트 기획 API (수동 제어)
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/event/tabs", methods=["POST"])
def event_get_tabs():
    """소스 xlsx의 탭 목록 반환 (날짜형 탭 전체)."""
    data   = request.json
    source = data.get("source_path", "").strip()
    if not source:
        return jsonify(ok=False, message="source_path가 필요합니다.")
    try:
        source, _ = resolve_source(source)
    except Exception as e:
        return jsonify(ok=False, message=f"파일 접근 실패: {e}")
    if not Path(source).exists():
        return jsonify(ok=False, message=f"파일 없음: {source}")
    try:
        import openpyxl as _xl, re as _re
        wb = _xl.load_workbook(source, read_only=True)
        all_tabs   = wb.sheetnames[:]
        date_tabs  = sorted([s for s in all_tabs if _re.match(r'^\d{6}$', s)])
        wb.close()
        return jsonify(ok=True, all_tabs=all_tabs, date_tabs=date_tabs)
    except Exception as e:
        return jsonify(ok=False, message=f"xlsx 파싱 오류: {e}")


@app.route("/api/event/create-tabs", methods=["POST"])
def event_create_tabs():
    data = request.json
    source      = data.get("source_path", "")
    new_tabs    = data.get("new_tabs", [])
    ref_tabs    = data.get("ref_tabs", [])
    output_name = data.get("output_filename", f"이벤트기획_{datetime.now().strftime('%Y%m%d')}.xlsx")
    keywords    = data.get("keywords", [])
    genre       = data.get("genre", "")
    market      = data.get("market", "글로벌")

    if not source or not new_tabs:
        return jsonify(ok=False, message="소스 경로와 생성 탭명이 필요합니다.")
    if ref_tabs and len(ref_tabs) != len(new_tabs):
        return jsonify(ok=False, message="생성 탭 수와 참조 탭 수가 다릅니다.")

    # Google Drive URL → 로컬 파일 다운로드
    try:
        source, downloaded = resolve_source(source)
    except Exception as e:
        return jsonify(ok=False, message=f"소스 파일 로드 실패: {e}")
    if not Path(source).exists():
        return jsonify(ok=False, message=f"파일을 찾을 수 없습니다: {source}")

    write_current_project(source_xlsx=source)
    output_path = EP_FILE / output_name
    config = {
        "genre": genre, "target_month": datetime.now().strftime("%Y-%m"),
        "genre_phrases": keywords,
        "event_name_replacements": {tab: [] for tab in new_tabs},
    }
    (EP_WORK / "event_names_config.json").write_text(
        json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (HANDOFF_DIR / "event-planner_input.json").write_text(
        json.dumps({
            "source_path": source, "target_month": datetime.now().strftime("%Y-%m"),
            "market": market, "genre": genre, "new_tab_names": new_tabs,
            "ref_tab_names": ref_tabs, "output_path": str(output_path),
            "genre_phrases": keywords,
        }, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    # generate_event_names.py — 제목 패턴 생성 (탭 생성 전)
    try:
        _dt0 = datetime.strptime("20" + new_tabs[0], "%Y%m%d")
        _target_month = _dt0.strftime("%Y-%m")
    except Exception:
        _target_month = datetime.now().strftime("%Y-%m")
    gen_cmd = [
        sys.executable, str(SCRIPTS_DIR / "generate_event_names.py"),
        "--source",       source,
        "--new-tabs",     ",".join(new_tabs),
        "--ref-tabs",     ",".join(ref_tabs) if ref_tabs else ",".join(new_tabs),
        "--target-month", _target_month,
        "--genre",        genre,
        "--phrases",      ",".join(keywords) if keywords else "",
        "--work-dir",     str(EP_WORK),
    ]
    run_script(gen_cmd)  # 실패해도 날짜 치환은 정상 진행

    cmd = [sys.executable, str(SCRIPTS_DIR / "create_tabs.py"),
           source, str(output_path), ",".join(new_tabs)]
    if ref_tabs:
        cmd.append(",".join(ref_tabs))

    ok, log = run_script(cmd)
    if ok:
        return jsonify(ok=True, message=f"탭 생성 완료: {output_name}",
                       output_path=str(output_path), log=log)
    return jsonify(ok=False, message="탭 생성 실패", log=log)


@app.route("/api/event/scan-rewards", methods=["POST"])
def event_scan_rewards():
    data   = request.json
    mode   = data.get("mode", "source")
    target = data.get("target_path", "")
    if not target or not Path(target).exists():
        return jsonify(ok=False, message=f"파일이 없습니다: {target}")
    out_key = "reward_by_event.json" if mode == "source" else "reward_new_tabs.json"
    cmd = [sys.executable, str(SCRIPTS_DIR / "scan_rewards_by_event.py"),
           target, str(EP_WORK / out_key)]
    ok, log = run_script(cmd)
    return jsonify(ok=ok, message="스캔 완료" if ok else "스캔 실패", log=log)


@app.route("/api/event/recommend-rewards", methods=["POST"])
def event_recommend_rewards():
    if not (EP_WORK / "reward_by_event.json").exists():
        return jsonify(ok=False, message="소스 탭 보상 스캔을 먼저 실행하세요.")
    if not (EP_WORK / "reward_new_tabs.json").exists():
        return jsonify(ok=False, message="신규 탭 보상 스캔을 먼저 실행하세요.")
    ok1, log1 = run_script([sys.executable, str(SCRIPTS_DIR / "recommend_rewards.py"),
                             "--per-event"])
    if not ok1:
        return jsonify(ok=False, message="보상 추천 실패", log=log1)
    ok2, log2 = run_script([sys.executable, str(SCRIPTS_DIR / "_prep_sequential_review.py")])
    if not ok2:
        return jsonify(ok=False, message="리뷰 큐 생성 실패", log=log2)
    stats = {}
    queue_file = EP_WORK / "reward_review_queue.json"
    if queue_file.exists():
        q = json.loads(queue_file.read_text(encoding="utf-8"))
        stats = {"total_sections": q.get("total_sections", 0),
                 "sections_with_changes": q.get("sections_with_changes", 0)}
    return jsonify(ok=True, message="보상 추천 완료", stats=stats, log=log1 + "\n" + log2)


@app.route("/api/event/apply-rewards", methods=["POST"])
def event_apply_rewards():
    data        = request.json
    output_path = data.get("output_path", "")
    changes     = data.get("changes", [])
    if not output_path or not Path(output_path).exists():
        return jsonify(ok=False, message="output xlsx 경로를 확인하세요.")
    if not changes:
        return jsonify(ok=False, message="변경 내용이 없습니다.")
    changes_file = EP_WORK / "reward_changes_manual.json"
    changes_file.write_text(json.dumps(changes, ensure_ascii=False, indent=2), encoding="utf-8")
    cmd = [sys.executable, str(SCRIPTS_DIR / "apply_reward_changes.py"),
           "--xlsx", output_path, "--changes", str(changes_file)]
    ok, log = run_script(cmd)
    return jsonify(ok=ok, message="변경 적용 완료" if ok else "적용 실패", log=log)


@app.route("/api/event/analyze-patterns", methods=["POST"])
def event_analyze_patterns():
    data   = request.json
    source = data.get("source_path", "")
    output = data.get("output_path", "")
    tabs   = data.get("tabs", "")
    if not all([source, output, tabs]):
        return jsonify(ok=False, message="소스 xlsx, output xlsx, 탭명을 모두 입력하세요.")
    for p in [source, output]:
        if not Path(p).exists():
            return jsonify(ok=False, message=f"파일이 없습니다: {p}")
    cmd = [sys.executable, str(SCRIPTS_DIR / "analyze_event_patterns.py"),
           source, output, tabs.replace(" ", "")]
    ok, log = run_script(cmd)
    analysis = {}
    analysis_file = EP_WORK / "event_pattern_analysis.json"
    if analysis_file.exists():
        try:
            analysis = json.loads(analysis_file.read_text(encoding="utf-8"))
        except Exception:
            pass
    return jsonify(ok=ok, message="갭 분석 완료" if ok else "갭 분석 경고 (블로킹 안 함)",
                   analysis=analysis, log=log)


@app.route("/api/event/date-patterns", methods=["POST"])
def event_date_patterns():
    """날짜별 이벤트 패턴 분석 — Google Sheets URL 또는 xlsx 경로 지원."""
    data   = request.json
    source = data.get("source_path", "").strip()
    if not source:
        return jsonify(ok=False, message="source_path 가 필요합니다.")

    # Google Sheets / Drive URL → 로컬 xlsx 다운로드
    try:
        source_local, _ = resolve_source(source)
    except Exception as e:
        return jsonify(ok=False, message=f"파일 로드 실패: {e}")

    if not Path(source_local).exists():
        return jsonify(ok=False, message=f"파일 없음: {source_local}")

    write_current_project(source_xlsx=source_local)

    cmd = [sys.executable, str(SCRIPTS_DIR / "analyze_date_patterns.py"), source_local]
    ok, log = run_script(cmd)

    result = {}
    out_file = EP_WORK / "date_pattern_analysis.json"
    if out_file.exists():
        try:
            result = json.loads(out_file.read_text(encoding="utf-8"))
        except Exception:
            pass

    return jsonify(
        ok=ok,
        message="날짜별 패턴 분석 완료" if ok else "분석 경고 (결과 확인 필요)",
        analysis=result,
        log=log,
    )


@app.route("/api/event/upload-gsheets", methods=["POST"])
def event_upload_gsheets():
    data        = request.json
    output_path = data.get("output_path", "")
    target_url  = data.get("target_url", "")
    if not output_path or not Path(output_path).exists():
        return jsonify(ok=False, message="업로드할 xlsx 파일이 없습니다.")
    cmd = [sys.executable, str(SCRIPTS_DIR / "upload_to_gsheets.py"), output_path]
    if target_url:
        cmd += ["--target-url", target_url]
    ok, log = run_script(cmd)
    return jsonify(ok=ok, message="업로드 완료" if ok else "업로드 실패", log=log)


@app.route("/api/event/files")
def event_files():
    files = []
    if EP_FILE.exists():
        for f in sorted(EP_FILE.glob("*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True):
            files.append({"name": f.name, "size_kb": f.stat().st_size // 1024,
                          "path": str(f)})
    return jsonify(files=files)


@app.route("/api/event/download")
def event_download():
    path = request.args.get("path", "")
    p = Path(path)
    if not p.exists() or not p.is_file():
        return jsonify(ok=False, message="파일이 없습니다."), 404
    return send_file(str(p), as_attachment=True, download_name=p.name)


# ═══════════════════════════════════════════════════════════════════════════════
# 현지화 번역 API (수동 제어)
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/localizer/load-glossary", methods=["POST"])
def localizer_load_glossary():
    data       = request.json
    gsheet_url = data.get("gsheet_url", "")
    excel_path = data.get("excel_path", "")
    if not gsheet_url and not excel_path:
        return jsonify(ok=False, message="Google Sheets URL 또는 Excel 경로가 필요합니다.")

    # Google Drive 파일 링크 → 로컬 xlsx로 다운로드 (Sheets URL은 fetch_gsheet.py가 처리)
    from gdrive_utils import is_drive_url, is_sheets_url
    if excel_path and is_drive_url(excel_path) and not is_sheets_url(excel_path):
        try:
            local, _ = resolve_source(excel_path)
            excel_path = local
        except Exception as e:
            return jsonify(ok=False, message=f"Drive 파일 다운로드 실패: {e}")

    config = {"gsheet_url": gsheet_url, "excel_path": excel_path,
              "column_mapping": {"source": "원어", "target_prefix": ""}}
    (GL_DIR / "glossary_config.json").write_text(
        json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    logs = []
    if gsheet_url:
        ok, log = run_script([sys.executable, str(SCRIPTS_DIR / "fetch_gsheet.py"), gsheet_url])
        logs.append(f"[GSheets] {'OK' if ok else 'FAIL'}: {log[:300]}")
        if not ok:
            logs.append("→ Excel 단독 모드로 계속")
    if excel_path and Path(excel_path).exists():
        ok, log = run_script([sys.executable, str(SCRIPTS_DIR / "read_excel.py"), excel_path])
        logs.append(f"[Excel] {'OK' if ok else 'FAIL'}: {log[:300]}")
    ok, log = run_script([sys.executable, str(SCRIPTS_DIR / "merge_glossary.py")])
    logs.append(f"[Merge] {'OK' if ok else 'FAIL'}: {log[:300]}")
    if not ok:
        return jsonify(ok=False, message="병합 실패", log="\n".join(logs))
    stats = {}
    cache_file = GL_DIR / "glossary_cache.json"
    if cache_file.exists():
        cache = json.loads(cache_file.read_text(encoding="utf-8"))
        terms     = cache.get("terms", {})
        conflicts = cache.get("conflicts_log", [])
        stats = {"term_count": len(terms), "conflict_count": len(conflicts),
                 "generated_at": cache.get("generated_at", "")[:19]}
    return jsonify(ok=True, message="용어집 로드 완료", stats=stats, log="\n".join(logs))


@app.route("/api/localizer/glossary-preview")
def localizer_glossary_preview():
    cache_file = GL_DIR / "glossary_cache.json"
    if not cache_file.exists():
        return jsonify(ok=False, terms={}, message="용어집 캐시 없음")
    cache = json.loads(cache_file.read_text(encoding="utf-8"))
    terms = cache.get("terms", {})
    preview = {k: v for k, v in list(terms.items())[:30]}
    return jsonify(ok=True, terms=preview, total=len(terms),
                   generated_at=cache.get("generated_at", "")[:19])


@app.route("/api/localizer/match-terms", methods=["POST"])
def localizer_match_terms():
    data         = request.json
    source_text  = data.get("source_text", "")
    text_type    = data.get("text_type")
    target_langs = data.get("target_languages", ["ko", "ja", "en", "zh"])
    if not source_text:
        return jsonify(ok=False, message="번역할 텍스트가 없습니다.")
    cache_file = GL_DIR / "glossary_cache.json"
    if not cache_file.exists():
        return jsonify(ok=False, message="용어집을 먼저 로드하세요.")
    req = {"source_text": source_text, "text_type_hint": text_type,
           "target_languages": target_langs,
           "glossary_cache_path": str(cache_file)}
    req_file = GL_DIR / "translate_request.json"
    req_file.write_text(json.dumps(req, ensure_ascii=False, indent=2), encoding="utf-8")
    ok, log = run_script([sys.executable, str(SCRIPTS_DIR / "match_terms.py")])
    if not ok:
        return jsonify(ok=False, message="용어집 매칭 실패", log=log)
    req_data = json.loads(req_file.read_text(encoding="utf-8"))
    return jsonify(ok=True, matched_terms=req_data.get("matched_terms", {}),
                   unregistered_terms=req_data.get("unregistered_terms", []), log=log)


@app.route("/api/localizer/save-translation", methods=["POST"])
def localizer_save_translation():
    data         = request.json
    translations = data.get("translations", {})
    text_type    = data.get("text_type", "dialogue")
    unregistered = data.get("unregistered_terms", [])
    if not translations:
        return jsonify(ok=False, message="번역 결과가 없습니다.")
    result_data = {"text_type": text_type, "translations": translations,
                   "unregistered_terms": unregistered,
                   "validation_status": "pass", "retry_count": 0}
    (GL_DIR / "translate_result.json").write_text(
        json.dumps(result_data, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    ok, log = run_script([sys.executable, str(SCRIPTS_DIR / "buffer_manager.py"), "add"])
    return jsonify(ok=ok, message="버퍼에 추가 완료" if ok else "버퍼 저장 실패", log=log)


@app.route("/api/localizer/buffer")
def localizer_buffer():
    buffer_file = GL_DIR / "session_buffer.json"
    if not buffer_file.exists():
        return jsonify(ok=True, entries=[], total=0)
    buf = json.loads(buffer_file.read_text(encoding="utf-8"))
    entries = buf.get("entries", [])
    return jsonify(ok=True, entries=entries[-20:], total=len(entries))


@app.route("/api/localizer/buffer/clear", methods=["POST"])
def localizer_buffer_clear():
    ok, log = run_script([sys.executable, str(SCRIPTS_DIR / "buffer_manager.py"), "clear"])
    return jsonify(ok=ok, message="버퍼 초기화 완료" if ok else "초기화 실패", log=log)


@app.route("/api/localizer/export", methods=["POST"])
def localizer_export():
    data = request.json
    fmt  = data.get("format", "xlsx")
    buffer_file = GL_DIR / "session_buffer.json"
    if not buffer_file.exists():
        return jsonify(ok=False, message="세션 버퍼가 비어 있습니다.")
    buf = json.loads(buffer_file.read_text(encoding="utf-8"))
    if not buf.get("entries"):
        return jsonify(ok=False, message="버퍼에 번역 항목이 없습니다.")
    ok, log = run_script([sys.executable, str(SCRIPTS_DIR / "export_xlsx.py"),
                           "--format", fmt])
    if not ok:
        return jsonify(ok=False, message="Export 실패", log=log)
    export_files = sorted(GL_DIR.glob("translation_export_*.*"),
                          key=lambda p: p.stat().st_mtime, reverse=True)
    files = [{"name": f.name, "path": str(f)} for f in export_files[:3]]
    return jsonify(ok=True, message="Export 완료", files=files, log=log)


@app.route("/api/localizer/files")
def localizer_files():
    files = []
    if GL_DIR.exists():
        for f in sorted(GL_DIR.glob("translation_export_*.*"),
                        key=lambda x: x.stat().st_mtime, reverse=True):
            files.append({"name": f.name, "size_kb": f.stat().st_size // 1024,
                          "path": str(f)})
    return jsonify(files=files)


@app.route("/api/localizer/download")
def localizer_download():
    path = request.args.get("path", "")
    p = Path(path)
    if not p.exists() or not p.is_file():
        return jsonify(ok=False, message="파일이 없습니다."), 404
    return send_file(str(p), as_attachment=True, download_name=p.name)


# ── 프로젝트 관리 API ──────────────────────────────────────────────────────────

@app.route("/api/projects")
def get_projects():
    """프로젝트 목록 및 설정 반환."""
    return jsonify(projects=_load_projects())


@app.route("/api/project/config", methods=["POST"])
def update_project_config():
    """프로젝트 설정 업데이트 (장르, 마켓 등)."""
    data       = request.json or {}
    project_id = data.get("project_id", "")
    projects   = _load_projects()

    if project_id not in projects:
        return jsonify(ok=False, message=f"알 수 없는 프로젝트: {project_id}")

    # 허용 필드만 업데이트
    for field in ("genre", "market", "display_name", "description"):
        if field in data:
            projects[project_id][field] = data[field]

    # 장르·마켓 모두 설정됐으면 configured=True
    p = projects[project_id]
    if p.get("genre") and p.get("market"):
        p["configured"] = True

    _save_projects(projects)
    return jsonify(ok=True, message="프로젝트 설정 저장 완료", project=p)


# ── 실행 ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    PORT = 5050

    def open_browser():
        time.sleep(1.2)
        webbrowser.open(f"http://localhost:{PORT}")

    threading.Thread(target=open_browser, daemon=True).start()
    print(f"\n멀티 에이전트 PM 서버 시작: http://localhost:{PORT}\n")
    app.run(host="0.0.0.0", port=PORT, debug=False)
